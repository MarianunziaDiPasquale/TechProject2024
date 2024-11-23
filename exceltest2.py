import sqlite3
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, Protection
from openpyxl.drawing.image import Image
from datetime import datetime


def crea_file_excel_con_estetica_da_db(table_name, sconto):
    """
    Crea un file Excel utilizzando i codici presi dalla tabella specificata in liste_personalizzate.db
    e confronta i dati con la tabella prodotti in MergedDatabase.db, applicando uno sconto uniforme.
    Aggiunge anche un titolo con il nome del file e uno spazio per un'immagine.
    """
    # Nome del file con la data di oggi
    data_oggi = datetime.now().strftime("%Y-%m-%d")
    output_file = f"output_{data_oggi}.xlsx"
    merged_db_path = 'Database_Utilities/database/MergedDatabase.db'
    liste_personalizzate_db_path = 'Database_Utilities/database/liste_personalizzate.db'
    logo_path = 'resources/LogoCINCOTTI.jpg'

    # Connessione ai database SQLite
    conn_liste = sqlite3.connect(liste_personalizzate_db_path)
    conn_merged = sqlite3.connect(merged_db_path)
    cursor_liste = conn_liste.cursor()
    cursor_merged = conn_merged.cursor()

    # Recupera i codici dalla tabella specificata
    query_codici = f"SELECT Col1 FROM '{table_name}'"
    cursor_liste.execute(query_codici)
    lista_codici = [row[0] for row in cursor_liste.fetchall()]

    # Crea una nuova cartella di lavoro (workbook) per Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Ordine"

    # Imposta il titolo prima di unire le celle
    titolo_cell = sheet.cell(row=1, column=1)
    titolo_cell.value = f"BARONE-AUSONIA_{data_oggi}.xlsx"
    titolo_cell.font = Font(bold=True, size=14)
    titolo_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Unisci le celle per il titolo
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    # Aggiungi l'immagine, dimensionata per occupare 8 colonne e 6 righe
    logo = Image(logo_path)
    logo.width = 535  # Imposta larghezza a 800 pixel
    logo.height = 135  # Imposta altezza a 150 pixel per adattarsi a 6 righe
    sheet.add_image(logo, "A2")
    sheet.merge_cells(start_row=2, start_column=1, end_row=8, end_column=7)

    # Blocca la riga del titolo in modo che non sia modificabile
    for cell in sheet["1:1"]:
        cell.protection = Protection(locked=True)

    # Stile per le celle
    bordo_sottile = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))
    riempimento_celle_giallo = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    riempimento_verde = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    riempimento_bianco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    riempimento_rosso = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    font_stile = Font(name="Calibri", size=12, bold=False)
    font_rosso = Font(name="Calibri", size=12, bold=False, color="FF0000")
    font_blu = Font(name="Calibri", size=12, bold=True, color="0066CC", underline="single")
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Intestazioni delle colonne, posizionate sotto l'immagine (riga 8)
    intestazioni = ["Codice", "Descrizione", "Composizione Cartone", "Prezzo Unitario", "Sconto (%)",
                    "Prezzo Scontato (€)", "Quantità"]
    for col_num, header in enumerate(intestazioni, start=1):
        cell = sheet.cell(row=9, column=col_num)
        cell.value = header
        cell.fill = riempimento_celle_giallo
        cell.border = bordo_sottile
        cell.alignment = alignment_center
        cell.font = font_rosso if header in ["Sconto (%)", "Prezzo Scontato (€)"] else Font(bold=True)

    # Recupera e aggiungi i dati dal database MergedDatabase
    query_prodotti = "SELECT Descrizione, COMPOSIZIONE_CARTONE, PREZZO_VENDITA FROM prodotti WHERE Codice = ?"

    # Imposta larghezza colonne per evitare che il testo venga tagliato
    larghezze_colonne = [15, 40, 20, 15, 12, 18, 10, 15]  # Larghezze personalizzate per ciascuna colonna
    for i, col_width in enumerate(larghezze_colonne, start=1):
        col_lettera = openpyxl.utils.get_column_letter(i)
        sheet.column_dimensions[col_lettera].width = col_width


    for idx, codice in enumerate(lista_codici):
        cursor_merged.execute(query_prodotti, (codice,))
        risultato_db = cursor_merged.fetchone()

        if risultato_db:
            descrizione, composizione_cartone, prezzo_unitario = risultato_db

            # Aggiungi la riga con i dati
            nuova_riga = [codice, descrizione, composizione_cartone, prezzo_unitario, sconto, "", "", ""]
            sheet.append(nuova_riga)

            # Colore alternato della bandiera italiana
            riempimento_riga = riempimento_verde if idx % 3 == 0 else (
                riempimento_bianco if idx % 3 == 1 else riempimento_rosso)

            # Applica la formattazione alle celle
            for col in ['A', 'B', 'C', 'D', 'E', 'F', "G"]:
                cell = sheet[f'{col}{sheet.max_row}']
                cell.border = bordo_sottile
                cell.fill = riempimento_riga
                cell.font = font_stile
                cell.alignment = alignment_center
                if col == 'D':
                    cell.font = font_rosso
                if col == 'F':
                    cell.font = font_blu

            # Colonna F (Prezzo Scontato) calcolata con lo sconto applicato al prezzo unitario
            sheet[f'F{sheet.max_row}'] = f'=D{sheet.max_row}*(1 - E{sheet.max_row}/100)'



        else:
            print(f"Codice {codice} non trovato nel database MergedDatabase.")

    # Protezione delle colonne
    for row in sheet.iter_rows(min_row=9, max_row=sheet.max_row):
        for cell in row:
            cell.protection = Protection(locked=cell.column != 7)

    # Protezione del foglio
    sheet.protection.set_password('password')
    workbook.save(output_file)

    # Chiude le connessioni ai database
    conn_liste.close()
    conn_merged.close()
    print(f"File Excel salvato come {output_file}")


# Parametri di esempio
table_name = "112IN"  # Nome della tabella nel database liste_personalizzate
sconto = 10  # Sconto uniforme per tutte le righe
crea_file_excel_con_estetica_da_db(table_name, sconto)
