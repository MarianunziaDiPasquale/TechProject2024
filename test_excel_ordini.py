import sqlite3
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, Protection
from openpyxl.drawing.image import Image
from datetime import datetime

def crea_file_excel_con_estetica(lista_codici, sconti):
    data_oggi = datetime.now().strftime("%Y-%m-%d")
    output_file = f"output_{data_oggi}.xlsx"
    merged_db_path = '/mnt/data/MergedDatabase.db'
    liste_personalizzate_db_path = '/mnt/data/liste_personalizzate.db'
    logo_path = 'resources/LogoCINCOTTI.jpg'



    output_file = "file_excel/output.xlsx"
    db_path = "Database_Utilities/Database/MergedDatabase.db"
    query = f"SELECT Descrizione, COMPOSIZIONE_CARTONE, PREZZO_VENDITA FROM prodotti WHERE Codice = ?"


    # Connessione al database SQLite
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Crea una nuova cartella di lavoro (workbook) per Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Ordine"

    # Definisci i bordi e altri stili
    bordo_sottile = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

    # Definisci il riempimento e il font
    riempimento_celle_giallo = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    riempimento_verde = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    riempimento_bianco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    riempimento_rosso = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    font_stile = Font(name="Calibri", size=12, bold=False)
    font_rosso = Font(name="Calibri", size=12, bold=False, color="FF0000")
    font_blu = Font(name="Calibri", size=12, bold=True, color="0066CC",
                    underline="single")  # Blu Chiaro per la colonna F
    font_verde_bold_sottolineato = Font(name="Calibri", size=12, bold=True, color="00FF00", underline="single")
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Imposta l'intestazione delle colonne (con la nuova colonna "Composizione Cartone" e "Prezzo Totale")
    intestazioni = ["Codice", "Descrizione", "Composizione Cartone", "Prezzo Unitario", "Sconto (%)", "Prezzo Scontato (€)",
                    "Quantità"]
    sheet.append(intestazioni)

    # Imposta larghezza colonne per evitare che il testo venga tagliato
    larghezze_colonne = [15, 40, 20, 15, 12, 18, 10, 15]  # Larghezze personalizzate per ciascuna colonna
    for i, col_width in enumerate(larghezze_colonne, start=1):
        col_lettera = openpyxl.utils.get_column_letter(i)
        sheet.column_dimensions[col_lettera].width = col_width

    # Applica lo sfondo giallo alla prima riga (intestazioni)
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        cell = sheet[f'{col}1']
        cell.fill = riempimento_celle_giallo
        cell.border = bordo_sottile
        cell.alignment = alignment_center

        # Imposta il font rosso per "Sconto (%)" e "Prezzo Scontato", altrimenti nero
        if col in ['E', 'F']:  # Colonne "Sconto (%)" e "Prezzo Scontato"
            cell.font = font_rosso
        else:
            cell.font = Font(bold=True)  # Font nero e in grassetto per altre intestazioni

    # Aggiungi dati per ogni codice nella lista
    for idx, codice in enumerate(lista_codici):
        # Recupera i dati dal database
        query = f"SELECT Descrizione, COMPOSIZIONE_CARTONE, PREZZO_VENDITA FROM prodotti WHERE Codice = ?"
        cursor.execute(query, (codice,))
        risultato_db = cursor.fetchone()

        if risultato_db:
            descrizione, composizione_cartone, prezzo_unitario = risultato_db
            sconto = sconti[idx]
            prezzo_scontato = prezzo_unitario * (1 - sconto / 100)

            # Aggiungi una riga al foglio Excel con i dati (includi "Composizione Cartone")
            nuova_riga = [codice, descrizione, composizione_cartone, prezzo_unitario, sconto, prezzo_scontato, "",
                          ""]  # Quantità e Prezzo Totale vuoti
            sheet.append(nuova_riga)

            # Alterna i colori della bandiera italiana (verde, bianco, rosso) per ogni riga
            if idx % 3 == 0:
                riempimento_riga = riempimento_verde
            elif idx % 3 == 1:
                riempimento_riga = riempimento_bianco
            else:
                riempimento_riga = riempimento_rosso

            # Applica la formattazione estetica solo a queste nuove righe, esclusa la colonna "Quantità"
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:  # Esclusa la colonna "Quantità"
                cell = sheet[f'{col}{sheet.max_row}']
                cell.border = bordo_sottile
                cell.fill = riempimento_riga
                cell.font = font_stile
                cell.alignment = alignment_center

                # Font rosso per la colonna D (Sconto)
                if col == 'D':
                    cell.font = font_rosso

                # Font blu chiaro per la colonna F (Prezzo Scontato)
                if col == 'F':
                    cell.font = font_blu



            # Colonna "Quantità" ha i bordi ma senza riempimento
            cell_quantita = sheet[f'G{sheet.max_row}']
            cell_quantita.border = bordo_sottile

            # Inserisci la formula per calcolare il Prezzo Totale (colonna H)


        else:
            print(f"Codice {codice} non trovato nel database.")

    # Protezione delle colonne: blocca tutte le colonne tranne la colonna "Quantità"
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            if cell.column != 7:  # La colonna "Quantità" è la settima
                cell.protection = Protection(locked=True)
            else:
                cell.protection = Protection(locked=False)

    # Protezione del foglio
    sheet.protection.set_password('password')

    # Salva il file Excel
    workbook.save(output_file)
    print(f"File Excel salvato come {output_file}")

    # Chiude la connessione al database
    conn.close()


lista_codici = ["C111", "BRESAOLA", "AFF2", "C105", "SAM", "SPRITE"]
sconti = [10, 15, 20, 10, 10, 10]
crea_file_excel_con_estetica(lista_codici, sconti)