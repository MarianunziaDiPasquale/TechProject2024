import os
from tkinter.filedialog import asksaveasfilename
import customtkinter as ctk
from tkinter import ttk, Menu
import tkinter as tk

from tkcalendar import Calendar

from Database_Utilities.crud_clienti import get_all_clienti_names, get_cliente_info_by_name
import openpyxl
from tkinter import messagebox

import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, Protection
from popup_functions import open_add_popup
from Database_Utilities.crud_fornitori import get_all_prodotti
from Database_Utilities.connection import _connection

def get_client_id_by_ragione_sociale(ragione_sociale):
    # Connect to the database
    db_path = 'Database_Utilities/Database/Magazzino.db'
    connection = sqlite3.connect(db_path)
    cursor = connection.cursor()

    # Query to fetch the ID based on the "Ragione sociale"
    query = "SELECT ID FROM clienti WHERE `Ragione sociale` = ?"
    cursor.execute(query, (ragione_sociale,))
    result = cursor.fetchone()

    # Close the connection
    connection.close()

    # Return the ID if found, else return a message indicating no match
    return result[0] if result else "No client found with the given Ragione sociale."
dashboard_font_size = 14

def pick_date(event,popup, entry):  #
    global cal, date_window
    date_window = tk.Toplevel()
    date_window.grab_set()
    date_window.title("Scegli una data")
    date_window.geometry(
        f"260x230+{popup.winfo_x() + entry.winfo_x() + 128 * 3}+{popup.winfo_y() + entry.winfo_y() + entry.winfo_height() - 70}")
    cal = Calendar(date_window, selectmode="day", date_pattern="dd/mm/yy")
    cal.place(x=0, y=0)
    submit_button = tk.Button(date_window, text="Submit", command=lambda: grab_date(entry))
    submit_button.place(x=80, y=200)


def grab_date(entry):
    entry.delete(0, 'end')
    entry.insert(0, cal.get_date())
    date_window.destroy()
# Add this function to dashboard_clienti.py or an appropriate module
def get_prodotti_by_cliente(cliente_id):
    conn = _connection()
    cursor = conn.cursor()

    # Querying a table named after the cliente_id
    query = f"SELECT `col1` FROM `{cliente_id}`"
    cursor.execute(query)
    prodotti = cursor.fetchall()
    conn.close()
    lista = []
    for x in prodotti:
        lista.append(x[0])
    return lista  # This returns a list of col1


def get_product_descriptions(prodotti):
    risultato = {}
    conn = _connection()
    cursor = conn.cursor()
    query = "SELECT `Codice`, `Descrizione` FROM `prodotti` WHERE `Codice` = %s"


    for codice in prodotti:
        cursor.execute(query, (str(codice),))
        risultato_db = cursor.fetchone()

        if risultato_db:
            # Formatta ogni prodotto come "Codice - Descrizione"
            risultato[codice] = f"{risultato_db[0]} - {risultato_db[1]}"
        else:
            risultato[codice] = None

    conn.close()
    return risultato


def center_window(window, width, height):
    window.update_idletasks()
    width = width
    height = height
    x = (window.winfo_screenwidth() // 2) - (width // 2)
    y = (window.winfo_screenheight() // 2) - (height // 2)
    window.geometry(f'{width}x{height}+{x}+{y}')
def export_clienti_to_excel(selected_clienti):
    if not selected_clienti:
        messagebox.showwarning("Nessun Dato", "Non ci sono clienti selezionati da esportare.")
        return

    if len(selected_clienti) == 1:
        default_filename = f"Info_Cliente_{selected_clienti[0]}.xlsx"
    else:
        default_filename = f"Info_Clienti.xlsx"
    file_path = asksaveasfilename(defaultextension=".xlsx",
                                  filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                  initialfile=default_filename)
    # Verifica se l'utente ha scelto un percorso
    if not file_path:
        messagebox.showwarning("Salvataggio Annullato", "Salvataggio del file Excel annullato.")
        return

    workbook = openpyxl.Workbook()

    for cliente in selected_clienti:
        info = get_cliente_info_by_name(cliente)

        # Crea un nuovo foglio per ogni cliente
        sheet = workbook.create_sheet(title=cliente)

        headers = ["Ragione sociale", "Indirizzo", "CAP", "Città", "Nazione", "Partita IVA", "Telefono", "Email", "Zona", "Giorni chiusura", "Orari di scarico", "Condizioni pagamento", "Sconto", "Agente"]
        sheet.append(headers)

        # Aggiungi i dati del cliente se esistono, altrimenti lascia solo le intestazioni
        if info:
            sheet.append([info.get(header, "N/A") for header in headers])
        else:
            # Aggiungi una riga vuota se non ci sono dati, lasciando solo l'intestazione
            sheet.append(["Nessun dato disponibile"])

    # Rimuove il foglio predefinito solo se ci sono altri fogli visibili
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook["Sheet"]

    workbook.save(file_path)
    messagebox.showinfo("Excel Generato", f"Il file Excel è stato salvato come {os.path.basename(file_path)}")

def export_liste_prsn_to_excel(selected_lista):
    if not selected_lista:
        messagebox.showwarning("Nessun Dato", "Non ci sono clienti selezionati da esportare.")
        return

    if len(selected_lista.get()) == 1:
        default_filename = f"Info_Lista_Cliente_{selected_lista[0]}.xlsx"
    else:
        default_filename = f"Info_Lista_Clienti.xlsx"
    file_path = asksaveasfilename(defaultextension=".xlsx",
                                  filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                  initialfile=default_filename)
    # Verifica se l'utente ha scelto un percorso
    if not file_path:
        messagebox.showwarning("Salvataggio Annullato", "Salvataggio del file Excel annullato.")
        return

    workbook = openpyxl.Workbook()

    cliente = get_client_id_by_ragione_sociale(selected_lista.get())
    # modificare con funzione per recuperare info lista
    info = get_lista_info(cliente)

    # Crea un nuovo foglio per ogni cliente
    sheet = workbook.create_sheet(title=cliente)

    # modificare con colonne lista personalizzata
    headers = ["Ragione sociale", "ID", "Prodotto", "Quantità"]
    sheet.append(headers)

    # Aggiungi i dati del cliente se esistono, altrimenti lascia solo le intestazioni
    if info:
        for row in info:
            sheet.append(row)
    else:
        # Aggiungi una riga vuota se non ci sono dati, lasciando solo l'intestazione
        sheet.append(["Nessun dato disponibile"])




    # Rimuove il foglio predefinito solo se ci sono altri fogli visibili
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook["Sheet"]

    workbook.save(file_path)
    messagebox.showinfo("Excel Generato", f"Il file Excel è stato salvato come {os.path.basename(file_path)}")

def crea_file_excel_con_estetica(lista_codici, sconti, nome_tabella, db_path="database.db", output_file="output.xlsx"):
    conn = _connection()
    cursor = conn.cursor()

    # Crea una nuova cartella di lavoro (workbook) per Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Ordine"

    # Definisci i bordi e altri stili per l'esempio
    bordo_sottile = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

    # Definisci il riempimento e il font
    riempimento_celle_giallo = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    font_stile = Font(name="Calibri", size=12, bold=False)
    font_rosso = Font(name="Calibri", size=12, bold=False, color="FF0000")
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Imposta l'intestazione delle colonne
    intestazioni = ["Codice", "Descrizione", "Prezzo Unitario", "Sconto (%)", "Prezzo Scontato", "Quantità"]
    sheet.append(intestazioni)

    # Applica lo sfondo giallo alla prima riga (intestazioni)
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = sheet[f'{col}1']
        cell.fill = riempimento_celle_giallo
        cell.border = bordo_sottile
        cell.alignment = alignment_center

        # Imposta il font rosso per "Sconto (%)" e "Prezzo Scontato", altrimenti nero
        if col in ['D', 'E']:  # Colonne "Sconto (%)" e "Prezzo Scontato"
            cell.font = font_rosso
        else:
            cell.font = Font(bold=True)  # Font nero e in grassetto per altre intestazioni

    # Aggiungi dati per ogni codice nella lista
    for idx, codice in enumerate(lista_codici):
        # Recupera i dati dal database
        query = f"SELECT `descrizione`, `prezzo` FROM `{nome_tabella}` WHERE `codice` = %s"
        cursor.execute(query, (codice,))
        risultato_db = cursor.fetchone()

        if risultato_db:
            descrizione, prezzo_unitario = risultato_db
            sconto = sconti[idx]
            prezzo_scontato = prezzo_unitario * (1 - sconto / 100)

            # Aggiungi una riga al foglio Excel con i dati
            nuova_riga = [codice, descrizione, prezzo_unitario, sconto, prezzo_scontato, ""]  # Quantità vuota
            sheet.append(nuova_riga)

            # Applica la formattazione estetica solo a queste nuove righe
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                cell = sheet[f'{col}{sheet.max_row}']
                cell.border = bordo_sottile
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Bianco per le righe dei dati
                cell.font = font_stile
                cell.alignment = alignment_center

        else:
            print(f"Codice {codice} non trovato nel database.")

    # Protezione delle colonne: blocca tutte le colonne tranne la colonna "Quantità"
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            if cell.column != 6:  # La colonna "Quantità" è la sesta
                cell.protection = Protection(locked=True)
            else:
                cell.protection = Protection(locked=False)

    # Protezione del foglio
    sheet.protection.set_password('password')

    # Salva il file Excel
    workbook.save(output_file)
    print(f"File Excel salvato come {output_file}")

    # Chiude la connessione al database
    conn.close()

def open_lista_selection_popup():
    popup = tk.Toplevel()
    popup.title("Seleziona la lista ")
    popup.geometry("700x700")  # Dimensioni compatte

    instruction_label = ctk.CTkLabel(popup, text="Seleziona il cliente di cui vuoi esportare la lista personalizzata :", font=('Arial', dashboard_font_size), text_color="black")
    instruction_label.pack(pady=10)

    container_frame = tk.Frame(popup)
    container_frame.pack(fill="both", expand=True, padx=60)

    clienti = get_all_clienti_names()  # Da modificare
    selected_cliente = tk.StringVar()
    combobox = ttk.Combobox(container_frame, textvariable=selected_cliente, values=clienti,
                            font=('Arial', dashboard_font_size))
    combobox.configure(width=35)
    combobox.option_add('*TCombobox*Listbox*Font', ('Arial', dashboard_font_size))
    combobox.pack(pady=5)
    '''

    canvas = tk.Canvas(container_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(container_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.create_window((0, 0), window=scrollable_frame, anchor="n")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Variabile per tenere traccia dei clienti selezionati
    selected_clienti = []

    def on_checkbutton_toggle(cliente, var):
        if var.get():
            if cliente not in selected_clienti:
                selected_clienti.append(cliente)
        else:
            if cliente in selected_clienti:
                selected_clienti.remove(cliente)

    def toggle_select_all():
        all_selected = all(var.get() for var in selections.values())
        if all_selected:
            for var in selections.values():
                var.set(False)
            selected_clienti.clear()
        else:
            selected_clienti.clear()
            for cliente in clienti:
                var = selections[cliente]
                var.set(True)
                selected_clienti.append(cliente)

    # Aggiungi il pulsante Seleziona/Deseleziona tutto
    select_all_button = ctk.CTkButton(popup, text="Seleziona Tutto", command=toggle_select_all, width=120, height=30)
    select_all_button.pack(pady=5)

    # Dizionario per tracciare le variabili di stato dei checkbutton
    selections = {}

    clienti = get_all_clienti_names()
    for cliente in clienti:
        var = tk.BooleanVar()
        selections[cliente] = var
        check = tk.Checkbutton(scrollable_frame, text=cliente, variable=var,font=('Arial', 14),
                               command=lambda c=cliente, v=var: on_checkbutton_toggle(c, v))
        check.pack(anchor="w")
    '''

    entry_width = 20
    sconto_var = tk.StringVar()
    sconto_label = tk.Label(container_frame, text="Sconto:", font=("Arial", dashboard_font_size))
    sconto_label.pack(pady=5)
    sconto_entry = tk.Entry(container_frame, width=entry_width, textvariable=sconto_var, font=("Arial", dashboard_font_size))
    sconto_entry.pack(pady=5)


    # Pulsante di conferma
    confirm_button = ctk.CTkButton(popup, text="Conferma", command=lambda: [popup.destroy(), export_liste_prsn_to_excel(selected_cliente)], width=120, height=30, font = ('Arial', dashboard_font_size))
    confirm_button.pack(pady=10)

    popup.mainloop()
def open_clienti_selection_popup():
    popup = tk.Toplevel()
    popup.title("Seleziona Clienti da Esportare")
    popup.geometry("700x700")  # Dimensioni compatte

    instruction_label = ctk.CTkLabel(popup, text="Seleziona i clienti che vuoi esportare:", font=('Arial', dashboard_font_size), text_color="black")
    instruction_label.pack(pady=10)

    container_frame = tk.Frame(popup)
    container_frame.pack(fill="both", expand=True, padx=60)

    canvas = tk.Canvas(container_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(container_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.create_window((0, 0), window=scrollable_frame, anchor="n")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Variabile per tenere traccia dei clienti selezionati
    selected_clienti = []

    def on_checkbutton_toggle(cliente, var):
        if var.get():
            if cliente not in selected_clienti:
                selected_clienti.append(cliente)
        else:
            if cliente in selected_clienti:
                selected_clienti.remove(cliente)

    def toggle_select_all():
        all_selected = all(var.get() for var in selections.values())
        if all_selected:
            for var in selections.values():
                var.set(False)
            selected_clienti.clear()
        else:
            selected_clienti.clear()
            for cliente in clienti:
                var = selections[cliente]
                var.set(True)
                selected_clienti.append(cliente)

    # Aggiungi il pulsante Seleziona/Deseleziona tutto
    select_all_button = ctk.CTkButton(popup, text="Seleziona Tutto", command=toggle_select_all, width=120, height=30)
    select_all_button.pack(pady=5)

    # Dizionario per tracciare le variabili di stato dei checkbutton
    selections = {}

    clienti = get_all_clienti_names()
    for cliente in clienti:
        var = tk.BooleanVar()
        selections[cliente] = var
        check = tk.Checkbutton(scrollable_frame, text=cliente, variable=var,font=('Arial', dashboard_font_size),
                               command=lambda c=cliente, v=var: on_checkbutton_toggle(c, v))
        check.pack(anchor="w")

    # Pulsante di conferma
    confirm_button = ctk.CTkButton(popup, text="Conferma", command=lambda: [popup.destroy(), export_clienti_to_excel(selected_clienti)], width=120, height=30)
    confirm_button.pack(pady=10)

    popup.mainloop()
def get_cliente_info(cliente):
    info = get_cliente_info_by_name(cliente)
    return info

def get_lista_info(cliente):
    info = get_prodotti_by_cliente(cliente)
    return info

def show_cliente_info(cliente, tree, table_frame):
    print(cliente + "cliente")
    for row in tree.get_children():
        tree.delete(row)
    info = get_cliente_info(cliente)
    if info:
        for key, value in info.items():
            print(value)
            tree.insert("", tk.END, values=[f"{key}: {value}"])
    tree.cliente = cliente # Save the current client in the tree
    table_frame.pack(pady=10, fill="both", expand=True)


def show_lista_info(cliente, tree_1, table_frame_1):
    for row in tree_1.get_children():
        tree_1.delete(row)

    # Get the client ID using the selected client name
    cliente_info = get_cliente_info_by_name(cliente)
    if cliente_info:
        cliente_id = cliente_info.get("ID")  # Assuming the ID is stored under the "ID" column

        # Get the list of products from the personalized table (table named after cliente_id)
        lista_prodotti = get_prodotti_by_cliente(cliente_id)
        print(lista_prodotti)


        if lista_prodotti:
            descrizioni_prodotti = get_product_descriptions(lista_prodotti)

            # Display the products in the Treeview
            for prodotto in lista_prodotti:
                prodotto_id = prodotto
                descrizione = descrizioni_prodotti[prodotto]
                tree_1.insert("", tk.END, values=(prodotto_id, descrizione))
        else:
            tree_1.insert("", tk.END, values=("Nessun prodotto disponibile", ""))

    table_frame_1.pack(pady=10, fill="both", expand=True)


def update_combobox():
    value = combobox.get().lower()
    if value == '':
        combobox['values'] = clienti
    else:
        data = [item for item in clienti if value in item.lower()]
        combobox['values'] = data
        combobox.event_generate('<Down>')


def update_combobox_prs(event, combobox_prodotto):
    prodotti_aggiornati = get_all_prodotti()
    typed_text = combobox_prodotto.get().lower()

    # Filtra i prodotti sia per codice che per descrizione
    prodotti_filtrati = [prod for prod in prodotti_aggiornati if typed_text in prod.lower()]

    cursor_position = combobox_prodotto.index(tk.INSERT)
    combobox_prodotto['values'] = prodotti_filtrati
    combobox_prodotto.set(typed_text)
    combobox_prodotto.icursor(cursor_position)
    combobox_prodotto.event_generate('<Down>')


def on_keyrelease(event):
    combobox.after(2000, update_combobox)

def copy_selection(treeview):
    selected_items = treeview.selection()  # Ottiene tutti gli elementi selezionati
    copied_data = []
    for item in selected_items:
        item_values = treeview.item(item, "values")
        copied_data.append("\t".join(map(str, item_values)))  # Unisce i valori di ciascuna riga con una tabulazione
    clipboard_text = "\n".join(copied_data)  # Unisce tutte le righe con un ritorno a capo
    treeview.clipboard_clear()  # Pulisce gli appunti
    treeview.clipboard_append(clipboard_text)  # Aggiunge il testo agli appunti
    messagebox.showinfo("Copiato", "I testi selezionati sono stati copiati negli appunti.")


def setup_context_menu(treeview):
    # Creare un menu contestuale
    context_menu = Menu(treeview, tearoff=0)
    context_menu.add_command(label="Copia", command=lambda: copy_selection(treeview), font=('Arial', dashboard_font_size))

    def on_right_click(event):
        # Mostrare il menu contestuale
        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    treeview.bind("<Button-3>", on_right_click)  # <Button-3> è il clic del pulsante destro del mouse

# Aggiungere il supporto per la selezione trascinando
def on_mouse_drag(event, treeview):
    # Identifica l'item su cui si trova il cursore
    item = treeview.identify_row(event.y)
    if item:
        treeview.selection_add(item)

def on_double_click_1(event, tree_1):
    item = tree_1.selection()[0]
    values = tree_1.item(item, "values")
    if values:
        current_id = values[0]
        current_prodotto = values[1]
        show_action_dialog_1(current_prodotto, lambda action: handle_action_1(action, current_id,current_prodotto, tree))

def on_double_click(event, tree):
    # Get the current cliente saved in the tree
    cliente = tree.cliente
    print(cliente)
    if cliente:
        # Retrieve the customer's full info using the cliente name
        cliente_info = get_cliente_info(cliente)
        print(cliente_info)

        if cliente_info:
            # Show the action dialog with the customer's info
            show_action_dialog(cliente, lambda action: handle_action(
                action, tree,
                cliente_info.get("Ragione sociale", ""),
                cliente_info.get("2° riga rag. sociale", ""),
                cliente_info.get("Indirizzo", ""),
                cliente_info.get("CAP", ""),
                cliente_info.get("Città", ""),
                cliente_info.get("Nazione", ""),
                cliente_info.get("Partita iva estero", ""),
                cliente_info.get("Esente IVA", ""),
                cliente_info.get("Telefono", ""),
                cliente_info.get("Email", ""),
                cliente_info.get("Zona", ""),
                cliente_info.get("Giorni di chiusura ", ""),
                cliente_info.get("Orari di scarico", ""),
                cliente_info.get("Condizioni pagamamento", ""),
                cliente_info.get("Sconto", ""),
                cliente_info.get("Agente 1", ""),
                cliente_info.get("ID", "")
            ))
        else:
            messagebox.showwarning("Attenzione", "Nessun dato disponibile per il cliente selezionato.")
    else:
        messagebox.showwarning("Attenzione", "Nessun cliente selezionato.")

button_font = ("Arial", dashboard_font_size)  # Font più grande per i pulsanti
button_width = 15  # Larghezza maggiore per i pulsanti
button_height = 2  # Altezza maggiore per i pulsanti
def show_action_dialog(ragione_sociale, callback):
    dialog = tk.Toplevel()
    dialog.title("Scegli Azione")
    dialog.grab_set()  # Ottieni il focus sulla finestra di dialogo
    dialog.transient()  # Rendi la finestra di dialogo modale

    label = tk.Label(dialog, text=f"Vuoi modificare o eliminare '{ragione_sociale}'?",
                         font=("Arial", dashboard_font_size))
    label.pack(pady=10)

    button_frame = tk.Frame(dialog)
    button_frame.pack(pady=10)

    def on_modify():
        dialog.destroy()
        callback("modify")

    def on_delete():
        dialog.destroy()
        callback("delete")

    modify_button = tk.Button(button_frame, text="Modifica", font=button_font, width=button_width,
                              height=button_height, command=on_modify)
    modify_button.grid(row=0, column=0, padx=10, pady=10)

    delete_button = tk.Button(button_frame, text="Elimina", font=button_font, width=button_width, height=button_height,
                              command=on_delete)
    delete_button.grid(row=0, column=1, padx=10, pady=10)

    center_window(dialog,800,400)
    dialog.wait_window()  # Attendi la chiusura della finestra di dialogo

def show_action_dialog_1(current_prodotto, callback):
    dialog = tk.Toplevel()
    dialog.title("Scegli Azione")
    dialog.grab_set()  # Ottieni il focus sulla finestra di dialogo
    dialog.transient()  # Rendi la finestra di dialogo modale

    label = tk.Label(dialog, text=f"Vuoi modificare o eliminare '{current_prodotto}'?",font=("Arial", dashboard_font_size))
    label.pack(pady=10)

    button_frame = tk.Frame(dialog)
    button_frame.pack(pady=10)

    def on_modify():
        dialog.destroy()
        callback("modify")

    def on_delete():
        dialog.destroy()
        callback("delete")

    modify_button = tk.Button(button_frame, text="Nuova quantità", font=button_font, width=button_width,
                              height=button_height, command=on_modify)
    modify_button.grid(row=0, column=0, padx=10, pady=10)

    delete_button = tk.Button(button_frame, text="Elimina", font=button_font, width=button_width, height=button_height,
                              command=on_delete)
    delete_button.grid(row=0, column=1, padx=10, pady=10)

    center_window(dialog,600,300)
    dialog.wait_window()  # Attendi la chiusura della finestra di dialogo

def handle_action(action,tree, ragione_sociale, seconda_riga, indirizzo, cap, citta, nazione, partita_iva,esente_iva, telefono, email, zona, giorni_chiusura, orari_scarico, condizioni_pagamento, sconto, agente, id_cliente):
    if action == "modify":
        details = ask_details(ragione_sociale, f"Inserisci le nuove info di '{ragione_sociale}':",seconda_riga, indirizzo, cap, citta, nazione, partita_iva,esente_iva, telefono, email, zona, giorni_chiusura, orari_scarico, condizioni_pagamento, sconto, agente, id_cliente)
        if details and details['quantity'] is not None:
            #inserisci funzione per modificare a db il cliente
            show_cliente_info(tree, tree.table_frame)
            messagebox.showinfo("Modifica Prodotto", f"Hai modificato '{ragione_sociale}'.")
    elif action == "delete":
        confirm = messagebox.askokcancel("Conferma Eliminazione", f"Sei sicuro di voler eliminare '{ragione_sociale}'?")
        if confirm:
            #inserisci funzione per far delete cliente
            show_cliente_info(tree.table_frame)
            messagebox.showinfo("Eliminazione Prodotto", f"Il cliente '{ragione_sociale}' è stato eliminato.")

def ask_details(ragione_sociale, prompt,seconda_riga, indirizzo, cap, citta, nazione, partita_iva,esente_iva, telefono, email, zona, giorni_chiusura, orari_scarico, condizioni_pagamento, sconto, agente, id_cliente):
    dialog = tk.Toplevel()
    dialog.title("Scegli le nuove info del cliente")
    dialog.grab_set()
    dialog.transient()

    label = tk.Label(dialog, text=prompt, font=("Arial", dashboard_font_size))
    label.pack(pady=10)

    entry_width = 40

    input_vars = {
        "ragione sociale": tk.StringVar(value=ragione_sociale),
        "seconda riga": tk.StringVar(value=seconda_riga),
        "indirizzo": tk.StringVar(value=indirizzo),
        "cap": tk.StringVar(value=cap),
        "citta": tk.StringVar(value=citta),
        "nazione": tk.StringVar(value=nazione),
        "partita iva": tk.StringVar(value=partita_iva),
        "esente iva": tk.StringVar(value=esente_iva),
        "telefono": tk.StringVar(value=telefono),
        "email": tk.StringVar(value=email),
        "zona": tk.StringVar(value=zona),
        "giorni chiusura": tk.StringVar(value=giorni_chiusura),
        "orari scarico": tk.StringVar(value=orari_scarico),
        "condizioni pagamento": tk.StringVar(value=condizioni_pagamento),
        "sconto": tk.DoubleVar(value=sconto),
        "agente": tk.StringVar(value=agente),
        "id cliente": tk.IntVar(value=id_cliente),
    }

    container_frame = tk.Frame(dialog)
    container_frame.pack(fill="both", expand=True, padx=60)

    canvas = tk.Canvas(container_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(container_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.create_window((0, 0), window=scrollable_frame, anchor="n")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Crea le entry per ciascun campo
    for label_text, var in input_vars.items():
        label = tk.Label(scrollable_frame, text=label_text.capitalize() + ":", font=("Arial", dashboard_font_size))
        label.pack(pady=5)
        if label_text == "condizioni pagamento":
            clienti = get_all_clienti_names()  # Da modificare
            condizione_selezionata = tk.StringVar()
            condizione = ttk.Combobox(scrollable_frame, textvariable=condizione_selezionata, values=clienti,
                                      font=("Arial", dashboard_font_size))
            condizione.configure(width=35)
            condizione.option_add('*TCombobox*Listbox*Font', ('Arial', dashboard_font_size))
            condizione.pack(pady=5)
        elif label_text == "esente iva":
            esenzioni = ["si","no"]  # Da modificare
            esente_selezionata = tk.StringVar()
            esenzione = ttk.Combobox(scrollable_frame, textvariable=esente_selezionata, values=esenzioni,
                                      font=("Arial", dashboard_font_size))
            esenzione.configure(width=35)
            esenzione.option_add('*TCombobox*Listbox*Font', ('Arial', dashboard_font_size))
            esenzione.pack(pady=5)
        elif label_text == "agente":
            agenti = get_all_clienti_names()  # Da modificare
            agente_selezionato = tk.StringVar()
            agente = ttk.Combobox(scrollable_frame, textvariable=agente_selezionato, values=agenti,
                                      font=("Arial", dashboard_font_size))
            agente.configure(width=35)
            agente.option_add('*TCombobox*Listbox*Font', ('Arial', dashboard_font_size))
            agente.pack(pady=5)
        else:
            entry = tk.Entry(scrollable_frame, width=entry_width, textvariable=var, font=("Arial", dashboard_font_size))
            entry.pack(pady=5)

    label = tk.Label(scrollable_frame, text="Data prima lista", font=("Arial", dashboard_font_size))
    label.pack(pady=5)
    start_date_entry = tk.Entry(scrollable_frame, width=12, font=('Arial', dashboard_font_size))
    start_date_entry.insert(0, "dd/mm/yy")
    start_date_entry.bind("<1>", lambda event: pick_date(event, scrollable_frame, start_date_entry))
    start_date_entry.pack(pady=5)

    label = tk.Label(scrollable_frame, text="Data ultima lista", font=("Arial", dashboard_font_size))
    label.pack(pady=5)
    end_date_entry = tk.Entry(scrollable_frame, width=12, font=('Arial', dashboard_font_size))
    end_date_entry.insert(0, "dd/mm/yy")
    end_date_entry.bind("<1>", lambda event: pick_date(event, scrollable_frame, end_date_entry))
    end_date_entry.pack(pady=5)

    def on_confirm():
        dialog.destroy()
        dialog.details = {label: var.get() for label, var in input_vars.items()}

    confirm_button = ctk.CTkButton(dialog, text="Conferma",  font=("Arial", dashboard_font_size) , command=on_confirm, width=120, height=30)
    confirm_button.pack(pady=10)

    center_window(dialog,700,700)
    dialog.wait_window()
    return getattr(dialog, 'details', None)

def handle_action_1(action, current_ragione_sociale,current_id,current_prodotto,current_quantita, tree):
    if action == "modify":
        details = ask_details_1(current_ragione_sociale,f"Inserisci il nuovo id, descrizione e quantità per '{current_prodotto}':", current_id,current_prodotto,current_quantita)
        if details and details['quantity'] is not None:
            #update_record_andria(product, details['quantity'], details['cartons'])
            show_lista_info(current_ragione_sociale, tree, tree.table_frame)
            messagebox.showinfo("Aggiunta Prodotto", f"Hai aggiunto {details['quantity']} prodotti di '{current_prodotto}'.")
    elif action == "delete":
        confirm = messagebox.askokcancel("Conferma Eliminazione", f"Sei sicuro di voler eliminare '{current_prodotto}'?")
        if confirm:
            #delete_record_andria(product)
            show_lista_info(current_ragione_sociale, tree, tree.table_frame)
            messagebox.showinfo("Eliminazione Prodotto", f"Il prodotto '{current_prodotto}' è stato eliminato.")

def ask_details_1(current_ragione_sociale, prompt,current_id,current_prodotto,current_quantita):
    dialog = tk.Toplevel()
    dialog.title("Scegli le nuove info della lista personalizzata")
    dialog.grab_set()
    dialog.transient()

    label = tk.Label(dialog, text=prompt, font=("Arial", 14))
    label.pack(pady=10)

    # Inizializza i campi con i valori correnti
    id_var = tk.IntVar(value=current_id)
    prodotto_var = tk.IntVar(value=current_prodotto)
    quantity_var = tk.IntVar(value=current_quantita)

    id_label = tk.Label(dialog, text=f"Id di '{current_prodotto}':", font=("Arial", dashboard_font_size))
    id_label.pack(pady=5)
    id_entry = tk.Entry(dialog, textvariable=id_var, font=("Arial", 14))
    id_entry.pack(pady=5)

    prodotto_label = tk.Label(dialog, text=f"Descrizione di '{current_prodotto}':", font=("Arial", dashboard_font_size))
    prodotto_label.pack(pady=5)
    prodotto_entry = tk.Entry(dialog, textvariable=prodotto_var, font=("Arial", dashboard_font_size))
    prodotto_entry.pack(pady=5)

    quantity_label = tk.Label(dialog, text=f"Quantità di '{current_prodotto}':",  font= ("Arial", dashboard_font_size))
    quantity_label.pack(pady=5)
    quantity_entry = tk.Entry(dialog, textvariable=quantity_var, font=("Arial", dashboard_font_size))
    quantity_entry.pack(pady=5)

    def on_confirm():
        dialog.destroy()
        dialog.details = {'id': id_var.get(), 'products': prodotto_var.get(), 'quantity': quantity_var.get()}

    confirm_button = ctk.CTkButton(dialog, text="Conferma",  font=("Arial", dashboard_font_size) , command=on_confirm, width=120, height=30)
    confirm_button.pack(pady=10)

    center_window(dialog,600,400)
    dialog.wait_window()
    return getattr(dialog, 'details', None)

def show_dashboard3(parent_frame):
    global combobox, clienti, tree, tree_1
    for widget in parent_frame.winfo_children():
        widget.destroy()

    def open_font_size_popup():
        """Open a popup to choose the font size and reload the dashboard with the new size."""
        popup = tk.Toplevel()
        popup.title("Scegli la dimensione del font")
        popup.geometry("500x250")
        popup.transient()  # Make it modal

        # Label for font size selection
        label = tk.Label(popup, text="Seleziona la dimensione del font:", font=("Arial", dashboard_font_size))
        label.pack(pady=10)

        # Scale widget to select font size
        font_size_var = tk.IntVar(value=dashboard_font_size)
        # Label to display the current slider value on top of the slider handle
        value_display = ctk.CTkLabel(popup, text=str(dashboard_font_size), font=("Arial", dashboard_font_size))
        value_display.place(relx=0.5, rely=0.35, anchor="center")  # Initial position
        # CTkSlider to select font size with inverted color appearance
        # Define the slider's range
        slider_min = 10
        slider_max = 30
        font_slider = ctk.CTkSlider(
            popup,
            from_=10,
            to=30,
            number_of_steps=20,
            fg_color="white",
            progress_color= parent_frame_color,
            command=lambda value: update_slider_value(value) # Sync slider value to font_size_var
        )
        font_slider.set(dashboard_font_size)  # Set initial slider position
        font_slider.pack(pady=10)

        def update_slider_value(value):
            """Update the label text and position to follow the slider handle."""
            font_size_var.set(int(value))  # Update the IntVar with the new slider value
            value_display.configure(text=str(int(value)))  # Update the display label text
            # Position the display label above the slider handle
            slider_pos = font_slider.get()
            display_x = 20 + (slider_pos - slider_min) / (slider_max - slider_min) * 240
            value_display.place(x=display_x, y=60)

        def apply_font_size():
            global dashboard_font_size
            dashboard_font_size = int(font_slider.get())
            popup.destroy()
            # Clear the existing dashboard and reload it with the new font size
            for widget in parent_frame.winfo_children():
                widget.destroy()  # Remove all existing widgets from parent_frame
            show_dashboard3(parent_frame)

        # Button to confirm font size selection
        apply_button = ctk.CTkButton(popup, text="Applica", font=("Arial", dashboard_font_size), command=apply_font_size)
        apply_button.pack(pady=10)

    parent_frame_color = parent_frame.cget("fg_color")

    # Crea un frame per contenere sia il titolo che la legenda
    top_frame = ctk.CTkFrame(parent_frame, corner_radius=5, fg_color=parent_frame_color)
    top_frame.pack(fill="x", pady=0)  # Si espande orizzontalmente

    # Usa il layout grid nel top_frame
    top_frame.columnconfigure(0, weight=0)  # La colonna 0 non si espande
    top_frame.columnconfigure(1, weight=1)  # La colonna 1 si espande

    # Aggiungi la legenda all'interno del top_frame, allineata a destra
    legend_label = ctk.CTkLabel(top_frame, text="Unità di misura:\n- L: liters\n- Kg: kilograms\n- Pz: pieces",
                                font=('Arial', dashboard_font_size), justify="left")
    legend_label.grid(row=0, column=0, sticky="w", padx=10)

    instruction_label = ctk.CTkLabel(top_frame, text="Scegliere Cliente:", font=('Arial', dashboard_font_size))
    instruction_label.grid(row=0, column=1, padx=10)

    clienti = get_all_clienti_names()
    selected_cliente = tk.StringVar()

    # Frame per allineare il menù a tendina e i pulsanti "Esporta in Excel"
    search_frame = ctk.CTkFrame(parent_frame,corner_radius=5)
    search_frame.pack(pady=10)

    # Style for larger font in Combobox
    style = ttk.Style()
    style.configure("TCombobox", font=('Arial', dashboard_font_size))  # Adjust font size as needed
    style.configure("TCombobox*Listbox*Font", font=('Arial', dashboard_font_size))  # Ensure larger font size for dropdown menu items
    
    # Creazione del menù a tendina con ricerca incrementale
    combobox = ttk.Combobox(search_frame, textvariable=selected_cliente, values=clienti, font=('Arial', dashboard_font_size))
    combobox.configure(width=30)
    combobox.pack(side="left", padx=10)

    # Abilita la ricerca incrementale nel menù a tendina
    combobox.bind('<KeyRelease>', on_keyrelease)
    combobox.pack(pady=10)

    def on_cliente_selected(event):
        cliente = selected_cliente.get()
        show_cliente_info(cliente, tree, table_frame)
        show_lista_info(cliente, tree_1, table_frame_1)
        mostra_frame()

    '''
    def on_cliente_selected(event):
        if listbox.curselection():  # Controlla se c'è una selezione
            cliente = listbox.get(listbox.curselection())
            selected_cliente.set(cliente)  # Memorizza il cliente selezionato nell'Entry
            # Nasconde la Listbox dopo la selezione
            listbox_window.withdraw()
            show_cliente_info(cliente, tree, table_frame)
            show_lista_info(cliente, tree_1, table_frame_1)
            mostra_frame()

    # Funzione per aggiornare la Listbox in base alla ricerca
    def update_listbox(*args):
        search_term = entry.get().lower()  # Prende il testo inserito nella barra di ricerca
        listbox.delete(0, tk.END)  # Svuota la Listbox

        if search_term == "":  # Se la barra di ricerca è vuota, nascondi la Listbox
            listbox_window.withdraw()
            return

        # Filtra i clienti e aggiorna la Listbox
        filtered_clienti = [cliente for cliente in clienti if search_term in cliente.lower()]

        if filtered_clienti:
            # Mostra la Listbox sotto la Entry solo se ci sono risultati
            listbox_window.geometry(f"400x200+{parent_frame.winfo_x() + search_frame.winfo_x() + entry.winfo_x() + 20}+{parent_frame.winfo_y() + search_frame.winfo_y() +entry.winfo_y() + search_frame.winfo_height() + entry.winfo_height() + 40}")
            listbox_window.deiconify()

            for cliente in filtered_clienti:
                listbox.insert(tk.END, cliente)
        else:
            listbox_window.withdraw()
            # Nasconde la Listbox se non ci sono risultati

    # Creazione della barra di ricerca (Entry)
    entry = tk.Entry(search_frame, textvariable=selected_cliente, font=('Arial', 16))
    entry.pack(side="left", padx=10,pady=10)
    entry.config(width=33)
    entry.bind('<KeyRelease>', update_listbox)  # Ogni volta che si digita, aggiorna la Listbox

    listbox_window = tk.Toplevel(search_frame)
    listbox_window.wm_overrideredirect(True)  # Rimuove bordi e titoli della finestra
    listbox_window.geometry("400x200")
    listbox_window.withdraw()

    # Creazione della Listbox per visualizzare i risultati (inizialmente nascosta)
    listbox = tk.Listbox(listbox_window, font=('Arial', 16), height=8, selectmode=tk.SINGLE)
    listbox.pack(side="left", fill="both", expand=True)


    # Quando si seleziona un cliente dalla Listbox
    listbox.bind('<<ListboxSelect>>', on_cliente_selected)
    '''
    def mostra_frame():
        table_frame.pack(side="left", padx=5, pady=10, fill="both", expand=True)  # Mostra il primo frame
        table_frame_1.pack(side="left", pady=10, fill="both", expand=True)  # Mostra il secondo frame affiancato

    combobox.bind("<<ComboboxSelected>>", on_cliente_selected)

    # Create a style for the Listbox within the Combobox dropdown
    combobox.option_add('*TCombobox*Listbox*Font', ('Arial', dashboard_font_size))

    # Pulsante "Esporta Excel Fornitori" accanto al pulsante prodotti
    export_fornitori_button = ctk.CTkButton(search_frame, text="Esporta Excel Clienti", command=open_clienti_selection_popup, font=('Arial', dashboard_font_size))
    export_fornitori_button.pack(side="left", padx=10)

    export_lista_button = ctk.CTkButton(search_frame, text="Esporta Excel lista prs", command=open_lista_selection_popup, font=('Arial', dashboard_font_size))
    export_lista_button.pack(side="left", padx=10)

    add_cliente_button = ctk.CTkButton(search_frame, text="Aggiungi Cliente",command = lambda:open_add_popup("Cliente"), font=('Arial', dashboard_font_size))
    add_cliente_button.pack(side="left", padx=10)

    modifica_lista_button = ctk.CTkButton(search_frame, text="Aggiungi a lista prs", command=lambda:open_add_popup("Lista"), font=('Arial', dashboard_font_size))
    modifica_lista_button.pack(side="left", padx=10)

    font_size_button = ctk.CTkButton(search_frame, text="Cambia Dimensione Font", font=("Arial", dashboard_font_size),
                                     command=open_font_size_popup, corner_radius=5)
    font_size_button.pack(side="left", padx=10)



    global table_frame, table_frame_1
    table_frame = ctk.CTkFrame(parent_frame, corner_radius=5)
    table_frame.pack(side="left",padx= 5,pady=10, fill="both", expand=True)
    table_frame.pack_forget()

    table_frame_1 = ctk.CTkFrame(parent_frame, corner_radius=5)
    table_frame_1.pack(side="left",pady=10, fill="both", expand=True)
    table_frame_1.pack_forget()

    table_title = ctk.CTkLabel(table_frame, text="Dati Cliente:", font=('Arial', dashboard_font_size, 'bold'))
    table_title.pack(pady=10)

    table_title_1 = ctk.CTkLabel(table_frame_1, text="Lista personalizzata:", font=('Arial', dashboard_font_size, 'bold'))
    table_title_1.pack(pady=10)

    # Create a frame to hold the table and scrollbars
    table_container = tk.Frame(table_frame, width=500, height=400)
    table_container.pack_propagate(False)  # Prevent the frame from resizing to fit its children
    table_container.pack(pady=10, fill="both", expand=True)

    table_container_1 = tk.Frame(table_frame_1, width=800, height=400)
    table_container_1.pack_propagate(False)  # Prevent the frame from resizing to fit its children
    table_container_1.pack(pady=10, fill="both", expand=True)

    # Use only one column for the Treeview to display data vertically
    global columns
    columns = ("",)

    global columns_1
    columns_1 = ("CODICE","DESCRIZIONE")

    global tree
    tree = ttk.Treeview(table_container, columns=columns, show="headings")
    tree.table_container = table_container # Save the frame in the treeview for reference
    setup_context_menu(tree)
    tree.bind("<B1-Motion>",lambda event: on_mouse_drag(event, tree))

    global tree_1
    tree_1 = ttk.Treeview(table_container_1, columns=columns_1, show="headings")
    tree_1.table_container = table_container_1  # Save the frame in the treeview for reference
    setup_context_menu(tree_1)
    tree_1.bind("<B1-Motion>", lambda event: on_mouse_drag(event, tree_1))


    style = ttk.Style()
    style.configure("Treeview",
                    rowheight=30,
                    font=('Arial', dashboard_font_size),
                    background="#f1f8e9",
                    foreground="#004d40",
                    fieldbackground="#f1f8e9",
                    bordercolor="#000000",
                    relief="solid",
                    borderwidth=1)
    style.configure("Treeview.Heading",
                    font=('Arial', dashboard_font_size, 'bold'),
                    background="#a5d6a7",
                    foreground="#004d40")
    style.map("Treeview",
              background=[('selected', '#c8e6c9')],
              foreground=[('selected', '#004d40')])

    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=1500, anchor="w")

    for col_1 in columns_1:
        tree_1.heading(col_1, text=col_1)
        tree_1.column(col_1, width=100, anchor="center")
    # Set a larger width to accommodate longer text

    yscrollbar = ttk.Scrollbar(table_container, orient=tk.VERTICAL, command=tree.yview)
    yscrollbar.pack(side="right", fill="y")
    tree.configure(yscroll=yscrollbar.set)

    yscrollbar_1 = ttk.Scrollbar(table_container_1, orient=tk.VERTICAL, command=tree.yview)
    yscrollbar_1.pack(side="right", fill="y")
    tree_1.configure(yscroll=yscrollbar_1.set)

    tree.pack(side="left", fill="both", expand=True)
    tree.bind("<Double-1>", lambda event: on_double_click(event, tree))

    tree_1.pack(side="left", fill="both", expand=True)
    tree_1.bind("<Double-1>", lambda event: on_double_click_1(event, tree_1))


# Other components of your application remain unchanged
