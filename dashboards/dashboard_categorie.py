import os
from tkinter.filedialog import asksaveasfilename

import customtkinter as ctk
from tkinter import ttk, Menu
import tkinter as tk
from Database_Utilities.crud_fornitori import get_all_fornitori, get_prodotti_by_fornitore_name, modify_prodotto, delete_prodotto
import openpyxl
from tkinter import messagebox

from popup_functions import open_remove_popup, open_add_popup

dashboard_font_size = 14  # Default font size
def center_window(window, width, height):
    window.update_idletasks()
    width = width
    height = height
    x = (window.winfo_screenwidth() // 2) - (width // 2)
    y = (window.winfo_screenheight() // 2) - (height // 2)
    window.geometry(f'{width}x{height}+{x}+{y}')
def export_categorie_to_excel(selected_fornitori):
    if not selected_fornitori:
        messagebox.showwarning("Nessun Dato", "Non ci sono categorie selezionate da esportare.")
        return

    if len(selected_fornitori) == 1:
        default_filename = f"Categorie_{selected_fornitori[0]}.xlsx"
    else:
        default_filename = f"Categorie_data.xlsx"
    file_path = asksaveasfilename(defaultextension=".xlsx",
                                  filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                  initialfile=default_filename)
    # Verifica se l'utente ha scelto un percorso
    if not file_path:
        messagebox.showwarning("Salvataggio Annullato", "Salvataggio del file Excel annullato.")
        return

    workbook = openpyxl.Workbook()

    for fornitore in selected_fornitori:
        prodotti = get_prodotti_by_fornitore_name(fornitore)

        # Crea un nuovo foglio per ogni fornitore
        sheet = workbook.create_sheet(title=fornitore)

        headers = ["Codice", "Descrizione", "ID Categoria", "Composizione Cartone", "Prezzo Vendita", "Prezzo Acquisto"]
        sheet.append(headers)

        # Aggiungi i prodotti se esistono, altrimenti lascia solo le intestazioni
        if prodotti:
            for prodotto in prodotti:
                sheet.append([prodotto['Codice'], prodotto['Descrizione'], prodotto['ID_FORNITORE'], prodotto['COMPOSIZIONE CARTONE'], prodotto['PREZZO VENDITA'], prodotto['PREZZO ACQUISTO']])
        else:
            # Aggiungi una riga vuota se non ci sono prodotti, lasciando solo l'intestazione
            sheet.append(["Nessun prodotto disponibile"])

    # Rimuove il foglio predefinito solo se ci sono altri fogli visibili
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook["Sheet"]

    workbook.save(file_path)
    messagebox.showinfo("Excel Generato", f"Il file Excel è stato salvato come {os.path.basename(file_path)}")

def open_fornitori_selection_popup():
    popup = tk.Toplevel()
    popup.title("Seleziona Categorie da Esportare")
    popup.geometry("700x700")  # Dimensioni compatte

    instruction_label = ctk.CTkLabel(popup, text="Seleziona le categorie che vuoi esportare:", font=("Arial", 14), text_color="black")
    instruction_label.pack(pady=10)

    container_frame = tk.Frame(popup)
    container_frame.pack(fill="both", expand=True, padx=60)

    canvas = tk.Canvas(container_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(container_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.create_window((0, 0), window=scrollable_frame, anchor="n")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Variabile per tenere traccia dei fornitori selezionati
    selected_fornitori = []

    def on_checkbutton_toggle(fornitore, var):
        if var.get():
            if fornitore not in selected_fornitori:
                selected_fornitori.append(fornitore)
        else:
            if fornitore in selected_fornitori:
                selected_fornitori.remove(fornitore)

    def toggle_select_all():
        all_selected = all(var.get() for var in selections.values())
        if all_selected:
            for var in selections.values():
                var.set(False)
            selected_fornitori.clear()
        else:
            selected_fornitori.clear()
            for fornitore in fornitori:
                var = selections[fornitore]
                var.set(True)
                selected_fornitori.append(fornitore)

    # Aggiungi il pulsante Seleziona/Deseleziona tutto
    select_all_button = ctk.CTkButton(popup, text="Seleziona Tutto", command=toggle_select_all, width=120, height=30)
    select_all_button.pack(pady=5)

    # Dizionario per tracciare le variabili di stato dei checkbutton
    selections = {}

    fornitori = get_all_fornitori()
    for fornitore in fornitori:
        var = tk.BooleanVar()
        selections[fornitore] = var
        check = tk.Checkbutton(scrollable_frame, text=fornitore, variable=var,font=("Arial", dashboard_font_size),
                               command=lambda f=fornitore, v=var: on_checkbutton_toggle(f, v))
        check.pack(anchor="w")

    # Pulsante di conferma
    confirm_button = ctk.CTkButton(popup, text="Conferma", command=lambda: [popup.destroy(), export_categorie_to_excel(selected_fornitori)], width=120, height=30)
    confirm_button.pack(pady=10)

    popup.mainloop()


def get_fornitore_info(fornitore):
    info = get_prodotti_by_fornitore_name(fornitore)
    return info


def show_fornitore_info(fornitore, tree, table_frame):
    # Verifica che la funzione venga eseguita solo una volta
    print(f"Popolando dati per il fornitore: {fornitore}")

    # Cancella tutte le righe esistenti nella treeview
    for row in tree.get_children():
        tree.delete(row)

    # Recupera i prodotti per il fornitore
    info = get_fornitore_info(fornitore)

    print(f"Numero di prodotti trovati: {len(info)}")

    # Inserisci i nuovi prodotti
    for prodotto in info:
        tree.insert("", tk.END, values=(
            prodotto['Codice'], prodotto['Descrizione'], prodotto['ID_FORNITORE'],
            prodotto['COMPOSIZIONE CARTONE'], prodotto['PREZZO VENDITA'], prodotto['PREZZO ACQUISTO'])
                    )

    tree.fornitore = fornitore  # Salva il fornitore corrente nella treeview
    table_frame.pack(pady=10, fill="both", expand=True)


def update_combobox():
    value = combobox.get().lower()
    if value == '':
        combobox['values'] = fornitori
    else:
        data = [item for item in fornitori if value in item.lower()]
        combobox['values'] = data
        combobox.event_generate('<Down>')

def on_keyrelease(event):
    combobox.after(2000, update_combobox)

def copy_selection(tree):
    selected_items = tree.selection()  # Ottiene tutti gli elementi selezionati
    copied_data = []
    for item in selected_items:
        item_values = tree.item(item, "values")
        copied_data.append("\t".join(map(str, item_values)))  # Unisce i valori di ciascuna riga con una tabulazione
    clipboard_text = "\n".join(copied_data)  # Unisce tutte le righe con un ritorno a capo
    tree.clipboard_clear()  # Pulisce gli appunti
    tree.clipboard_append(clipboard_text)  # Aggiunge il testo agli appunti
    messagebox.showinfo("Copiato", "I testi selezionati sono stati copiati negli appunti.")


def setup_context_menu(tree):
    # Creare un menu contestuale
    context_menu = Menu(tree, tearoff=0)
    context_menu.add_command(label="Copia", command=lambda: copy_selection(tree), font=("Arial", dashboard_font_size))

    def on_right_click(event):
        # Mostrare il menu contestuale
        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    tree.bind("<Button-3>", on_right_click)  # <Button-3> è il clic del pulsante destro del mouse

def on_mouse_drag(event):
    # Identifica l'item su cui si trova il cursore
    item = tree.identify_row(event.y)
    if item:
        tree.selection_add(item)

def setup_treeview(tree):
    columns = ("CODICE", "DESCRIZIONE", "CATEGORIA", "COMPOSIZIONE CARTONE", "PREZZO VENDITA", "PREZZO ACQUISTO")
    tree['columns'] = columns
    sort_order = {col: False for col in columns}  # False = Ascendente, True = Discendente

    def treeview_sort_column(tree, col, col_type):
        """Funzione per ordinare le colonne."""
        nonlocal sort_order
        reverse = sort_order[col]
        # Determina il tipo di dato per corretta comparazione nel sort
        if col_type == 'number':
            converter = float
        else:
            converter = str

        l = [(tree.set(k, col), k) for k in tree.get_children('')]
        l.sort(reverse=reverse, key=lambda x: converter(x[0]))

        for index, (val, k) in enumerate(l):
            tree.move(k, '', index)

        sort_order[col] = not reverse
        tree.heading(col, text=col, command=lambda: treeview_sort_column(tree, col, col_type))

    # Configurazione delle intestazioni e delle colonne
    tree.heading("CODICE", text="CODICE", anchor="w",
                 command=lambda: treeview_sort_column(tree, "CODICE", 'text'))
    tree.heading("DESCRIZIONE", text="DESCRIZIONE", anchor="w",
                 command=lambda: treeview_sort_column(tree, "DESCRIZIONE", 'text'))
    tree.heading("CATEGORIA", text="CATEGORIA", anchor="w",
                 command=lambda: treeview_sort_column(tree, "CATEGORIA", 'text'))
    tree.heading("COMPOSIZIONE CARTONE", text="COMPOSIZIONE CARTONE", anchor="w",
                 command=lambda: treeview_sort_column(tree, "COMPOSIZIONE CARTONE", 'text'))
    tree.heading("PREZZO VENDITA", text="PREZZO VENDITA", anchor="w",
                 command=lambda: treeview_sort_column(tree, "PREZZO VENDITA", 'number'))
    tree.heading("PREZZO ACQUISTO", text="PREZZO ACQUISTO", anchor="w",
                 command=lambda: treeview_sort_column(tree, "PREZZO ACQUISTO", 'number'))

    # Impostazione della larghezza delle colonne (adattabile in base alle tue necessità)
    tree.column("CODICE", width=120, anchor="center")
    tree.column("DESCRIZIONE", width=150, anchor="center")
    tree.column("CATEGORIA", width=120, anchor="center")
    tree.column("COMPOSIZIONE CARTONE", width=150, anchor="center")
    tree.column("PREZZO VENDITA", width=120, anchor="center")
    tree.column("PREZZO ACQUISTO", width=120, anchor="center")

    tree.pack(fill="both", expand=True)

def on_double_click(event, tree):
    item = tree.selection()[0]
    values = tree.item(item, "values")
    if values:
        # Recupera i dati salvati
        product = values[0]
        current_descrizione = values[1]
        current_composizione_cartone = values[3]
        current_prezzo_vendita = values[4]
        current_prezzo_acquisto = values[5]
        show_action_dialog(product, lambda action: handle_action(action, product,current_descrizione,current_composizione_cartone,current_prezzo_vendita,current_prezzo_acquisto, tree))

button_font = ("Arial", dashboard_font_size)  # Font più grande per i pulsanti
button_width = 15  # Larghezza maggiore per i pulsanti
button_height = 2  # Altezza maggiore per i pulsanti
def show_action_dialog(product_name, callback):
    dialog = tk.Toplevel()
    dialog.title("Scegli Azione")
    dialog.grab_set()  # Ottieni il focus sulla finestra di dialogo
    dialog.transient()  # Rendi la finestra di dialogo modale

    label = tk.Label(dialog, text=f"Vuoi modificare o eliminare '{product_name}'?",
                         font=("Arial", 16))
    label.pack(pady=10)

    button_frame = tk.Frame(dialog)
    button_frame.pack(pady=10)

    def on_modify():
        dialog.destroy()
        callback("modify")

    def on_delete():
        dialog.destroy()
        callback("delete")

    modify_button = tk.Button(button_frame, text="Modifica", font=button_font, width=button_width,
                              height=button_height, command=on_modify)
    modify_button.grid(row=0, column=0, padx=10, pady=10)

    delete_button = tk.Button(button_frame, text="Elimina", font=button_font, width=button_width, height=button_height,
                              command=on_delete)
    delete_button.grid(row=0, column=1, padx=10, pady=10)

    center_window(dialog,600,300)
    dialog.wait_window()  # Attendi la chiusura della finestra di dialogo

def handle_action(action, product,current_descrizione,current_composizione_cartone,current_prezzo_vendita,current_prezzo_acquisto, tree):
    if action == "modify":
        details = ask_details(product, f"Inserisci le nuove info per '{product}':", current_descrizione,
                              current_composizione_cartone, current_prezzo_vendita, current_prezzo_acquisto)
        if details and details['descrizione'] is not None:
            # Richiamiamo la funzione modify_prodotto per modificare il prodotto nel database
            modify_prodotto(product, details['descrizione'], details['composizione_cartoni'], details['prezzo vendita'],
                            details['prezzo acquisto'])
            show_fornitore_info(tree.fornitore, tree, tree.table_container)  # Aggiorniamo la tabella con i nuovi dati
            messagebox.showinfo("Modifica Prodotto", f"Hai modificato il prodotto '{product}'.")

    elif action == "delete":

        confirm = messagebox.askokcancel("Conferma Eliminazione",
                                         f"Sei sicuro di voler eliminare il prodotto '{product}'?")

        if confirm:
            # Richiamiamo la funzione delete_prodotto per eliminare il prodotto dal database
            delete_prodotto(product)

            show_fornitore_info(tree.fornitore, tree,
                                tree.table_container)  # Aggiorniamo la tabella per riflettere l'eliminazione
            messagebox.showinfo("Eliminazione Prodotto", f"Il prodotto '{product}' è stato eliminato.")


def ask_details(product, prompt, current_descrizione, current_composizione_cartoni, current_prezzo_vendita,
                current_prezzo_acquisto):
    dialog = tk.Toplevel()
    dialog.title("Scegli le nuove info della categoria")
    dialog.grab_set()
    dialog.transient()

    label = tk.Label(dialog, text=prompt, font=("Arial", dashboard_font_size))
    label.pack(pady=10)

    entry_width = 40

    # Inizializza i campi con i valori correnti, usando StringVar per i campi di testo
    descrizione_var = tk.StringVar(value=current_descrizione)
    composizione_cartoni_var = tk.StringVar(value=current_composizione_cartoni)

    # Gestiamo "vuoto" come condizione per prezzo
    prezzo_vendita_var = tk.DoubleVar(value=float(current_prezzo_vendita) if current_prezzo_vendita != "vuoto" else 0.0)
    prezzo_acquisto_var = tk.DoubleVar(
        value=float(current_prezzo_acquisto) if current_prezzo_acquisto != "vuoto" else 0.0)

    # Entry per descrizione (usiamo StringVar)
    descrizione_label = tk.Label(dialog, text="Descrizione:", font=("Arial", dashboard_font_size))
    descrizione_label.pack(pady=5)
    descrizione_entry = tk.Entry(dialog, width=entry_width, textvariable=descrizione_var, font=("Arial", dashboard_font_size))
    descrizione_entry.pack(pady=5)

    # Entry per composizione cartoni (usiamo StringVar)
    composizione_cartoni_label = tk.Label(dialog, text="Composizione Cartoni:", font=("Arial", dashboard_font_size))
    composizione_cartoni_label.pack(pady=5)
    composizione_cartoni_entry = tk.Entry(dialog, width=entry_width, textvariable=composizione_cartoni_var,
                                          font=("Arial", dashboard_font_size))
    composizione_cartoni_entry.pack(pady=5)

    # Entry per prezzo di vendita (usiamo DoubleVar)
    prezzo_vendita_label = tk.Label(dialog, text="Prezzo Vendita:", font=("Arial", dashboard_font_size))
    prezzo_vendita_label.pack(pady=5)
    prezzo_vendita_entry = tk.Entry(dialog, width=entry_width, textvariable=prezzo_vendita_var, font=("Arial", dashboard_font_size))
    prezzo_vendita_entry.pack(pady=5)

    # Entry per prezzo di acquisto (usiamo DoubleVar)
    prezzo_acquisto_label = tk.Label(dialog, text="Prezzo Acquisto:", font=("Arial", dashboard_font_size))
    prezzo_acquisto_label.pack(pady=5)
    prezzo_acquisto_entry = tk.Entry(dialog, width=entry_width, textvariable=prezzo_acquisto_var, font=("Arial", dashboard_font_size))
    prezzo_acquisto_entry.pack(pady=5)

    def on_confirm():
        dialog.destroy()
        dialog.details = {
            'descrizione': descrizione_var.get(),
            'composizione_cartoni': composizione_cartoni_var.get(),
            'prezzo vendita': prezzo_vendita_var.get(),
            'prezzo acquisto': prezzo_acquisto_var.get(),
        }

    confirm_button = ctk.CTkButton(dialog, text="Conferma", font=("Arial", dashboard_font_size), command=on_confirm, width=120, height=30)
    confirm_button.pack(pady=10)

    center_window(dialog, 600, 500)
    dialog.wait_window()
    return getattr(dialog, 'details', None)

def show_dashboard2(parent_frame):
    global combobox, fornitori, tree
    for widget in parent_frame.winfo_children():
        widget.destroy()

    def open_font_size_popup():
        """Open a popup to choose the font size and reload the dashboard with the new size."""
        popup = tk.Toplevel()
        popup.title("Scegli la dimensione del font")
        popup.geometry("500x250")
        popup.transient()  # Make it modal

        # Label for font size selection
        label = tk.Label(popup, text="Seleziona la dimensione del font:", font=("Arial", dashboard_font_size))
        label.pack(pady=10)

        # Scale widget to select font size
        font_size_var = tk.IntVar(value=dashboard_font_size)
        # Label to display the current slider value on top of the slider handle
        value_display = ctk.CTkLabel(popup, text=str(dashboard_font_size), font=("Arial", 12))
        value_display.place(relx=0.5, rely=0.35, anchor="center")  # Initial position
        # CTkSlider to select font size with inverted color appearance
        slider_min = 10
        slider_max = 30
        font_slider = ctk.CTkSlider(
            popup,
            from_=10,
            to=30,
            number_of_steps=20,
            fg_color="white",
            progress_color= parent_frame_color,
            command=lambda value: update_slider_value(value) # Sync slider value to font_size_var
        )
        font_slider.set(dashboard_font_size)  # Set initial slider position
        font_slider.pack(pady=10)

        def update_slider_value(value):
            """Update the label text and position to follow the slider handle."""
            font_size_var.set(int(value))  # Update the IntVar with the new slider value
            value_display.configure(text=str(int(value)))  # Update the display label text
            # Position the display label above the slider handle
            slider_pos = font_slider.get()
            display_x = 20 + (slider_pos - slider_min) / (slider_max - slider_min) * 240
            value_display.place(x=display_x, y=60)

        def apply_font_size():
            global dashboard_font_size
            dashboard_font_size = int(font_slider.get())
            popup.destroy()
            # Clear the existing dashboard and reload it with the new font size
            for widget in parent_frame.winfo_children():
                widget.destroy()  # Remove all existing widgets from parent_frame
            show_dashboard2(parent_frame)

        # Button to confirm font size selection
        apply_button = ctk.CTkButton(popup, text="Applica", font=("Arial", dashboard_font_size), command=apply_font_size)
        apply_button.pack(pady=10)

    parent_frame_color = parent_frame.cget("fg_color")

    # Crea un frame per contenere sia il titolo che la legenda
    top_frame = ctk.CTkFrame(parent_frame, corner_radius=5, fg_color=parent_frame_color)
    top_frame.pack(fill="x", pady=0)  # Si espande orizzontalmente

    top_frame.columnconfigure(0, weight=0)  # La colonna 0 non si espande
    top_frame.columnconfigure(1, weight=1)  # La colonna 1 si espande

    # Aggiungi l'etichetta "Scegli una città" all'interno del top_frame, allineata a sinistra
    label = ctk.CTkLabel(top_frame, text="Scegli una categoria:", font=('Arial', dashboard_font_size))
    label.grid(row=0, column=1, padx=10)

    # Aggiungi la legenda all'interno del top_frame, allineata a destra
    legend_label = ctk.CTkLabel(top_frame, text="Unità di misura:\n- L: liters\n- Kg: kilograms\n- Pz: pieces",
                                font=('Arial', dashboard_font_size), justify="left")
    legend_label.grid(row=0, column=0, sticky="w", padx=10)

    fornitori = get_all_fornitori()
    selected_fornitore = tk.StringVar()

    # Frame per allineare il menù a tendina e i pulsanti "Esporta in Excel"
    search_frame = ctk.CTkFrame(parent_frame,corner_radius=5)
    search_frame.pack(pady=10)

    # Creazione del menù a tendina con ricerca incrementale
    combobox = ttk.Combobox(search_frame, textvariable=selected_fornitore, values=fornitori, font=('Arial', dashboard_font_size))
    combobox.pack(side="left", padx=10,pady=10)

    # Abilita la ricerca incrementale nel menù a tendina
    combobox.bind('<KeyRelease>', on_keyrelease)

    def on_fornitore_selected(event):
        fornitore = selected_fornitore.get()
        show_fornitore_info(fornitore, tree, table_frame)

    combobox.bind("<<ComboboxSelected>>", on_fornitore_selected)

    # Aumenta la larghezza del menù a tendina
    combobox.configure(width=20)

    # Create a style for the Listbox within the Combobox dropdown
    combobox.option_add('*TCombobox*Listbox*Font', ('Arial', dashboard_font_size))
    '''
    def on_fornitore_selected(event):
        if listbox.curselection():  # Controlla se c'è una selezione
            fornitore = listbox.get(listbox.curselection())
            selected_fornitore.set(fornitore)  # Memorizza il cliente selezionato nell'Entry
            # Nasconde la Listbox dopo la selezione
            listbox_window.withdraw()
            show_fornitore_info(fornitore, tree, table_frame)

    # Funzione per aggiornare la Listbox in base alla ricerca
    def update_listbox(*args):
        search_term = entry.get().lower()  # Prende il testo inserito nella barra di ricerca
        listbox.delete(0, tk.END)  # Svuota la Listbox

        if search_term == "":  # Se la barra di ricerca è vuota, nascondi la Listbox
            listbox_window.withdraw()
            return

        # Filtra i clienti e aggiorna la Listbox
        filtered_fornitori = [fornitore for fornitore in fornitori if search_term in fornitore.lower()]

        if filtered_fornitori:
            # Mostra la Listbox sotto la Entry solo se ci sono risultati
            listbox_window.geometry(f"300x200+{parent_frame.winfo_x() + search_frame.winfo_x() + entry.winfo_x() + 20}+{parent_frame.winfo_y() + search_frame.winfo_y() +entry.winfo_y() + search_frame.winfo_height() + entry.winfo_height() + 40}")
            listbox_window.deiconify()

            for fornitore in filtered_fornitori:
                listbox.insert(tk.END, fornitore)
        else:
            listbox_window.withdraw()
            #listbox.place_forget()
            # Nasconde la Listbox se non ci sono risultati

    # Creazione della barra di ricerca (Entry)
    entry = tk.Entry(search_frame, textvariable=selected_fornitore, font=('Arial', 16))
    entry.pack(side="left", padx=10,pady=10)
    entry.config(width=25)
    entry.bind('<KeyRelease>', update_listbox)  # Ogni volta che si digita, aggiorna la Listbox

    listbox_window = tk.Toplevel(search_frame)
    listbox_window.wm_overrideredirect(True)  # Rimuove bordi e titoli della finestra
    listbox_window.geometry("300x200")
    listbox_window.withdraw()

    # Creazione della Listbox per visualizzare i risultati (inizialmente nascosta)
    listbox = tk.Listbox(listbox_window, font=('Arial', 16), height=8, selectmode=tk.SINGLE)
    listbox.pack(side="left", fill="both", expand=True)


    # Quando si seleziona un cliente dalla Listbox
    listbox.bind('<<ListboxSelect>>', on_fornitore_selected)
    '''

    # Pulsante "Esporta Excel Fornitori" accanto al pulsante prodotti
    export_fornitori_button = ctk.CTkButton(search_frame, text="Esporta Excel Categoria", command=open_fornitori_selection_popup, font=("Arial", dashboard_font_size))
    export_fornitori_button.pack(side="left", padx=10)


    add_product_button = ctk.CTkButton(search_frame, text="Aggiungi prodotto",command = lambda: open_add_popup("Prodotto"), font=("Arial", dashboard_font_size))
    add_product_button.pack(side="left", padx=10)


    add_fornitore_button = ctk.CTkButton(search_frame, text="Aggiungi Categoria",command = lambda:open_add_popup("Fornitore"), font=("Arial", dashboard_font_size))
    add_fornitore_button.pack(side="left", padx=10)

    font_size_button = ctk.CTkButton(search_frame, text="Cambia Dimensione Font", font=("Arial", dashboard_font_size), command=open_font_size_popup, corner_radius=5)
    font_size_button.pack(side="left", padx=10)

    global table_frame
    table_frame = ctk.CTkFrame(parent_frame, corner_radius=5)
    table_frame.pack_forget()

    table_title = ctk.CTkLabel(table_frame, text=" Categoria:", font=('Arial', dashboard_font_size, 'bold'))
    table_title.pack(pady=10)

    # Create a frame to hold the table and scrollbars
    table_container = tk.Frame(table_frame, width=1600, height=400)
    table_container.pack_propagate(False)  # Prevent the frame from resizing to fit its children
    table_container.pack(pady=10, fill="both", expand=True)

    global columns
    columns = ("CODICE", "DESCRIZIONE", "CATEGORIA", "COMPOSIZIONE CARTONE", "PREZZO VENDITA", "PREZZO ACQUISTO")

    global tree
    tree = ttk.Treeview(table_container, columns=columns,style= "dash2.Treeview", show="headings", selectmode="extended")
    tree.table_container = table_container  # Salva il frame nella treeview per riferimento
    setup_treeview(tree)
    setup_context_menu(tree)
    tree.bind("<B1-Motion>", on_mouse_drag)

    style = ttk.Style()
    style.configure("dash2.Treeview",
                    rowheight=30,
                    font=('Arial', dashboard_font_size),
                    background="#f1f8e9",
                    foreground="#004d40",
                    fieldbackground="#f1f8e9",
                    bordercolor="#000000",
                    relief="solid",
                    borderwidth=1)
    style.configure("dash2.Treeview.Heading",
                    font=('Arial', dashboard_font_size, 'bold'),
                    background="#a5d6a7",
                    foreground="#004d40")
    style.map("dash2.Treeview",
              background=[('selected', '#c8e6c9')],
              foreground=[('selected', '#004d40')])

    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=250, anchor="center")  # Adjust width as needed

    yscrollbar = ttk.Scrollbar(tree, orient=tk.VERTICAL, command=tree.yview)
    yscrollbar.pack(side="right", fill="y")
    tree.configure(yscroll=yscrollbar.set)

    xscrollbar = ttk.Scrollbar(tree, orient=tk.HORIZONTAL, command=tree.xview)
    xscrollbar.pack(side="bottom", fill="x")
    tree.configure(xscroll=xscrollbar.set)

    tree.pack(side="left", fill="both", expand=True)
    tree.bind("<Double-1>", lambda event: on_double_click(event, tree))
