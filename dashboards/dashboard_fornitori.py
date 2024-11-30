import datetime
import os
import sqlite3
from tkinter.filedialog import asksaveasfilename

import customtkinter as ctk
from tkinter import ttk, Menu, filedialog
import tkinter as tk

from fpdf import FPDF

from Database_Utilities.crud_clienti import get_all_clienti_names, get_cliente_info_by_name
import openpyxl
from tkinter import messagebox

dashboard_font_size = 14
def center_window(window, width, height):
    window.update_idletasks()
    width = width
    height = height
    x = (window.winfo_screenwidth() // 2) - (width // 2)
    y = (window.winfo_screenheight() // 2) - (height // 2)
    window.geometry(f'{width}x{height}+{x}+{y}')
def export_clienti_to_excel(selected_clienti):
    if not selected_clienti:
        messagebox.showwarning("Nessun Dato", "Non ci sono fornitori selezionati da esportare.")
        return

    if len(selected_clienti) == 1:
        default_filename = f"Info_Fornitore_{selected_clienti[0]}.xlsx"
    else:
        default_filename = f"Info_Fornitori.xlsx"
    file_path = asksaveasfilename(defaultextension=".xlsx",
                                  filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                  initialfile=default_filename)
    # Verifica se l'utente ha scelto un percorso
    if not file_path:
        messagebox.showwarning("Salvataggio Annullato", "Salvataggio del file Excel annullato.")
        return

    workbook = openpyxl.Workbook()

    for cliente in selected_clienti:
        info = get_cliente_info_by_name(cliente)

        # Crea un nuovo foglio per ogni cliente
        sheet = workbook.create_sheet(title=cliente)

        headers = ["Ragione sociale", "Indirizzo", "CAP", "Città", "Nazione", "Partita IVA", "Telefono", "Email", "Zona", "Giorni chiusura", "Orari di scarico", "Condizioni pagamento", "Sconto", "Agente"]
        sheet.append(headers)

        # Aggiungi i dati del cliente se esistono, altrimenti lascia solo le intestazioni
        if info:
            sheet.append([info.get(header, "N/A") for header in headers])
        else:
            # Aggiungi una riga vuota se non ci sono dati, lasciando solo l'intestazione
            sheet.append(["Nessun dato disponibile"])

    # Rimuove il foglio predefinito solo se ci sono altri fogli visibili
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook["Sheet"]

    workbook.save(file_path)
    messagebox.showinfo("Excel Generato", f"Il file Excel è stato salvato come {os.path.basename(file_path)}")
def open_clienti_selection_popup():
    popup = tk.Toplevel()
    popup.title("Seleziona Fornitori da Esportare")
    popup.geometry("700x700")  # Dimensioni compatte

    instruction_label = ctk.CTkLabel(popup, text="Seleziona i fornitori che vuoi esportare:", font=('Arial', dashboard_font_size), text_color="black")
    instruction_label.pack(pady=10)

    container_frame = tk.Frame(popup)
    container_frame.pack(fill="both", expand=True, padx=60)

    canvas = tk.Canvas(container_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(container_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.create_window((0, 0), window=scrollable_frame, anchor="n")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Variabile per tenere traccia dei clienti selezionati
    selected_clienti = []

    def on_checkbutton_toggle(cliente, var):
        if var.get():
            if cliente not in selected_clienti:
                selected_clienti.append(cliente)
        else:
            if cliente in selected_clienti:
                selected_clienti.remove(cliente)

    def toggle_select_all():
        all_selected = all(var.get() for var in selections.values())
        if all_selected:
            for var in selections.values():
                var.set(False)
            selected_clienti.clear()
        else:
            selected_clienti.clear()
            for cliente in clienti:
                var = selections[cliente]
                var.set(True)
                selected_clienti.append(cliente)

    # Aggiungi il pulsante Seleziona/Deseleziona tutto
    select_all_button = ctk.CTkButton(popup, text="Seleziona Tutto", command=toggle_select_all, width=120, height=30)
    select_all_button.pack(pady=5)

    # Dizionario per tracciare le variabili di stato dei checkbutton
    selections = {}

    clienti = get_all_clienti_names()
    for cliente in clienti:
        var = tk.BooleanVar()
        selections[cliente] = var
        check = tk.Checkbutton(scrollable_frame, text=cliente, variable=var,font=('Arial', dashboard_font_size),
                               command=lambda c=cliente, v=var: on_checkbutton_toggle(c, v))
        check.pack(anchor="w")

    # Pulsante di conferma
    confirm_button = ctk.CTkButton(popup, text="Conferma",font=('Arial',dashboard_font_size), command=lambda: [popup.destroy(), export_clienti_to_excel(selected_clienti)], width=120, height=30)
    confirm_button.pack(pady=10)

    popup.mainloop()
def get_cliente_info(cliente):
    info = get_cliente_info_by_name(cliente)
    return info

def show_cliente_info(cliente, tree, table_frame):
    for row in tree.get_children():
        tree.delete(row)
    info = get_cliente_info(cliente)
    if info:
        for key, value in info.items():
            tree.insert("", tk.END, values=[f"{key}: {value}"])
    tree.cliente = cliente # Save the current client in the tree
    table_frame.pack(pady=10, fill="both", expand=True)
def update_combobox():
    value = combobox.get().lower()
    if value == '':
        combobox['values'] = clienti
    else:
        data = [item for item in clienti if value in item.lower()]
        combobox['values'] = data
        combobox.event_generate('<Down>')

def on_keyrelease(event):
    combobox.after(2000, update_combobox)

def copy_selection(tree):
    selected_items = tree.selection()  # Ottiene tutti gli elementi selezionati
    copied_data = []
    for item in selected_items:
        item_values = tree.item(item, "values")
        copied_data.append("\t".join(map(str, item_values)))  # Unisce i valori di ciascuna riga con una tabulazione
    clipboard_text = "\n".join(copied_data)  # Unisce tutte le righe con un ritorno a capo
    tree.clipboard_clear()  # Pulisce gli appunti
    tree.clipboard_append(clipboard_text)  # Aggiunge il testo agli appunti
    messagebox.showinfo("Copiato", "I testi selezionati sono stati copiati negli appunti.")


def setup_context_menu(tree):
    # Creare un menu contestuale
    context_menu = Menu(tree, tearoff=0)
    context_menu.add_command(label="Copia", command=lambda: copy_selection(tree), font=('Arial', dashboard_font_size))

    def on_right_click(event):
        # Mostrare il menu contestuale
        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    tree.bind("<Button-3>", on_right_click)  # <Button-3> è il clic del pulsante destro del mouse

# Aggiungere il supporto per la selezione trascinando
def on_mouse_drag(event):
    # Identifica l'item su cui si trova il cursore
    item = tree.identify_row(event.y)
    if item:
        tree.selection_add(item)


def on_double_click(event, tree):
    # Get the current cliente saved in the tree
    cliente = tree.cliente
    print(cliente)
    if cliente:
        # Retrieve the customer's full info using the cliente name
        cliente_info = get_cliente_info(cliente)
        print(cliente_info)

        if cliente_info:
            # Show the action dialog with the customer's info
            show_action_dialog(cliente, lambda action: handle_action(
                action, tree,
                cliente_info.get("Ragione sociale", ""),
                cliente_info.get("2° riga rag. sociale", ""),
                cliente_info.get("Indirizzo", ""),
                cliente_info.get("CAP", ""),
                cliente_info.get("Città", ""),
                cliente_info.get("Nazione", ""),
                cliente_info.get("Partita iva estero", ""),
                cliente_info.get("Telefono", ""),
                cliente_info.get("Email", ""),
                cliente_info.get("Zona", ""),
                cliente_info.get("Giorni di chiusura ", ""),
                cliente_info.get("Orari di scarico", ""),
                cliente_info.get("Condizioni pagamamento", ""),
                cliente_info.get("Sconto", ""),
                cliente_info.get("Agente 1", ""),
                cliente_info.get("ID", "")
            ))
        else:
            messagebox.showwarning("Attenzione", "Nessun dato disponibile per il fornitore selezionato.")
    else:
        messagebox.showwarning("Attenzione", "Nessun cliente selezionato.")
button_font = ("Arial", dashboard_font_size)  # Font più grande per i pulsanti
button_width = 15  # Larghezza maggiore per i pulsanti
button_height = 2  # Altezza maggiore per i pulsanti
def show_action_dialog(ragione_sociale, callback):
    dialog = tk.Toplevel()
    dialog.title("Scegli Azione")
    dialog.grab_set()  # Ottieni il focus sulla finestra di dialogo
    dialog.transient()  # Rendi la finestra di dialogo modale

    label = tk.Label(dialog, text=f"Vuoi modificare o eliminare il fornitore '{ragione_sociale}'?",
                         font=("Arial", dashboard_font_size))
    label.pack(pady=10)

    button_frame = tk.Frame(dialog)
    button_frame.pack(pady=10)

    def on_modify():
        dialog.destroy()
        callback("modify")

    def on_delete():
        dialog.destroy()
        callback("delete")

    modify_button = tk.Button(button_frame, text="Modifica", font=button_font, width=button_width,
                              height=button_height, command=on_modify)
    modify_button.grid(row=0, column=0, padx=10, pady=10)

    delete_button = tk.Button(button_frame, text="Elimina", font=button_font, width=button_width, height=button_height,
                              command=on_delete)
    delete_button.grid(row=0, column=1, padx=10, pady=10)

    center_window(dialog,800,400)
    dialog.wait_window()  # Attendi la chiusura della finestra di dialogo

def handle_action(action,tree, ragione_sociale, seconda_riga, indirizzo, cap, citta, nazione, partita_iva, telefono, email, zona, giorni_chiusura, orari_scarico, condizioni_pagamento, sconto, agente, id_cliente):
    if action == "modify":
        details = ask_details(ragione_sociale, f"Inserisci le nuove info di '{ragione_sociale}':",seconda_riga, indirizzo, cap, citta, nazione, partita_iva, telefono, email, zona, giorni_chiusura, orari_scarico, condizioni_pagamento, sconto, agente, id_cliente)
        if details and details['quantity'] is not None:
            #inserisci funzione per modificare a db il cliente
            show_cliente_info(tree, tree.table_frame)
            messagebox.showinfo("Modifica Prodotto", f"Hai modificato '{ragione_sociale}'.")
    elif action == "delete":
        confirm = messagebox.askokcancel("Conferma Eliminazione", f"Sei sicuro di voler eliminare '{ragione_sociale}'?")
        if confirm:
            #inserisci funzione per far delete cliente
            show_cliente_info(tree.table_frame)
            messagebox.showinfo("Eliminazione Prodotto", f"Il fornitore '{ragione_sociale}' è stato eliminato.")

def ask_details(ragione_sociale, prompt,seconda_riga, indirizzo, cap, citta, nazione, partita_iva, telefono, email, zona, giorni_chiusura, orari_scarico, condizioni_pagamento, sconto, agente, id_cliente):
    dialog = tk.Toplevel()
    dialog.title("Scegli le nuove info del fornitore")
    dialog.grab_set()
    dialog.transient()

    label = tk.Label(dialog, text=prompt, font=("Arial", dashboard_font_size))
    label.pack(pady=10)

    entry_width = 40

    input_vars = {
        "ragione sociale": tk.StringVar(value=ragione_sociale),
        "seconda riga": tk.StringVar(value=seconda_riga),
        "indirizzo": tk.StringVar(value=indirizzo),
        "cap": tk.StringVar(value=cap),
        "citta": tk.StringVar(value=citta),
        "nazione": tk.StringVar(value=nazione),
        "partita iva": tk.StringVar(value=partita_iva),
        "telefono": tk.StringVar(value=telefono),
        "email": tk.StringVar(value=email),
        "zona": tk.StringVar(value=zona),
        "giorni chiusura": tk.StringVar(value=giorni_chiusura),
        "orari scarico": tk.StringVar(value=orari_scarico),
        "condizioni pagamento": tk.StringVar(value=condizioni_pagamento),
        "sconto": tk.DoubleVar(value=sconto),
        "agente": tk.StringVar(value=agente),
        "id cliente": tk.IntVar(value=id_cliente),
    }

    container_frame = tk.Frame(dialog)
    container_frame.pack(fill="both", expand=True, padx=60)

    canvas = tk.Canvas(container_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(container_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.create_window((0, 0), window=scrollable_frame, anchor="n")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Crea le entry per ciascun campo
    for label_text, var in input_vars.items():
        label = tk.Label(scrollable_frame, text=label_text.capitalize() + ":", font=("Arial", dashboard_font_size))
        label.pack(pady=5)
        entry = tk.Entry(scrollable_frame, width=entry_width, textvariable=var, font=("Arial", dashboard_font_size))
        entry.pack(pady=5)

    def on_confirm():
        dialog.destroy()
        dialog.details = {label: var.get() for label, var in input_vars.items()}

    confirm_button = tk.Button(dialog, text="Conferma",  font=("Arial", dashboard_font_size) , command=on_confirm)
    confirm_button.pack(pady=10)

    center_window(dialog,700,700)
    dialog.wait_window()
    return getattr(dialog, 'details', None)


def show_dashboard6(parent_frame):
    global combobox, clienti, tree
    for widget in parent_frame.winfo_children():
        widget.destroy()

    def open_font_size_popup():
        """Open a popup to choose the font size and reload the dashboard with the new size."""
        popup = tk.Toplevel()
        popup.title("Scegli la dimensione del font")
        popup.geometry("300x150")
        popup.transient()  # Make it modal

        # Label for font size selection
        label = tk.Label(popup, text="Seleziona la dimensione del font:", font=("Arial", dashboard_font_size))
        label.pack(pady=10)

        # Scale widget to select font size
        font_size_var = tk.IntVar(value=dashboard_font_size)
        # Label to display the current slider value on top of the slider handle
        value_display = ctk.CTkLabel(popup, text=str(dashboard_font_size), font=("Arial", dashboard_font_size))
        value_display.place(relx=0.5, rely=0.35, anchor="center")  # Initial position
        # CTkSlider to select font size with inverted color appearance
        font_slider = ctk.CTkSlider(
            popup,
            from_=10,
            to=30,
            number_of_steps=20,
            fg_color="white",
            progress_color= parent_frame_color,
            command=lambda value: update_slider_value(value) # Sync slider value to font_size_var
        )
        font_slider.set(dashboard_font_size)  # Set initial slider position
        font_slider.pack(pady=10)

        def update_slider_value(value):
            """Update the label text and position to follow the slider handle."""
            font_size_var.set(int(value))  # Update the IntVar with the new slider value
            value_display.configure(text=str(int(value)))  # Update the display label text
            # Position the display label above the slider handle
            slider_pos = font_slider.get()
            display_x = 20 + (slider_pos - font_slider.cget("from")) / (
                        font_slider.cget("to") - font_slider.cget("from")) * 240
            value_display.place(x=display_x, y=60)

        def apply_font_size():
            global dashboard_font_size
            dashboard_font_size = int(font_slider.get())
            popup.destroy()
            # Clear the existing dashboard and reload it with the new font size
            for widget in parent_frame.winfo_children():
                widget.destroy()  # Remove all existing widgets from parent_frame
            show_dashboard6(parent_frame)

        # Button to confirm font size selection
        apply_button = ctk.CTkButton(popup, text="Applica", font=("Arial", dashboard_font_size), command=apply_font_size)
        apply_button.pack(pady=10)

    parent_frame_color = parent_frame.cget("fg_color")

    # Crea un frame per contenere sia il titolo che la legenda
    top_frame = ctk.CTkFrame(parent_frame, corner_radius=5, fg_color=parent_frame_color)
    top_frame.pack(fill="x", pady=0)  # Si espande orizzontalmente

    # Usa il layout grid nel top_frame
    top_frame.columnconfigure(0, weight=0)  # La colonna 0 non si espande
    top_frame.columnconfigure(1, weight=1)  # La colonna 1 si espande

    # Aggiungi la legenda all'interno del top_frame, allineata a destra
    legend_label = ctk.CTkLabel(top_frame, text="Unità di misura:\n- L: liters\n- Kg: kilograms\n- Pz: pieces",
                                font=('Arial', dashboard_font_size), justify="left")
    legend_label.grid(row=0, column=0, sticky="w", padx=10)

    instruction_label = ctk.CTkLabel(top_frame, text="Scegliere Fornitore:", font=('Arial', dashboard_font_size))
    instruction_label.grid(row=0, column=1, padx=10)

    clienti = get_all_clienti_names()
    selected_cliente = tk.StringVar()

    # Frame per allineare il menù a tendina e i pulsanti "Esporta in Excel"
    search_frame = ctk.CTkFrame(parent_frame,corner_radius=5)
    search_frame.pack(pady=10)

    # Style for larger font in Combobox
    style = ttk.Style()
    style.configure("TCombobox", font=('Arial', dashboard_font_size))  # Adjust font size as needed
    style.configure("TCombobox*Listbox*Font", font=('Arial', dashboard_font_size))  # Ensure larger font size for dropdown menu items

    # Creazione del menù a tendina con ricerca incrementale
    combobox = ttk.Combobox(search_frame, textvariable=selected_cliente, values=clienti, font=('Arial', dashboard_font_size))
    combobox.configure(width=35)
    combobox.pack(side="left", padx=10)

    # Abilita la ricerca incrementale nel menù a tendina
    combobox.bind('<KeyRelease>', on_keyrelease)
    combobox.pack(pady=10)

    def on_cliente_selected(event):
        cliente = selected_cliente.get()
        show_cliente_info(cliente, tree, table_frame)

    combobox.bind("<<ComboboxSelected>>", on_cliente_selected)

    # Create a style for the Listbox within the Combobox dropdown
    combobox.option_add('*TCombobox*Listbox*Font', ('Arial', dashboard_font_size))
    '''
    def on_cliente_selected(event):
        if listbox.curselection():  # Controlla se c'è una selezione
            cliente = listbox.get(listbox.curselection())
            selected_cliente.set(cliente)  # Memorizza il cliente selezionato nell'Entry
            # Nasconde la Listbox dopo la selezione
            listbox_window.withdraw()
            show_cliente_info(cliente, tree, table_frame)

    # Funzione per aggiornare la Listbox in base alla ricerca
    def update_listbox(*args):
        search_term = entry.get().lower()  # Prende il testo inserito nella barra di ricerca
        listbox.delete(0, tk.END)  # Svuota la Listbox

        if search_term == "":  # Se la barra di ricerca è vuota, nascondi la Listbox
            listbox_window.withdraw()
            return

        # Filtra i clienti e aggiorna la Listbox
        filtered_clienti = [cliente for cliente in clienti if search_term in cliente.lower()]

        if filtered_clienti:
            # Mostra la Listbox sotto la Entry solo se ci sono risultati
            listbox_window.geometry(f"400x200+{parent_frame.winfo_x() + search_frame.winfo_x() + entry.winfo_x() + 20}+{parent_frame.winfo_y() + search_frame.winfo_y() +entry.winfo_y() + search_frame.winfo_height() + entry.winfo_height() + 40}")
            listbox_window.deiconify()

            for cliente in filtered_clienti:
                listbox.insert(tk.END, cliente)
        else:
            listbox_window.withdraw()
            # Nasconde la Listbox se non ci sono risultati

    # Creazione della barra di ricerca (Entry)
    entry = tk.Entry(search_frame, textvariable=selected_cliente, font=('Arial', 16))
    entry.pack(side="left", padx=10,pady=10)
    entry.config(width=33)
    entry.bind('<KeyRelease>', update_listbox)  # Ogni volta che si digita, aggiorna la Listbox

    listbox_window = tk.Toplevel(search_frame)
    listbox_window.wm_overrideredirect(True)  # Rimuove bordi e titoli della finestra
    listbox_window.geometry("400x200")
    listbox_window.withdraw()

    # Creazione della Listbox per visualizzare i risultati (inizialmente nascosta)
    listbox = tk.Listbox(listbox_window, font=('Arial', 16), height=8, selectmode=tk.SINGLE)
    listbox.pack(side="left", fill="both", expand=True)


    # Quando si seleziona un cliente dalla Listbox
    listbox.bind('<<ListboxSelect>>', on_cliente_selected)
    '''

    # Pulsante "Esporta Excel Fornitori" accanto al pulsante prodotti
    export_fornitori_button = ctk.CTkButton(search_frame, text="Esporta Excel Fornitori", command=open_clienti_selection_popup, font=('Arial', dashboard_font_size))
    export_fornitori_button.pack(side="left", padx=10)

    font_size_button = ctk.CTkButton(search_frame, text="Cambia Dimensione Font", font=("Arial", dashboard_font_size),
                                     command=open_font_size_popup, corner_radius=5)
    font_size_button.pack(side="left", padx=10)

    global table_frame
    table_frame = ctk.CTkFrame(parent_frame, corner_radius=5)
    table_frame.pack_forget()

    table_title = ctk.CTkLabel(table_frame, text="Dati Fornitore:", font=('Arial', dashboard_font_size, 'bold'))
    table_title.pack(pady=10)

    # Create a frame to hold the table and scrollbars
    table_container = tk.Frame(table_frame, width=1600, height=400)
    table_container.pack_propagate(False)  # Prevent the frame from resizing to fit its children
    table_container.pack(pady=10, fill="both", expand=True)

    # Use only one column for the Treeview to display data vertically
    global columns
    columns = ("",)

    global tree
    tree = ttk.Treeview(table_container, columns=columns, show="headings")
    tree.table_container = table_container # Save the frame in the treeview for reference
    setup_context_menu(tree)
    tree.bind("<B1-Motion>", on_mouse_drag)


    style = ttk.Style()
    style.configure("Treeview",
                    rowheight=30,
                    font=('Arial', dashboard_font_size),
                    background="#f1f8e9",
                    foreground="#004d40",
                    fieldbackground="#f1f8e9",
                    bordercolor="#000000",
                    relief="solid",
                    borderwidth=1)
    style.configure("Treeview.Heading",
                    font=('Arial', dashboard_font_size, 'bold'),
                    background="#a5d6a7",
                    foreground="#004d40")
    style.map("Treeview",
              background=[('selected', '#c8e6c9')],
              foreground=[('selected', '#004d40')])

    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=1500, anchor="w")  # Set a larger width to accommodate longer text

    yscrollbar = ttk.Scrollbar(table_container, orient=tk.VERTICAL, command=tree.yview)
    yscrollbar.pack(side="right", fill="y")
    tree.configure(yscroll=yscrollbar.set)

    tree.pack(side="left", fill="both", expand=True)
    tree.bind("<Double-1>", lambda event: on_double_click(event, tree))
    def fetch_orders():
        conn = sqlite3.connect('Database_Utilities/Database/MergedDatabase.db')
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM storico_ordini")
        orders = cursor.fetchall()
        conn.close()
        return orders

    orders = fetch_orders()

    def update_treeview(order_list):
        for row in tree.get_children():
            tree.delete(row)
        for order in order_list:
            tree.insert("", tk.END, values=order)

    update_treeview(orders)

    def fetch_invoice_data(invoice_number):
        conn = sqlite3.connect('resources/orders_fattura_1.db')
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM orders_fattura WHERE id = '{invoice_number}'")
        column_names = [description[0] for description in cursor.description]
        invoice_tuple = cursor.fetchone()
        conn.close()
        if invoice_tuple:
            invoice = dict(zip(column_names, invoice_tuple))
            return invoice
        return None

    def fetch_invoice_products(invoice_number):
        conn = sqlite3.connect('resources/orders_fattura_1.db')
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM orders_fattura WHERE id = '{invoice_number}'")
        column_names = [description[0] for description in cursor.description]
        product_rows = cursor.fetchall()
        conn.close()
        products = [dict(zip(column_names, product_row)) for product_row in product_rows]
        return products

    def print_return_document():
        selected_item = tree.selection()
        if selected_item:
            order = tree.item(selected_item, "values")
            generate_return_pdf(order)
        else:
            messagebox.showwarning("Seleziona Ordine",
                                   "Per favore, seleziona un ordine per stampare la bolla di reso.")


    class PDFReturn(FPDF):
        def header(self):

            # Aggiungi il logo in alto a sinistra
            self.image('C:/PROGETTO/myapp/resources/LogoCINCOTTI.jpg', 8, 2, 77, 44,
                       'JPG')  # Modifica il percorso e le dimensioni

            # Aggiungi l'immagine in alto a destra
            self.image('C:/PROGETTO/myapp/resources/Logojpeg.jpg', 120, 8, 77, 33,
                       'JPG')  # Modifica il percorso e le dimensioni

            self.ln(35)

        def footer(self):
            # Aggiungi l'immagine in basso al centro
            self.set_y(-40)  # Posiziona a 30 mm dal fondo della pagina
            self.image('C:/PROGETTO/myapp/resources/LOGOMOZZABELLAECINCOTTI.jpg', 75, self.get_y(), 70, 40,
                       'JPG')  # Modifica il percorso e le dimensioni

        def draw_table(self, data, col_widths, headers=None, bold_headers=False):
            if headers:
                self.set_font('Arial', 'B', 10) if bold_headers else self.set_font('Arial', '', 10)
                for header, width in zip(headers, col_widths):
                    self.cell(width, 8, header, 1, 0, 'C')
                self.ln()

            self.set_font('Arial', '', 10)

            for row in data:
                row_heights = [self.calculate_cell_height(datum, width) for datum, width in zip(row, col_widths)]
                max_row_height = max(row_heights)

                x_start = self.get_x()
                y_start = self.get_y()
                for datum, width in zip(row, col_widths):
                    self.multi_cell(width, 6, datum if datum.strip() != "" else "-", 1, 'L', False)
                    self.set_xy(x_start + width, y_start)
                    x_start += width
                self.ln(max_row_height)  # Move down to start the next row

        def calculate_cell_height(self, text, width):
            """ Calculate the necessary cell height for the text based on the column width. """
            self.set_font('Arial', '', 10)
            num_lines = max(1, self.get_string_width(text) / width)  # Ensure at least one line
            return 6 * (text.count('\n') + num_lines)  # Line height times the number of lines

        def invoice_header(self, invoice):
            self.set_font('Arial', '', 10)
            data = [
                f"Document: Bolla di reso", f"n. {invoice['number']}", f"Data: {invoice['Date_1']}"
            ]
            col_widths = [70, 40, 70]  # Regola le larghezze delle colonne come necessario
            self.draw_table([data], col_widths)

        def customer_details(self, invoice):
            data = [
                ["Goods destination:\n" + invoice["customer_phone"], "Customer:\n" + invoice["customer_address"]],
            ]
            col_widths = [70, 110]  # Regola le larghezze delle colonne come necessario
            self.draw_table(data, col_widths, bold_headers=True)

        def product_table(self, products):
            headers = ["Article", "Description", "Quantity", "Price", "Discount", "Amount"]
            data = [
                [product['article'], product['product_description'], str(product['product_quantity']),
                 str(product['product_price']), str(product['product_discount']), str(product['product_amount'])]
                for product in products
            ]
            col_widths = [20, 70, 20, 20, 20, 30]
            self.set_font('Arial', 'B', 10)
            self.cell(0, 8, 'Product Details', 0, 1, 'L')
            self.draw_table(data, col_widths, headers, bold_headers=True)

        def total_details(self, invoice):
            data = [
                ["Total Quantity:\n" + str(invoice["total_quantity"]), "IVA:\n" + " ", "Mittente:\n" + invoice["sender_name"]],
                ["Causale trasporto:\n" + "Reso", "Trasporto a mezzo:\n" + "Vettore ", "Destinatario:\n" + invoice["recipient"]],
                ["Imponibile:\n" + str(invoice["total_amount"]), "Importo IVA:\n" + str(invoice["iva_amount"]),
                 "Totale Fattura EURO:\n" + str(invoice["total_invoice_euro"])],
            ]
            col_widths = [60, 60, 60]
            self.set_font('Arial', 'B', 10)
            self.cell(0, 8, 'Total Details', 0, 1, 'L')
            self.draw_table(data, col_widths, bold_headers=True)

    def generate_return_pdf(order):
        invoice_number = order[0]
        invoice = fetch_invoice_data(invoice_number)
        products = fetch_invoice_products(invoice_number)
        if not invoice:
            messagebox.showerror("Errore", "Dati della bolla di reso non trovati.")
            return

        # Popup window to edit and confirm invoice data
        def edit_and_confirm():
            def save_changes():
                # Update invoice data from the entry fields
                invoice["number"] = entry_number.get()
                invoice["Date_1"] = entry_date.get()
                invoice["customer_phone"] = entry_customer_phone.get()
                invoice["customer_address"] = entry_customer_address.get()
                invoice["recipient"] = entry_recipient.get()
                invoice["total_quantity"] = entry_total_quantity.get()
                invoice["total_amount"] = entry_total_amount.get()
                invoice["total_invoice_euro"] = entry_total_invoice_euro.get()

                # Close the popup window after saving changes
                popup.destroy()

                # Open a separate window to choose save path
                choose_save_path()

            # Ottiene la data corrente e la formatta come YYYYMMDD
            current_date = datetime.now().strftime("%Y%m%d")
            filename = f"Bolla_Reso_{current_date}"

            def choose_save_path():
                save_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")], initialfile=filename)
                if save_path:
                    pdf = PDFReturn()
                    pdf.add_page()
                    pdf.invoice_header(invoice)
                    pdf.customer_details(invoice)
                    pdf.product_table(products)
                    pdf.total_details(invoice)
                    pdf.output(save_path)
                    messagebox.showinfo("Successo", f"Fattura salvata come {os.path.basename(save_path)}")
                else:
                    messagebox.showwarning("Salvataggio Annullato", "Il salvataggio della fattura è stato annullato.")

            # Create the popup window
            popup = tk.Toplevel()
            popup.title("Modifica i dettagli della fattura")
            popup.geometry("800x600")  # Set the size of the popup window
            popup.resizable(True, True)  # Prevent resizing for consistent layout

            # Add padding and spacing between elements
            padding = {'padx': 10, 'pady': 5}
            # Set the width of entry fields and the font size
            entry_width = 40
            font_size = ("Arial", 14)  # Font family Arial, size 14

            # Create and grid labels and entries for each invoice field with larger font
            tk.Label(popup, text="Numero:", font=font_size).grid(row=0, column=0, **padding)
            entry_number = tk.Entry(popup, width=entry_width, font=font_size)
            entry_number.grid(row=0, column=1, **padding)
            entry_number.insert(0, invoice["number"])

            tk.Label(popup, text="Data:", font=font_size).grid(row=1, column=0, **padding)
            entry_date = tk.Entry(popup, width=entry_width, font=font_size)
            entry_date.grid(row=1, column=1, **padding)
            entry_date.insert(0, invoice["Date_1"])

            tk.Label(popup, text="Telefono Cliente:", font=font_size).grid(row=2, column=0, **padding)
            entry_customer_phone = tk.Entry(popup, width=entry_width, font=font_size)
            entry_customer_phone.grid(row=2, column=1, **padding)
            entry_customer_phone.insert(0, invoice["customer_phone"])

            tk.Label(popup, text="Indirizzo Cliente:", font=font_size).grid(row=3, column=0, **padding)
            entry_customer_address = tk.Entry(popup, width=entry_width, font=font_size)
            entry_customer_address.grid(row=3, column=1, **padding)
            entry_customer_address.insert(0, invoice["customer_address"])

            tk.Label(popup, text="Mittente:", font=font_size).grid(row=12, column=0, **padding)
            entry_sender_name = tk.Entry(popup, width=entry_width, font=font_size)
            entry_sender_name.grid(row=12, column=1, **padding)
            entry_sender_name.insert(0, invoice["sender_name"])

            tk.Label(popup, text="Destinatario:", font=font_size).grid(row=15, column=0, **padding)
            entry_recipient = tk.Entry(popup, width=entry_width, font=font_size)
            entry_recipient.grid(row=15, column=1, **padding)
            entry_recipient.insert(0, invoice["recipient"])

            tk.Label(popup, text="Quantità Totale:", font=font_size).grid(row=16, column=0, **padding)
            entry_total_quantity = tk.Entry(popup, width=entry_width, font=font_size)
            entry_total_quantity.grid(row=16, column=1, **padding)
            entry_total_quantity.insert(0, invoice["total_quantity"])

            tk.Label(popup, text="Importo IVA:", font=font_size).grid(row=17, column=0, **padding)
            entry_iva_amount = tk.Entry(popup, width=entry_width, font=font_size)
            entry_iva_amount.grid(row=17, column=1, **padding)
            entry_iva_amount.insert(0, invoice["iva_amount"])

            tk.Label(popup, text="Imponibile:", font=font_size).grid(row=18, column=0, **padding)
            entry_total_amount = tk.Entry(popup, width=entry_width, font=font_size)
            entry_total_amount.grid(row=18, column=1, **padding)
            entry_total_amount.insert(0, invoice["total_amount"])

            tk.Label(popup, text="Totale Fattura EURO:", font=font_size).grid(row=19, column=0, **padding)
            entry_total_invoice_euro = tk.Entry(popup, width=entry_width, font=font_size)
            entry_total_invoice_euro.grid(row=19, column=1, **padding)
            entry_total_invoice_euro.insert(0, invoice["total_invoice_euro"])

            # Add save changes button at the bottom
            tk.Button(popup, text="Salva modifiche e conferma", command=save_changes, font=font_size).grid(row=22, columnspan=2,
                                                                                           pady=30)

        edit_and_confirm()
        # Continue to generate the PDF after the popup is closed
        pdf = PDFReturn()
        pdf.add_page()
        pdf.invoice_header(invoice)
        pdf.customer_details(invoice)
        pdf.product_table(products)
        pdf.total_details(invoice)

    button_return = ctk.CTkButton(search_frame, text="Stampa Bolla di reso", command=print_return_document, font=('Arial', dashboard_font_size))
    button_return.pack(side="left", padx=10)

# Other components of your application remain unchanged
