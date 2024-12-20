import os
from tkinter.filedialog import asksaveasfilename

import customtkinter as ctk
from tkinter import ttk, Menu
import tkinter as tk
from Database_Utilities.crud_vettori import get_all_vettori_names, get_vettore_info_by_name, update_record_vettori, \
    delete_record_vettori
import openpyxl
from tkinter import messagebox

from dashboards.popup_functions import open_add_popup

dashboard_font_size = 14  # Default font size
def center_window(window, width, height):
    window.update_idletasks()
    width = width
    height = height
    x = (window.winfo_screenwidth() // 2) - (width // 2)
    y = (window.winfo_screenheight() // 2) - (height // 2)
    window.geometry(f'{width}x{height}+{x}+{y}')

def export_vettori_to_excel(selected_vettori):
    if not selected_vettori:
        messagebox.showwarning("Nessun Dato", "Non ci sono vettori selezionati da esportare.")
        return

    if len(selected_vettori) == 1:
        default_filename = f"Info_Vettore_{selected_vettori[0]}.xlsx"
    else:
        default_filename = f"Info_Vettori.xlsx"
    file_path = asksaveasfilename(defaultextension=".xlsx",
                                  filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                  initialfile=default_filename)
    if not file_path:
        messagebox.showwarning("Salvataggio Annullato", "Salvataggio del file Excel annullato.")
        return

    workbook = openpyxl.Workbook()

    for vettore in selected_vettori:
        info = get_vettore_info_by_name(vettore)  # Assume this function can be reused or similarly named for vettori

        # Create a new sheet for each vettore
        sheet = workbook.create_sheet(title=vettore)

        headers = ["Ragione sociale", "Indirizzo", "CAP", "Città", "Nazione", "Partita IVA", "Telefono", "Email",
                   "Zona", "Giorni chiusura", "Orari di scarico", "Condizioni pagamento", "Sconto", "Agente"]
        sheet.append(headers)

        if info:
            sheet.append([info.get(header, "N/A") for header in headers])
        else:
            sheet.append(["Nessun dato disponibile"])

    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook["Sheet"]

    workbook.save(file_path)
    messagebox.showinfo("Excel Generato", f"Il file Excel è stato salvato come {os.path.basename(file_path)}")

def open_vettori_selection_popup():
    popup = tk.Toplevel()
    popup.title("Seleziona Vettori da Esportare")
    popup.geometry("700x700")

    instruction_label = ctk.CTkLabel(popup, text="Seleziona i vettori che vuoi esportare:", font=('Arial', 14), text_color="black")
    instruction_label.pack(pady=10)

    container_frame = tk.Frame(popup)
    container_frame.pack(fill="both", expand=True, padx=60)

    canvas = tk.Canvas(container_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(container_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.create_window((0, 0), window=scrollable_frame, anchor="n")
    canvas.configure(yscrollcommand=scrollbar.set)

    selected_vettori = []

    def on_checkbutton_toggle(vettore, var):
        if var.get():
            if vettore not in selected_vettori:
                selected_vettori.append(vettore)
        else:
            if vettore in selected_vettori:
                selected_vettori.remove(vettore)

    def toggle_select_all():
        all_selected = all(var.get() for var in selections.values())
        if all_selected:
            for var in selections.values():
                var.set(False)
            selected_vettori.clear()
        else:
            selected_vettori.clear()
            for vettore in vettori:
                var = selections[vettore]
                var.set(True)
                selected_vettori.append(vettore)

    select_all_button = ctk.CTkButton(popup, text="Seleziona Tutto", command=toggle_select_all, width=120, height=30)
    select_all_button.pack(pady=5)

    selections = {}
    vettori = get_all_vettori_names()  # Update or rename this function as necessary
    for vettore in vettori:
        var = tk.BooleanVar()
        selections[vettore] = var
        check = tk.Checkbutton(scrollable_frame, text=vettore, variable=var, font=('Arial', dashboard_font_size),
                               command=lambda c=vettore, v=var: on_checkbutton_toggle(c, v))
        check.pack(anchor="w")

    confirm_button = ctk.CTkButton(popup, text="Conferma",font=('Arial', dashboard_font_size), command=lambda: [popup.destroy(), export_vettori_to_excel(selected_vettori)], width=120, height=30)
    confirm_button.pack(pady=10)
def get_vettore_info(vettore):
    # Presumiamo che `get_vettore_info_by_name` sia una funzione simile a quella dei clienti
    info = get_vettore_info_by_name(vettore)
    return info


def show_vettore_info(vettore, tree, table_frame):
    # Pulisce le righe esistenti nel treeview
    for row in tree.get_children():
        tree.delete(row)
    # Recupera le informazioni del vettore selezionato
    info = get_vettore_info(vettore)
    if info:
        # Inserisce le informazioni nel treeview
        for key, value in info.items():
            tree.insert("", tk.END, values=[f"{key}: {value}"])
    # Salva il vettore corrente nel tree per un utilizzo futuro
    tree.vettore = vettore
    # Mostra il frame della tabella
    table_frame.pack(pady=10, fill="both", expand=True)
def update_combobox():
    value = combobox.get().lower()
    if value == '':
        combobox['values'] = vettori
    else:
        data = [item for item in vettori if value in item.lower()]
        combobox['values'] = data
        combobox.event_generate('<Down>')

def on_keyrelease(event):
    combobox.after(2000, update_combobox)

def copy_selection(tree):
    selected_items = tree.selection()  # Ottiene tutti gli elementi selezionati
    copied_data = []
    for item in selected_items:
        item_values = tree.item(item, "values")
        copied_data.append("\t".join(map(str, item_values)))  # Unisce i valori di ciascuna riga con una tabulazione
    clipboard_text = "\n".join(copied_data)  # Unisce tutte le righe con un ritorno a capo
    tree.clipboard_clear()  # Pulisce gli appunti
    tree.clipboard_append(clipboard_text)  # Aggiunge il testo agli appunti
    messagebox.showinfo("Copiato", "I testi selezionati sono stati copiati negli appunti.")


def setup_context_menu(tree):
    # Creare un menu contestuale
    context_menu = Menu(tree, tearoff=0)
    context_menu.add_command(label="Copia", command=lambda: copy_selection(tree), font=('Arial', dashboard_font_size))

    def on_right_click(event):
        # Mostrare il menu contestuale
        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    tree.bind("<Button-3>", on_right_click)  # <Button-3> è il clic del pulsante destro del mouse

# Aggiungere il supporto per la selezione trascinando
def on_mouse_drag(event):
    # Identifica l'item su cui si trova il cursore
    item = tree.identify_row(event.y)
    if item:
        tree.selection_add(item)


def on_double_click(event, tree):
    # Get the current cliente saved in the tree
    vettore = tree.vettore
    print(vettore)
    if vettore:
        # Retrieve the customer's full info using the cliente name
        cliente_info = get_vettore_info(vettore)
        print(cliente_info)

        if cliente_info:
            # Show the action dialog with the vettore's info
            #qui passare solo campi vettore
            show_action_dialog(vettore, lambda action: handle_action(
                action, tree,
                cliente_info.get("id", ""),
                cliente_info.get("nome", ""),
                cliente_info.get("indirizzo", "")
            ))
        else:
            messagebox.showwarning("Attenzione", "Nessun dato disponibile per il vettore selezionato.")
    else:
        messagebox.showwarning("Attenzione", "Nessun vettore selezionato.")
button_font = ("Arial", dashboard_font_size)  # Font più grande per i pulsanti
button_width = 15  # Larghezza maggiore per i pulsanti
button_height = 2  # Altezza maggiore per i pulsanti
def show_action_dialog(ragione_sociale, callback):
    dialog = tk.Toplevel()
    dialog.title("Scegli Azione")
    dialog.grab_set()  # Ottieni il focus sulla finestra di dialogo
    dialog.transient()  # Rendi la finestra di dialogo modale

    label = tk.Label(dialog, text=f"Vuoi modificare o eliminare '{ragione_sociale}'?",
                         font=("Arial", dashboard_font_size))
    label.pack(pady=10)

    button_frame = tk.Frame(dialog)
    button_frame.pack(pady=10)

    def on_modify():
        dialog.destroy()
        callback("modify")

    def on_delete():
        dialog.destroy()
        callback("delete")

    modify_button = tk.Button(button_frame, text="Modifica", font=button_font, width=button_width,
                              height=button_height, command=on_modify)
    modify_button.grid(row=0, column=0, padx=10, pady=10)

    delete_button = tk.Button(button_frame, text="Elimina", font=button_font, width=button_width, height=button_height,
                              command=on_delete)
    delete_button.grid(row=0, column=1, padx=10, pady=10)

    center_window(dialog,800,400)
    dialog.wait_window()  # Attendi la chiusura della finestra di dialogo

def handle_action(action, tree, id, nome, indirizzo):
    if action == "modify":
        details = ask_details(id, f"Inserisci le nuove info di '{nome}':", nome, indirizzo)
        if details is not None:
            update_record_vettori(id, nome, indirizzo)
            show_vettore_info(tree, tree.table_frame)
            messagebox.showinfo("Modifica Prodotto", f"Hai modificato '{nome}'.")
    elif action == "delete":
        confirm = messagebox.askokcancel("Conferma Eliminazione", f"Sei sicuro di voler eliminare '{nome}'?")
        if confirm:
            delete_record_vettori(id)
            show_vettore_info(tree.table_frame)
            messagebox.showinfo("Eliminazione Prodotto", f"Il vettore '{nome}' è stato eliminato.")

def ask_details(id, prompt, nome, indirizzo):
    dialog = tk.Toplevel()
    dialog.title("Scegli le nuove info del vettore")
    dialog.grab_set()
    dialog.transient()

    label = tk.Label(dialog, text=prompt, font=("Arial", dashboard_font_size))
    label.pack(pady=10)

    entry_width = 40

    input_vars = {

        "id": tk.StringVar(value=id),
        "nome": tk.StringVar(value=nome),
        "indirizzo": tk.StringVar(value=indirizzo)

    }

    container_frame = tk.Frame(dialog)
    container_frame.pack(fill="both", expand=True, padx=60)

    canvas = tk.Canvas(container_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(container_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.create_window((0, 0), window=scrollable_frame, anchor="n")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Crea le entry per ciascun campo
    for label_text, var in input_vars.items():
        label = tk.Label(scrollable_frame, text=label_text.capitalize() + ":", font=("Arial", dashboard_font_size))
        label.pack(pady=5)
        entry = tk.Entry(scrollable_frame, width=entry_width, textvariable=var, font=("Arial", dashboard_font_size))
        entry.pack(pady=5)

    def on_confirm():
        dialog.destroy()
        dialog.details = {label: var.get() for label, var in input_vars.items()}

    confirm_button = tk.Button(dialog, text="Conferma",  font=("Arial", dashboard_font_size) , command=on_confirm)
    confirm_button.pack(pady=10)

    center_window(dialog,700,700)
    dialog.wait_window()
    return getattr(dialog, 'details', None)


def show_dashboard5(parent_frame):
    global combobox, vettori, tree
    for widget in parent_frame.winfo_children():
        widget.destroy()

    def open_font_size_popup():
        """Open a popup to choose the font size and reload the dashboard with the new size."""
        popup = tk.Toplevel()
        popup.title("Scegli la dimensione del font")
        popup.geometry("500x250")
        popup.transient()  # Make it modal

        # Label for font size selection
        label = tk.Label(popup, text="Seleziona la dimensione del font:", font=("Arial", dashboard_font_size))
        label.pack(pady=10)

        # Scale widget to select font size
        font_size_var = tk.IntVar(value=dashboard_font_size)
        # Label to display the current slider value on top of the slider handle
        value_display = ctk.CTkLabel(popup, text=str(dashboard_font_size), font=("Arial", dashboard_font_size))
        value_display.place(relx=0.5, rely=0.35, anchor="center")  # Initial position
        # CTkSlider to select font size with inverted color appearance
        slider_min = 10
        slider_max = 30
        font_slider = ctk.CTkSlider(
            popup,
            from_=10,
            to=30,
            number_of_steps=20,
            fg_color="white",
            progress_color= parent_frame_color,
            command=lambda value: update_slider_value(value) # Sync slider value to font_size_var
        )
        font_slider.set(dashboard_font_size)  # Set initial slider position
        font_slider.pack(pady=10)

        def update_slider_value(value):
            """Update the label text and position to follow the slider handle."""
            font_size_var.set(int(value))  # Update the IntVar with the new slider value
            value_display.configure(text=str(int(value)))  # Update the display label text
            # Position the display label above the slider handle
            slider_pos = font_slider.get()
            display_x = 20 + (slider_pos - slider_min) / (slider_max - slider_min) * 240
            value_display.place(x=display_x, y=60)

        def apply_font_size():
            global dashboard_font_size
            dashboard_font_size = int(font_slider.get())
            popup.destroy()
            # Clear the existing dashboard and reload it with the new font size
            for widget in parent_frame.winfo_children():
                widget.destroy()  # Remove all existing widgets from parent_frame
            show_dashboard5(parent_frame)

        # Button to confirm font size selection
        apply_button = ctk.CTkButton(popup, text="Applica", font=("Arial", dashboard_font_size), command=apply_font_size)
        apply_button.pack(pady=10)

    parent_frame_color = parent_frame.cget("fg_color")

    # Crea un frame per contenere sia il titolo che la legenda
    top_frame = ctk.CTkFrame(parent_frame, corner_radius=5, fg_color=parent_frame_color)
    top_frame.pack(fill="x", pady=0)  # Si espande orizzontalmente

    # Usa il layout grid nel top_frame
    top_frame.columnconfigure(0, weight=0)  # La colonna 0 non si espande
    top_frame.columnconfigure(1, weight=1)  # La colonna 1 si espande

    instruction_label = ctk.CTkLabel(top_frame, text="Scegliere Vettore:", font=('Arial', dashboard_font_size))
    instruction_label.grid(row=0, column=1, padx=10)

    vettori = get_all_vettori_names()
    selected_vettore = tk.StringVar()

    # Frame per allineare il menù a tendina e i pulsanti "Esporta in Excel"
    search_frame = ctk.CTkFrame(parent_frame,corner_radius=5)
    search_frame.pack(pady=10)

    # Style for larger font in Combobox
    style = ttk.Style()
    style.configure("TCombobox", font=('Arial', dashboard_font_size))  # Adjust font size as needed
    style.configure("TCombobox*Listbox*Font", font=('Arial', dashboard_font_size))  # Ensure larger font size for dropdown menu items

    # Creazione del menù a tendina con ricerca incrementale
    combobox = ttk.Combobox(search_frame, textvariable=selected_vettore, values=vettori, font=('Arial', dashboard_font_size))
    combobox.configure(width=35)
    combobox.pack(side="left", padx=10)

    # Abilita la ricerca incrementale nel menù a tendina
    combobox.bind('<KeyRelease>', on_keyrelease)
    combobox.pack(pady=10)

    def on_vettore_selected(event):
        vettore = selected_vettore.get()
        show_vettore_info(vettore, tree, table_frame)

    combobox.bind("<<ComboboxSelected>>", on_vettore_selected)

    # Create a style for the Listbox within the Combobox dropdown
    combobox.option_add('*TCombobox*Listbox*Font', ('Arial', dashboard_font_size))
    '''
    def on_vettore_selected(event):
        if listbox.curselection():  # Controlla se c'è una selezione
            vettore = listbox.get(listbox.curselection())
            selected_vettore.set(vettore)  # Memorizza il cliente selezionato nell'Entry
            # Nasconde la Listbox dopo la selezione
            listbox_window.withdraw()
            show_vettore_info(vettore, tree, table_frame)

    # Funzione per aggiornare la Listbox in base alla ricerca
    def update_listbox(*args):
        search_term = entry.get().lower()  # Prende il testo inserito nella barra di ricerca
        listbox.delete(0, tk.END)  # Svuota la Listbox

        if search_term == "":  # Se la barra di ricerca è vuota, nascondi la Listbox
            listbox_window.withdraw()
            return

        # Filtra i clienti e aggiorna la Listbox
        filtered_vettori = [vettore for vettore in vettori if search_term in vettore.lower()]

        if filtered_vettori:
            # Mostra la Listbox sotto la Entry solo se ci sono risultati
            listbox_window.geometry(f"300x200+{parent_frame.winfo_x() + search_frame.winfo_x() + entry.winfo_x() + 20}+{parent_frame.winfo_y() + search_frame.winfo_y() +entry.winfo_y() + search_frame.winfo_height() + entry.winfo_height() + 40}")
            listbox_window.deiconify()

            for vettore in filtered_vettori:
                listbox.insert(tk.END, vettore)
        else:
            listbox_window.withdraw()
            #listbox.place_forget()
            # Nasconde la Listbox se non ci sono risultati

    # Creazione della barra di ricerca (Entry)
    entry = tk.Entry(search_frame, textvariable=selected_vettore, font=('Arial', 16))
    entry.pack(side="left", padx=10,pady=10)
    entry.config(width=25)
    entry.bind('<KeyRelease>', update_listbox)  # Ogni volta che si digita, aggiorna la Listbox

    listbox_window = tk.Toplevel(search_frame)
    listbox_window.wm_overrideredirect(True)  # Rimuove bordi e titoli della finestra
    listbox_window.geometry("300x200")
    listbox_window.withdraw()

    # Creazione della Listbox per visualizzare i risultati (inizialmente nascosta)
    listbox = tk.Listbox(listbox_window, font=('Arial', 16), height=8, selectmode=tk.SINGLE)
    listbox.pack(side="left", fill="both", expand=True)


    # Quando si seleziona un cliente dalla Listbox
    listbox.bind('<<ListboxSelect>>', on_vettore_selected)
    '''

    # Pulsante "Esporta Excel Fornitori" accanto al pulsante prodotti
    export_vettori_button = ctk.CTkButton(search_frame, text="Esporta Excel Vettori", command =lambda: open_vettori_selection_popup(), font=('Arial', dashboard_font_size))
    export_vettori_button.pack(side="left", padx=10)

    add_cliente_button = ctk.CTkButton(search_frame, text="Aggiungi Vettore", command=lambda: open_add_popup("Vettore"),
                                       font=("Arial", dashboard_font_size))
    add_cliente_button.pack(side="left", padx=10)

    font_size_button = ctk.CTkButton(search_frame, text="Cambia Dimensione Font", font=("Arial", dashboard_font_size),
                                     command=open_font_size_popup, corner_radius=5)
    font_size_button.pack(side="left", padx=10)



    global table_frame
    table_frame = ctk.CTkFrame(parent_frame, corner_radius=5)
    table_frame.pack_forget()

    table_title = ctk.CTkLabel(table_frame, text="Dati Vettori:", font=('Arial', dashboard_font_size, 'bold'))
    table_title.pack(pady=10)

    # Create a frame to hold the table and scrollbars
    table_container = tk.Frame(table_frame, width=1600, height=400)
    table_container.pack_propagate(False)  # Prevent the frame from resizing to fit its children
    table_container.pack(pady=10, fill="both", expand=True)

    # Use only one column for the Treeview to display data vertically
    global columns
    columns = ("",)

    global tree
    tree = ttk.Treeview(table_container, columns=columns, show="headings")
    tree.table_container = table_container # Save the frame in the treeview for reference
    setup_context_menu(tree)
    tree.bind("<B1-Motion>", on_mouse_drag)


    style = ttk.Style()
    style.configure("Treeview",
                    rowheight=30,
                    font=('Arial', dashboard_font_size),
                    background="#f1f8e9",
                    foreground="#004d40",
                    fieldbackground="#f1f8e9",
                    bordercolor="#000000",
                    relief="solid",
                    borderwidth=1)
    style.configure("Treeview.Heading",
                    font=('Arial', dashboard_font_size, 'bold'),
                    background="#a5d6a7",
                    foreground="#004d40")
    style.map("Treeview",
              background=[('selected', '#c8e6c9')],
              foreground=[('selected', '#004d40')])

    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=1500, anchor="w")  # Set a larger width to accommodate longer text

    yscrollbar = ttk.Scrollbar(table_container, orient=tk.VERTICAL, command=tree.yview)
    yscrollbar.pack(side="right", fill="y")
    tree.configure(yscroll=yscrollbar.set)

    tree.pack(side="left", fill="both", expand=True)
    tree.bind("<Double-1>", lambda event: on_double_click(event, tree))

# Other components of your application remain unchanged
