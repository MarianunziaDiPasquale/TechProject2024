import os
from tkinter.filedialog import asksaveasfilename

import customtkinter as ctk
from tkinter import ttk, Menu
import tkinter as tk
from dashboards.Database_Utilities.crud_clienti import get_all_clienti_names, get_cliente_info_by_name
import openpyxl
from tkinter import messagebox
def center_window(window, width, height):
    window.update_idletasks()
    width = width
    height = height
    x = (window.winfo_screenwidth() // 2) - (width // 2)
    y = (window.winfo_screenheight() // 2) - (height // 2)
    window.geometry(f'{width}x{height}+{x}+{y}')
def export_clienti_to_excel(selected_clienti):
    if not selected_clienti:
        messagebox.showwarning("Nessun Dato", "Non ci sono clienti selezionati da esportare.")
        return

    if len(selected_clienti) == 1:
        default_filename = f"Info_Cliente_{selected_clienti[0]}.xlsx"
    else:
        default_filename = f"Info_Clienti.xlsx"
    file_path = asksaveasfilename(defaultextension=".xlsx",
                                  filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                  initialfile=default_filename)
    # Verifica se l'utente ha scelto un percorso
    if not file_path:
        messagebox.showwarning("Salvataggio Annullato", "Salvataggio del file Excel annullato.")
        return

    workbook = openpyxl.Workbook()

    for cliente in selected_clienti:
        info = get_cliente_info_by_name(cliente)

        # Crea un nuovo foglio per ogni cliente
        sheet = workbook.create_sheet(title=cliente)

        headers = ["Ragione sociale", "Indirizzo", "CAP", "Città", "Nazione", "Partita IVA", "Telefono", "Email", "Zona", "Giorni chiusura", "Orari di scarico", "Condizioni pagamento", "Sconto", "Agente"]
        sheet.append(headers)

        # Aggiungi i dati del cliente se esistono, altrimenti lascia solo le intestazioni
        if info:
            sheet.append([info.get(header, "N/A") for header in headers])
        else:
            # Aggiungi una riga vuota se non ci sono dati, lasciando solo l'intestazione
            sheet.append(["Nessun dato disponibile"])

    # Rimuove il foglio predefinito solo se ci sono altri fogli visibili
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook["Sheet"]

    workbook.save(file_path)
    messagebox.showinfo("Excel Generato", f"Il file Excel è stato salvato come {os.path.basename(file_path)}")

def export_liste_prsn_to_excel(selected_lista):
    if not selected_lista:
        messagebox.showwarning("Nessun Dato", "Non ci sono clienti selezionati da esportare.")
        return

    if len(selected_lista) == 1:
        default_filename = f"Info_Cliente_{selected_lista[0]}.xlsx"
    file_path = asksaveasfilename(defaultextension=".xlsx",
                                  filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                  initialfile=default_filename)
    # Verifica se l'utente ha scelto un percorso
    if not file_path:
        messagebox.showwarning("Salvataggio Annullato", "Salvataggio del file Excel annullato.")
        return

    workbook = openpyxl.Workbook()

    for cliente in selected_lista:

        # modificare con funzione per recuperare info lista
        info = get_cliente_info_by_name(selected_lista)

        # Crea un nuovo foglio per ogni cliente
        sheet = workbook.create_sheet(title=cliente)

        # modificare con colonne lista personalizzata
        headers = ["Ragione sociale", "Indirizzo", "CAP", "Città"]
        sheet.append(headers)

        # Aggiungi i dati del cliente se esistono, altrimenti lascia solo le intestazioni
        if info:
            sheet.append([info.get(header, "N/A") for header in headers])
        else:
            # Aggiungi una riga vuota se non ci sono dati, lasciando solo l'intestazione
            sheet.append(["Nessun dato disponibile"])

    # Rimuove il foglio predefinito solo se ci sono altri fogli visibili
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook["Sheet"]

    workbook.save(file_path)
    messagebox.showinfo("Excel Generato", f"Il file Excel è stato salvato come {os.path.basename(file_path)}")

def open_lista_selection_popup():
    popup = tk.Toplevel()
    popup.title("Seleziona la lista ")
    popup.geometry("700x700")  # Dimensioni compatte

    instruction_label = ctk.CTkLabel(popup, text="Seleziona i clienti di cui vuoi esportare la lista personalizzata :", font=('Arial', 14), text_color="black")
    instruction_label.pack(pady=10)

    container_frame = tk.Frame(popup)
    container_frame.pack(fill="both", expand=True, padx=60)

    canvas = tk.Canvas(container_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(container_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.create_window((0, 0), window=scrollable_frame, anchor="n")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Variabile per tenere traccia dei clienti selezionati
    selected_clienti = []

    def on_checkbutton_toggle(cliente, var):
        if var.get():
            if cliente not in selected_clienti:
                selected_clienti.append(cliente)
        else:
            if cliente in selected_clienti:
                selected_clienti.remove(cliente)

    def toggle_select_all():
        all_selected = all(var.get() for var in selections.values())
        if all_selected:
            for var in selections.values():
                var.set(False)
            selected_clienti.clear()
        else:
            selected_clienti.clear()
            for cliente in clienti:
                var = selections[cliente]
                var.set(True)
                selected_clienti.append(cliente)

    # Aggiungi il pulsante Seleziona/Deseleziona tutto
    select_all_button = ctk.CTkButton(popup, text="Seleziona Tutto", command=toggle_select_all, width=120, height=30)
    select_all_button.pack(pady=5)

    # Dizionario per tracciare le variabili di stato dei checkbutton
    selections = {}

    clienti = get_all_clienti_names()
    for cliente in clienti:
        var = tk.BooleanVar()
        selections[cliente] = var
        check = tk.Checkbutton(scrollable_frame, text=cliente, variable=var,font=('Arial', 14),
                               command=lambda c=cliente, v=var: on_checkbutton_toggle(c, v))
        check.pack(anchor="w")

    # Pulsante di conferma
    confirm_button = ctk.CTkButton(popup, text="Conferma", command=lambda: [popup.destroy(), export_liste_prsn_to_excel(selected_clienti)], width=120, height=30)
    confirm_button.pack(pady=10)

    popup.mainloop()
def open_clienti_selection_popup():
    popup = tk.Toplevel()
    popup.title("Seleziona Clienti da Esportare")
    popup.geometry("700x700")  # Dimensioni compatte

    instruction_label = ctk.CTkLabel(popup, text="Seleziona i clienti che vuoi esportare:", font=('Arial', 14), text_color="black")
    instruction_label.pack(pady=10)

    container_frame = tk.Frame(popup)
    container_frame.pack(fill="both", expand=True, padx=60)

    canvas = tk.Canvas(container_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(container_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.create_window((0, 0), window=scrollable_frame, anchor="n")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Variabile per tenere traccia dei clienti selezionati
    selected_clienti = []

    def on_checkbutton_toggle(cliente, var):
        if var.get():
            if cliente not in selected_clienti:
                selected_clienti.append(cliente)
        else:
            if cliente in selected_clienti:
                selected_clienti.remove(cliente)

    def toggle_select_all():
        all_selected = all(var.get() for var in selections.values())
        if all_selected:
            for var in selections.values():
                var.set(False)
            selected_clienti.clear()
        else:
            selected_clienti.clear()
            for cliente in clienti:
                var = selections[cliente]
                var.set(True)
                selected_clienti.append(cliente)

    # Aggiungi il pulsante Seleziona/Deseleziona tutto
    select_all_button = ctk.CTkButton(popup, text="Seleziona Tutto", command=toggle_select_all, width=120, height=30)
    select_all_button.pack(pady=5)

    # Dizionario per tracciare le variabili di stato dei checkbutton
    selections = {}

    clienti = get_all_clienti_names()
    for cliente in clienti:
        var = tk.BooleanVar()
        selections[cliente] = var
        check = tk.Checkbutton(scrollable_frame, text=cliente, variable=var,font=('Arial', 14),
                               command=lambda c=cliente, v=var: on_checkbutton_toggle(c, v))
        check.pack(anchor="w")

    # Pulsante di conferma
    confirm_button = ctk.CTkButton(popup, text="Conferma", command=lambda: [popup.destroy(), export_clienti_to_excel(selected_clienti)], width=120, height=30)
    confirm_button.pack(pady=10)

    popup.mainloop()
def get_cliente_info(cliente):
    info = get_cliente_info_by_name(cliente)
    return info

def show_cliente_info(cliente, tree, table_frame):
    for row in tree.get_children():
        tree.delete(row)
    info = get_cliente_info(cliente)
    if info:
        for key, value in info.items():
            tree.insert("", tk.END, values=[f"{key}: {value}"])
    tree.cliente = cliente # Save the current client in the tree
    table_frame.pack(pady=10, fill="both", expand=True)
def update_combobox():
    value = combobox.get().lower()
    if value == '':
        combobox['values'] = clienti
    else:
        data = [item for item in clienti if value in item.lower()]
        combobox['values'] = data
        combobox.event_generate('<Down>')

def on_keyrelease(event):
    combobox.after(2000, update_combobox)

def copy_selection(tree):
    selected_items = tree.selection()  # Ottiene tutti gli elementi selezionati
    copied_data = []
    for item in selected_items:
        item_values = tree.item(item, "values")
        copied_data.append("\t".join(map(str, item_values)))  # Unisce i valori di ciascuna riga con una tabulazione
    clipboard_text = "\n".join(copied_data)  # Unisce tutte le righe con un ritorno a capo
    tree.clipboard_clear()  # Pulisce gli appunti
    tree.clipboard_append(clipboard_text)  # Aggiunge il testo agli appunti
    messagebox.showinfo("Copiato", "I testi selezionati sono stati copiati negli appunti.")


def setup_context_menu(tree):
    # Creare un menu contestuale
    context_menu = Menu(tree, tearoff=0)
    context_menu.add_command(label="Copia", command=lambda: copy_selection(tree), font=('Arial', 14))

    def on_right_click(event):
        # Mostrare il menu contestuale
        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    tree.bind("<Button-3>", on_right_click)  # <Button-3> è il clic del pulsante destro del mouse

# Aggiungere il supporto per la selezione trascinando
def on_mouse_drag(event):
    # Identifica l'item su cui si trova il cursore
    item = tree.identify_row(event.y)
    if item:
        tree.selection_add(item)


def on_double_click(event, tree):
    # Get the current cliente saved in the tree
    cliente = tree.cliente
    print(cliente)
    if cliente:
        # Retrieve the customer's full info using the cliente name
        cliente_info = get_cliente_info(cliente)
        print(cliente_info)

        if cliente_info:
            # Show the action dialog with the customer's info
            show_action_dialog(cliente, lambda action: handle_action(
                action, tree,
                cliente_info.get("Ragione sociale", ""),
                cliente_info.get("2° riga rag. sociale", ""),
                cliente_info.get("Indirizzo", ""),
                cliente_info.get("CAP", ""),
                cliente_info.get("Città", ""),
                cliente_info.get("Nazione", ""),
                cliente_info.get("Partita iva estero", ""),
                cliente_info.get("Telefono", ""),
                cliente_info.get("Email", ""),
                cliente_info.get("Zona", ""),
                cliente_info.get("Giorni di chiusura ", ""),
                cliente_info.get("Orari di scarico", ""),
                cliente_info.get("Condizioni pagamamento", ""),
                cliente_info.get("Sconto", ""),
                cliente_info.get("Agente 1", ""),
                cliente_info.get("ID", "")
            ))
        else:
            messagebox.showwarning("Attenzione", "Nessun dato disponibile per il cliente selezionato.")
    else:
        messagebox.showwarning("Attenzione", "Nessun cliente selezionato.")
button_font = ("Arial", 12)  # Font più grande per i pulsanti
button_width = 15  # Larghezza maggiore per i pulsanti
button_height = 2  # Altezza maggiore per i pulsanti
def show_action_dialog(ragione_sociale, callback):
    dialog = tk.Toplevel()
    dialog.title("Scegli Azione")
    dialog.grab_set()  # Ottieni il focus sulla finestra di dialogo
    dialog.transient()  # Rendi la finestra di dialogo modale

    label = tk.Label(dialog, text=f"Vuoi modificare o eliminare '{ragione_sociale}'?",
                         font=("Arial", 16))
    label.pack(pady=10)

    button_frame = tk.Frame(dialog)
    button_frame.pack(pady=10)

    def on_modify():
        dialog.destroy()
        callback("modify")

    def on_delete():
        dialog.destroy()
        callback("delete")

    modify_button = tk.Button(button_frame, text="Modifica", font=button_font, width=button_width,
                              height=button_height, command=on_modify)
    modify_button.grid(row=0, column=0, padx=10, pady=10)

    delete_button = tk.Button(button_frame, text="Elimina", font=button_font, width=button_width, height=button_height,
                              command=on_delete)
    delete_button.grid(row=0, column=1, padx=10, pady=10)

    center_window(dialog,800,400)
    dialog.wait_window()  # Attendi la chiusura della finestra di dialogo

def handle_action(action,tree, ragione_sociale, seconda_riga, indirizzo, cap, citta, nazione, partita_iva, telefono, email, zona, giorni_chiusura, orari_scarico, condizioni_pagamento, sconto, agente, id_cliente):
    if action == "modify":
        details = ask_details(ragione_sociale, f"Inserisci le nuove info di '{ragione_sociale}':",seconda_riga, indirizzo, cap, citta, nazione, partita_iva, telefono, email, zona, giorni_chiusura, orari_scarico, condizioni_pagamento, sconto, agente, id_cliente)
        if details and details['quantity'] is not None:
            #inserisci funzione per modificare a db il cliente
            show_cliente_info(tree, tree.table_frame)
            messagebox.showinfo("Modifica Prodotto", f"Hai modificato '{ragione_sociale}'.")
    elif action == "delete":
        confirm = messagebox.askokcancel("Conferma Eliminazione", f"Sei sicuro di voler eliminare '{ragione_sociale}'?")
        if confirm:
            #inserisci funzione per far delete cliente
            show_cliente_info(tree.table_frame)
            messagebox.showinfo("Eliminazione Prodotto", f"Il cliente '{ragione_sociale}' è stato eliminato.")

def ask_details(ragione_sociale, prompt,seconda_riga, indirizzo, cap, citta, nazione, partita_iva, telefono, email, zona, giorni_chiusura, orari_scarico, condizioni_pagamento, sconto, agente, id_cliente):
    dialog = tk.Toplevel()
    dialog.title("Scegli le nuove info del prodotto")
    dialog.grab_set()
    dialog.transient()

    label = tk.Label(dialog, text=prompt, font=("Arial", 14))
    label.pack(pady=10)

    entry_width = 40

    input_vars = {
        "ragione sociale": tk.StringVar(value=ragione_sociale),
        "seconda riga": tk.StringVar(value=seconda_riga),
        "indirizzo": tk.StringVar(value=indirizzo),
        "cap": tk.StringVar(value=cap),
        "citta": tk.StringVar(value=citta),
        "nazione": tk.StringVar(value=nazione),
        "partita iva": tk.StringVar(value=partita_iva),
        "telefono": tk.StringVar(value=telefono),
        "email": tk.StringVar(value=email),
        "zona": tk.StringVar(value=zona),
        "giorni chiusura": tk.StringVar(value=giorni_chiusura),
        "orari scarico": tk.StringVar(value=orari_scarico),
        "condizioni pagamento": tk.StringVar(value=condizioni_pagamento),
        "sconto": tk.DoubleVar(value=sconto),
        "agente": tk.StringVar(value=agente),
        "id cliente": tk.IntVar(value=id_cliente),
    }

    container_frame = tk.Frame(dialog)
    container_frame.pack(fill="both", expand=True, padx=60)

    canvas = tk.Canvas(container_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(container_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.create_window((0, 0), window=scrollable_frame, anchor="n")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Crea le entry per ciascun campo
    for label_text, var in input_vars.items():
        label = tk.Label(scrollable_frame, text=label_text.capitalize() + ":", font=("Arial", 14))
        label.pack(pady=5)
        entry = tk.Entry(scrollable_frame, width=entry_width, textvariable=var, font=("Arial", 14))
        entry.pack(pady=5)

    def on_confirm():
        dialog.destroy()
        dialog.details = {label: var.get() for label, var in input_vars.items()}

    confirm_button = tk.Button(dialog, text="Conferma",  font=("Arial", 12) , command=on_confirm)
    confirm_button.pack(pady=10)

    center_window(dialog,700,700)
    dialog.wait_window()
    return getattr(dialog, 'details', None)


def show_dashboard3(parent_frame):
    global combobox, clienti, tree, tree_list
    for widget in parent_frame.winfo_children():
        widget.destroy()

    parent_frame_color = parent_frame.cget("fg_color")

    # Crea un frame per contenere sia il titolo che la legenda
    top_frame = ctk.CTkFrame(parent_frame, corner_radius=5, fg_color=parent_frame_color)
    top_frame.pack(fill="x", pady=0)  # Si espande orizzontalmente

    # Usa il layout grid nel top_frame
    top_frame.columnconfigure(0, weight=0)  # La colonna 0 non si espande
    top_frame.columnconfigure(1, weight=1)  # La colonna 1 si espande

    # Aggiungi la legenda all'interno del top_frame, allineata a destra
    legend_label = ctk.CTkLabel(top_frame, text="Unità di misura:\n- L: liters\n- Kg: kilograms\n- Pz: pieces",
                                font=('Arial', 12), justify="left")
    legend_label.grid(row=0, column=0, sticky="w", padx=10)

    instruction_label = ctk.CTkLabel(top_frame, text="Scegliere Cliente:", font=('Arial', 14))
    instruction_label.grid(row=0, column=1, padx=10)

    clienti = get_all_clienti_names()
    selected_cliente = tk.StringVar()

    # Frame per allineare il menù a tendina e i pulsanti "Esporta in Excel"
    search_frame = ctk.CTkFrame(parent_frame,corner_radius=5)
    search_frame.pack(pady=10)

    # Style for larger font in Combobox
    style = ttk.Style()
    style.configure("TCombobox", font=('Arial', 16))  # Adjust font size as needed
    style.configure("TCombobox*Listbox*Font", font=('Arial', 16))  # Ensure larger font size for dropdown menu items

    # Creazione del menù a tendina con ricerca incrementale
    combobox = ttk.Combobox(search_frame, textvariable=selected_cliente, values=clienti, font=('Arial', 16))
    combobox.configure(width=35)
    combobox.pack(side="left", padx=10)

    # Abilita la ricerca incrementale nel menù a tendina
    combobox.bind('<KeyRelease>', on_keyrelease)
    combobox.pack(pady=10)

    def on_cliente_selected(event):
        cliente = selected_cliente.get()
        show_cliente_info(cliente, tree, table_frame)

    combobox.bind("<<ComboboxSelected>>", on_cliente_selected)

    # Create a style for the Listbox within the Combobox dropdown
    combobox.option_add('*TCombobox*Listbox*Font', ('Arial', 16))

    # Pulsante "Esporta Excel Fornitori" accanto al pulsante prodotti
    export_fornitori_button = ctk.CTkButton(search_frame, text="Esporta Excel Clienti", command=open_clienti_selection_popup, font=('Arial', 14))
    export_fornitori_button.pack(side="left", padx=10)

    export_lista_button = ctk.CTkButton(search_frame, text="Esporta Excel lista prs", command=open_lista_selection_popup, font=('Arial', 14))
    export_lista_button.pack(side="left", padx=10)

    modifica_lista_button = ctk.CTkButton(search_frame, text="Modifica lista prs", command=open_clienti_selection_popup, font=('Arial', 14))
    modifica_lista_button.pack(side="left", padx=10)



    global table_frame, table_frame_1
    table_frame = ctk.CTkFrame(parent_frame, corner_radius=5)
    table_frame.pack_forget()
    table_frame_1 = ctk.CTkFrame(parent_frame, corner_radius=5)
    table_frame_1.pack_forget()

    table_title = ctk.CTkLabel(table_frame, text="Dati Cliente:", font=('Arial', 14, 'bold'))
    table_title.pack(pady=10)

    table_title_1 = ctk.CTkLabel(table_frame_1, text="Lista personalizzata:", font=('Arial', 14, 'bold'))
    table_title_1.pack(pady=10)

    # Create a frame to hold the table and scrollbars
    table_container = tk.Frame(table_frame, width=800, height=400)
    table_container.pack_propagate(False)  # Prevent the frame from resizing to fit its children
    table_container.pack(pady=10, fill="both", expand=True)

    table_container_1 = tk.Frame(table_frame_1, width=800, height=400)
    table_container_1.pack_propagate(False)  # Prevent the frame from resizing to fit its children
    table_container_1.pack(pady=10, fill="both", expand=True)

    # Use only one column for the Treeview to display data vertically
    global columns
    columns = ("",)

    global tree
    tree = ttk.Treeview(table_container, columns=columns, show="headings")
    tree.table_container = table_container # Save the frame in the treeview for reference
    setup_context_menu(tree)
    tree.bind("<B1-Motion>", on_mouse_drag)


    style = ttk.Style()
    style.configure("Treeview",
                    rowheight=30,
                    font=('Arial', 14),
                    background="#f1f8e9",
                    foreground="#004d40",
                    fieldbackground="#f1f8e9",
                    bordercolor="#000000",
                    relief="solid",
                    borderwidth=1)
    style.configure("Treeview.Heading",
                    font=('Arial', 16, 'bold'),
                    background="#a5d6a7",
                    foreground="#004d40")
    style.map("Treeview",
              background=[('selected', '#c8e6c9')],
              foreground=[('selected', '#004d40')])

    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=1500, anchor="w")  # Set a larger width to accommodate longer text

    yscrollbar = ttk.Scrollbar(table_container, orient=tk.VERTICAL, command=tree.yview)
    yscrollbar.pack(side="right", fill="y")
    tree.configure(yscroll=yscrollbar.set)

    tree.pack(side="left", fill="both", expand=True)
    tree.bind("<Double-1>", lambda event: on_double_click(event, tree))

# Other components of your application remain unchanged
