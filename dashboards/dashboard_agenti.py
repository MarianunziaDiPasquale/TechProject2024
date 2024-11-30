import os

from tkinter.filedialog import asksaveasfilename
import customtkinter as ctk
from tkinter import ttk, Menu, filedialog
import tkinter as tk
import datetime
from datetime import datetime

from fpdf import FPDF

#da cambiare in funzioni per agenti
from Database_Utilities.crud_agenti import get_all_clienti_names, get_cliente_info_by_name
import openpyxl
from tkinter import messagebox


#from dashboards.dashboard_clienti import get_lista_info
from db_prova_lista_personalizzata import get_prodotti_by_cliente
from popup_functions import open_add_popup

from Database_Utilities.connection import _connection

def get_clienti_by_agente(selected_agente):
    """Fetch all clients associated with the selected agent."""
    #print("check1")
    conn = _connection() # Update this with your database path
    cursor = conn.cursor()

    query = "SELECT `Ragione sociale` FROM `clienti` WHERE `Agente 1` = %s"
    cursor.execute(query, (selected_agente,))

    clienti = cursor.fetchall()
    #print(clienti)
    conn.close()

    return [cliente[0] for cliente in clienti]  # Returns a list of customer names


def center_window(window, width, height):
    window.update_idletasks()
    width = width
    height = height
    x = (window.winfo_screenwidth() // 2) - (width // 2)
    y = (window.winfo_screenheight() // 2) - (height // 2)
    window.geometry(f'{width}x{height}+{x}+{y}')

def export_agenti_to_excel(selected_agenti):
    if not selected_agenti:
        messagebox.showwarning("Nessun Dato", "Non ci sono agenti selezionati da esportare.")
        return

    if len(selected_agenti) == 1:
        default_filename = f"Info_Agente_{selected_agenti[0]}.xlsx"
    else:
        default_filename = f"Info_Agenti.xlsx"
    file_path = asksaveasfilename(defaultextension=".xlsx",
                                  filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                  initialfile=default_filename)
    # Verifica se l'utente ha scelto un percorso
    if not file_path:
        messagebox.showwarning("Salvataggio Annullato", "Salvataggio del file Excel annullato.")
        return

    workbook = openpyxl.Workbook()

    for agente in selected_agenti:
        #cambiare in funzione per agenti
        info = get_cliente_info_by_name(agente)

        # Crea un nuovo foglio per ogni cliente
        sheet = workbook.create_sheet(title=agente)

        #cambiare in colonne per agente
        headers = ["Ragione sociale", "Indirizzo", "CAP", "Città", "Nazione", "Partita IVA", "Telefono", "Email", "Zona", "Giorni chiusura", "Orari di scarico", "Condizioni pagamento", "Sconto", "Agente"]
        sheet.append(headers)

        # Aggiungi i dati del cliente se esistono, altrimenti lascia solo le intestazioni
        if info:
            sheet.append([info.get(header, "N/A") for header in headers])
        else:
            # Aggiungi una riga vuota se non ci sono dati, lasciando solo l'intestazione
            sheet.append(["Nessun dato disponibile"])

    # Rimuove il foglio predefinito solo se ci sono altri fogli visibili
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook["Sheet"]

    workbook.save(file_path)
    messagebox.showinfo("Excel Generato", f"Il file Excel è stato salvato come {os.path.basename(file_path)}")

def export_clienti_conn_to_excel(selected_agente_conn):
    if not selected_agente_conn:

        messagebox.showwarning("Nessun Dato", "Non ci sono agenti selezionati per clienti connessi da esportare.")
        return

    if len(selected_agente_conn) == 1:
        default_filename = f"Info_Agente_Clienti_Connessi_{selected_agente_conn[0]}.xlsx"
    file_path = asksaveasfilename(defaultextension=".xlsx",
                                  filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                  initialfile=default_filename)
    # Verifica se l'utente ha scelto un percorso
    if not file_path:
        messagebox.showwarning("Salvataggio Annullato", "Salvataggio del file Excel annullato.")
        return

    workbook = openpyxl.Workbook()

    for agente in selected_agente_conn:

        # modificare con funzione per recuperare clienti connessi
        info = get_lista_info(agente)

        # Crea un nuovo foglio per ogni cliente
        sheet = workbook.create_sheet(title=agente)

        # modificare con colonne clienti connessi
        headers = ["Ragione sociale", "Indirizzo", "CAP", "Città"]
        sheet.append(headers)

        # Aggiungi i dati del cliente se esistono, altrimenti lascia solo le intestazioni
        if info:
            for row in info:
                sheet.append(row)
        else:
            # Aggiungi una riga vuota se non ci sono dati, lasciando solo l'intestazione
            sheet.append(["Nessun dato disponibile"])

    # Rimuove il foglio predefinito solo se ci sono altri fogli visibili
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook["Sheet"]

    workbook.save(file_path)
    messagebox.showinfo("Excel Generato", f"Il file Excel è stato salvato come {os.path.basename(file_path)}")

def open_clienti_conn_popup():
    popup = tk.Toplevel()
    popup.title("Seleziona la lista ")
    popup.geometry("700x700")  # Dimensioni compatte

    instruction_label = ctk.CTkLabel(popup, text="Seleziona gli agenti di cui vuoi esportare i clienti connessi :", font=('Arial', 14), text_color="black")
    instruction_label.pack(pady=10)

    container_frame = tk.Frame(popup)
    container_frame.pack(fill="both", expand=True, padx=60)

    canvas = tk.Canvas(container_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(container_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.create_window((0, 0), window=scrollable_frame, anchor="n")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Variabile per tenere traccia dei clienti selezionati
    selected_agente_conn = []

    def on_checkbutton_toggle(agente, var):
        if var.get():
            if agente not in selected_agente_conn:
                selected_agente_conn.append(agente)
        else:
            if agente in selected_agente_conn:
                selected_agente_conn.remove(agente)

    def toggle_select_all():
        all_selected = all(var.get() for var in selections.values())
        if all_selected:
            for var in selections.values():
                var.set(False)
            selected_agente_conn.clear()
        else:
            selected_agente_conn.clear()
            for agente in agenti:
                var = selections[agente]
                var.set(True)
                selected_agente_conn.append(agente)

    # Aggiungi il pulsante Seleziona/Deseleziona tutto
    select_all_button = ctk.CTkButton(popup, text="Seleziona Tutto", command=toggle_select_all, width=120, height=30)
    select_all_button.pack(pady=5)

    # Dizionario per tracciare le variabili di stato dei checkbutton
    selections = {}


    #cambiare per recuperare clienti connessi
    agenti= get_all_clienti_names()
    for agente in agenti:
        var = tk.BooleanVar()
        selections[agente] = var
        check = tk.Checkbutton(scrollable_frame, text=agente, variable=var,font=('Arial', 14),
                               command=lambda c=agente, v=var: on_checkbutton_toggle(c, v))
        check.pack(anchor="w")

    # Pulsante di conferma
    confirm_button = ctk.CTkButton(popup, text="Conferma", command=lambda: [popup.destroy(), export_clienti_conn_to_excel(selected_agente_conn)], width=120, height=30)
    confirm_button.pack(pady=10)

    popup.mainloop()

def open_agenti_selection_popup():
    popup = tk.Toplevel()
    popup.title("Seleziona Agenti da Esportare")
    popup.geometry("700x700")  # Dimensioni compatte

    instruction_label = ctk.CTkLabel(popup, text="Seleziona gli agenti che vuoi esportare:", font=('Arial', 14), text_color="black")
    instruction_label.pack(pady=10)

    container_frame = tk.Frame(popup)
    container_frame.pack(fill="both", expand=True, padx=60)

    canvas = tk.Canvas(container_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(container_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.create_window((0, 0), window=scrollable_frame, anchor="n")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Variabile per tenere traccia dei clienti selezionati
    selected_agenti = []

    def on_checkbutton_toggle(agente, var):
        if var.get():
            if agente not in selected_agenti:
                selected_agenti.append(agente)
        else:
            if agente in selected_agenti:
                selected_agenti.remove(agente)

    def toggle_select_all():
        all_selected = all(var.get() for var in selections.values())
        if all_selected:
            for var in selections.values():
                var.set(False)
            selected_agenti.clear()
        else:
            selected_agenti.clear()
            for agente in agenti:
                var = selections[agente]
                var.set(True)
                selected_agenti.append(agente)

    # Aggiungi il pulsante Seleziona/Deseleziona tutto
    select_all_button = ctk.CTkButton(popup, text="Seleziona Tutto", command=toggle_select_all, width=120, height=30)
    select_all_button.pack(pady=5)

    # Dizionario per tracciare le variabili di stato dei checkbutton
    selections = {}

    #modificare per agente
    agenti = get_all_clienti_names()
    for agente in agenti:
        var = tk.BooleanVar()
        selections[agente] = var
        check = tk.Checkbutton(scrollable_frame, text=agente, variable=var,font=('Arial', 14),
                               command=lambda c=agente, v=var: on_checkbutton_toggle(c, v))
        check.pack(anchor="w")

    # Pulsante di conferma
    confirm_button = ctk.CTkButton(popup, text="Conferma", command=lambda: [popup.destroy(), export_agenti_to_excel(selected_agenti)], width=120, height=30)
    confirm_button.pack(pady=10)

    popup.mainloop()

def get_agente_info(agente):
    #cambiare funzione per agente
    info = get_cliente_info_by_name(agente)
    return info

def get_clienti_conn(agente):
    #cambiare con funzione che recupera per agente
    info = get_prodotti_by_cliente(agente)
    return info

def show_agente_info(agente, tree, table_frame):
    for row in tree.get_children():
        tree.delete(row)
    #cambiare per agente
    info = get_agente_info(agente)
    if info:
        for key, value in info.items():
            tree.insert("", tk.END, values=[f"{key}: {value}"])
    tree.agente = agente # Save the current client in the tree
    table_frame.pack(pady=10, fill="both", expand=True)

def show_agente_clienti_conn(agente, tree, table_frame):
    # Clear the table
    for row in tree.get_children():
        tree.delete(row)

    # Fetch the connected customers based on the selected agent
    clienti_conn = get_clienti_by_agente(agente)

    # Insert the customer names into the table
    for cliente_conn in clienti_conn:
        tree.insert("", tk.END, values=(cliente_conn,))

    tree.agente = agente  # Save the current agent in the tree
    table_frame.pack(pady=10, fill="both", expand=True)

def update_combobox():
    value = combobox.get().lower()
    if value == '':
        combobox['values'] = agenti
    else:
        data = [item for item in agenti if value in item.lower()]
        combobox['values'] = data
        combobox.event_generate('<Down>')

def on_keyrelease(event):
    combobox.after(2000, update_combobox)

def copy_selection(tree):
    selected_items = tree.selection()  # Ottiene tutti gli elementi selezionati
    copied_data = []
    for item in selected_items:
        item_values = tree.item(item, "values")
        copied_data.append("\t".join(map(str, item_values)))  # Unisce i valori di ciascuna riga con una tabulazione
    clipboard_text = "\n".join(copied_data)  # Unisce tutte le righe con un ritorno a capo
    tree.clipboard_clear()  # Pulisce gli appunti
    tree.clipboard_append(clipboard_text)  # Aggiunge il testo agli appunti
    messagebox.showinfo("Copiato", "I testi selezionati sono stati copiati negli appunti.")

def setup_context_menu(tree):
    # Creare un menu contestuale
    context_menu = Menu(tree, tearoff=0)
    context_menu.add_command(label="Copia", command=lambda: copy_selection(tree), font=('Arial', 14))

    def on_right_click(event):
        # Mostrare il menu contestuale
        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    tree.bind("<Button-3>", on_right_click)  # <Button-3> è il clic del pulsante destro del mouse

# Aggiungere il supporto per la selezione trascinando
def on_mouse_drag(event, tree):
    # Identifica l'item su cui si trova il cursore
    item = tree.identify_row(event.y)
    if item:
        tree.selection_add(item)


def on_double_click(event, tree):
    # Get the current cliente saved in the tree
    agente = tree.agente
    print(agente)
    if agente:
        # Retrieve the customer's full info using the cliente name
        #cambiare per agente
        cliente_info = get_agente_info(agente)
        print(cliente_info)

        if cliente_info:
            # Show the action dialog with the customer's info
            #cambiare in righe per agente
            show_action_dialog(agente, lambda action: handle_action(
                action, tree,agente,
                cliente_info.get("nome", ""),
                cliente_info.get("id", "")
            ))
        else:
            messagebox.showwarning("Attenzione", "Nessun dato disponibile per l'agente selezionato.")
    else:
        messagebox.showwarning("Attenzione", "Nessun agente selezionato.")

def on_double_click_1(event, tree_1):
    item = tree_1.selection()[0]
    values = tree_1.item(item, "values")
    if values:
        current_agente = values[0]
        current_id = values[1]
        current_cliente_connesso= values[2]
        current_quantita = values[3]
        show_action_dialog_1(current_cliente_connesso, lambda action: handle_action_1(action, current_agente,current_id,current_cliente_connesso,current_quantita, tree))

button_font = ("Arial", 12)  # Font più grande per i pulsanti
button_width = 15  # Larghezza maggiore per i pulsanti
button_height = 2  # Altezza maggiore per i pulsanti

def show_action_dialog(agente, callback):
    dialog = tk.Toplevel()
    dialog.title("Scegli Azione")
    dialog.grab_set()  # Ottieni il focus sulla finestra di dialogo
    dialog.transient()  # Rendi la finestra di dialogo modale

    label = tk.Label(dialog, text=f"Vuoi modificare o eliminare '{agente}'?",
                         font=("Arial", 16))
    label.pack(pady=10)

    button_frame = tk.Frame(dialog)
    button_frame.pack(pady=10)

    def on_modify():
        dialog.destroy()
        callback("modify")

    def on_delete():
        dialog.destroy()
        callback("delete")

    modify_button = tk.Button(button_frame, text="Modifica", font=button_font, width=button_width,
                              height=button_height, command=on_modify)
    modify_button.grid(row=0, column=0, padx=10, pady=10)

    delete_button = tk.Button(button_frame, text="Elimina", font=button_font, width=button_width, height=button_height,
                              command=on_delete)
    delete_button.grid(row=0, column=1, padx=10, pady=10)

    center_window(dialog,800,400)
    dialog.wait_window()  # Attendi la chiusura della finestra di dialogo

def show_action_dialog_1(current_cliente_connesso, callback):
    dialog = tk.Toplevel()
    dialog.title("Scegli Azione")
    dialog.grab_set()  # Ottieni il focus sulla finestra di dialogo
    dialog.transient()  # Rendi la finestra di dialogo modale

    label = tk.Label(dialog, text=f"Vuoi modificare o eliminare '{current_cliente_connesso}'?",font=("Arial", 16))
    label.pack(pady=10)

    button_frame = tk.Frame(dialog)
    button_frame.pack(pady=10)

    def on_modify():
        dialog.destroy()
        callback("modify")

    def on_delete():
        dialog.destroy()
        callback("delete")

    modify_button = tk.Button(button_frame, text="Nuova quantità", font=button_font, width=button_width,
                              height=button_height, command=on_modify)
    modify_button.grid(row=0, column=0, padx=10, pady=10)

    delete_button = tk.Button(button_frame, text="Elimina", font=button_font, width=button_width, height=button_height,
                              command=on_delete)
    delete_button.grid(row=0, column=1, padx=10, pady=10)

    center_window(dialog,600,300)
    dialog.wait_window()  # Attendi la chiusura della finestra di dialogo

def handle_action(action,tree,agente,nome, id):
    if action == "modify":
        details = ask_details(agente, f"Inserisci le nuove info di '{agente}':",nome, id )
        if details and details['quantity'] is not None:
            #inserisci funzione per modificare a db il cliente
            show_agente_info(tree, tree.table_frame)
            messagebox.showinfo("Modifica Prodotto", f"Hai modificato '{agente}'.")
    elif action == "delete":
        confirm = messagebox.askokcancel("Conferma Eliminazione", f"Sei sicuro di voler eliminare '{agente}'?")
        if confirm:
            #inserisci funzione per far delete cliente
            show_agente_info(tree.table_frame)
            messagebox.showinfo("Eliminazione Prodotto", f"L'agente '{agente}' è stato eliminato.")

def ask_details(agente, prompt,nome, id):
    dialog = tk.Toplevel()
    dialog.title("Scegli le nuove info dell'agente")
    dialog.grab_set()
    dialog.transient()

    label = tk.Label(dialog, text=prompt, font=("Arial", 14))
    label.pack(pady=10)

    entry_width = 40
    #cambiare per agente
    input_vars = {
        "agente": tk.StringVar(value=nome),
        "id cliente": tk.IntVar(value=id),
    }

    container_frame = tk.Frame(dialog)
    container_frame.pack(fill="both", expand=True, padx=60)

    canvas = tk.Canvas(container_frame)
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(container_frame, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    scrollable_frame = tk.Frame(canvas)
    scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

    canvas.create_window((0, 0), window=scrollable_frame, anchor="n")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Crea le entry per ciascun campo
    for label_text, var in input_vars.items():
        label = tk.Label(scrollable_frame, text=label_text.capitalize() + ":", font=("Arial", 14))
        label.pack(pady=5)
        entry = tk.Entry(scrollable_frame, width=entry_width, textvariable=var, font=("Arial", 14))
        entry.pack(pady=5)

    def on_confirm():
        dialog.destroy()
        dialog.details = {label: var.get() for label, var in input_vars.items()}

    confirm_button = ctk.CTkButton(dialog, text="Conferma",  font=("Arial", 12) , command=on_confirm, width=120, height=30)
    confirm_button.pack(pady=10)

    center_window(dialog,700,700)
    dialog.wait_window()
    return getattr(dialog, 'details', None)

    def print_invoice():
        selected_item = tree.selection()
        if selected_item:
            order = tree.item(selected_item, "values")
            generate_invoice_pdf(order)
        else:
            messagebox.showwarning("Seleziona Ordine", "Per favore, seleziona un ordine per stampare la fattura.")

    class PDF(FPDF):
        def header(self):

            # Aggiungi il logo in alto a sinistra
            self.image('resources/LogoCINCOTTI.jpg', 8, 2, 77, 44, 'JPG')  # Modifica il percorso e le dimensioni

            # Aggiungi l'immagine in alto a destra
            self.image('resources/Logojpeg.jpg', 120, 8, 77, 33, 'JPG')  # Modifica il percorso e le dimensioni

            self.ln(35)

        def footer(self):
            # Aggiungi l'immagine in basso al centro
            self.set_y(-40)  # Posiziona a 30 mm dal fondo della pagina
            self.image('resources/LOGOMOZZABELLAECINCOTTI.jpg', 75, self.get_y(), 70, 40,
                       'JPG')  # Modifica il percorso e le dimensioni

        def draw_table(self, data, col_widths, headers=None, bold_headers=False):
            if headers:
                if bold_headers:
                    self.set_font('Arial', 'B', 10)
                else:
                    self.set_font('Arial', '', 10)

                for header, width in zip(headers, col_widths):
                    self.cell(width, 8, header, 1, 0, 'C')
                self.ln()

            if bold_headers:
                self.set_font('Arial', '', 10)

            for row in data:
                # Prima passata per determinare l'altezza massima della riga
                row_heights = []
                for datum, width in zip(row, col_widths):
                    line_height = 6  # Altezza per linea, può essere aggiustata
                    num_lines = self.get_string_width(datum) / width  # Calcola il numero di righe necessarie
                    num_lines = datum.count('\n') + 1  # Contiamo le nuove linee
                    height = line_height * num_lines
                    row_heights.append(height)
                max_row_height = max(row_heights)  # Altezza massima tra tutte le celle della riga

                # Seconda passata per disegnare le celle con l'altezza della riga determinata
                x_start = self.get_x()  # Posizione iniziale X per la riga
                y_start = self.get_y()  # Posizione iniziale Y per la riga
                for datum, width in zip(row, col_widths):
                    # Disegna la cella utilizzando multi_cell per il testo con ritorni a capo
                    self.multi_cell(width, line_height, datum, 1, 'L', False)
                    # Posiziona il cursore all'inizio della prossima cella nella riga
                    self.set_xy(x_start + width, y_start)
                    x_start += width  # Avanza per la larghezza della cella
                self.ln(max_row_height)  # Salta alla prossima riga usando l'altezza massima

        def invoice_header(self, invoice):
            self.set_font('Arial', '', 10)
            data = [
                f"Document: INVOICE", f"n. {invoice['number']}", f"Data: {invoice['Date_1']}"
            ]
            col_widths = [70, 40, 70]  # Regola le larghezze delle colonne come necessario
            self.draw_table([data], col_widths)

        def customer_details(self, invoice):
            data = [
                ["Goods destination:\n" + invoice["customer_phone"], "Customer:\n" + invoice["customer_address"]],
            ]
            col_widths = [70, 110]  # Regola le larghezze delle colonne come necessario
            self.draw_table(data, col_widths, bold_headers=True)

        def customer_details_1(self, invoice):
            data = [
                ["ID:\n" + invoice["id"], "Email:\n" + invoice["customer_email"],"VAT:\n" + invoice["vat_number"],"Nation:\n" + invoice["nation"]],
            ]
            col_widths = [25,65,45,45]  # Regola le larghezze delle colonne come necessario
            self.draw_table(data, col_widths, bold_headers=True)

        def bank_details(self, invoice):
            data = [
                ["IBAN:\n"+invoice["iban"], "Swift:\n"+invoice["swift"]]
            ]
            col_widths = [90,90]
            self.set_font('Arial', '', 8)
            self.draw_table(data, col_widths, bold_headers=True)

        def agent_details(self, invoice):
            data = [
                ["Agent:\n"+ invoice["agent_name"],"Payment Condition:\n"+ invoice["payment_condition"],"Sender:\n"+ invoice["sender_name"], "Giorno di chiusura:\n"+ invoice["sender_giorno"]]
            ]
            col_widths = [55,35,45,45]
            self.set_font('Arial', '', 8)
            self.draw_table(data, col_widths, bold_headers=True)

        def product_table(self, products):
            headers = ["Article", "Description", "Quantity", "Price", "Discount", "Amount"]
            data = [
                [product['article'], product['product_description'], str(product['product_quantity']),
                 str(product['product_price']), str(product['product_discount']), str(product['product_amount'])]
                for product in products
            ]
            col_widths = [20, 70, 20, 20, 20, 30]
            self.set_font('Arial', 'B', 10)
            self.cell(0, 8, 'Product Details', 0, 1, 'L')
            self.draw_table(data, col_widths, headers, bold_headers=True)

        def total_details(self, invoice):
            data = [
                ["Total Quantity:\n" + str(invoice["total_quantity"]),"IVA:\n" + str(invoice["iva_amount"]), "Sender:\n" + invoice["sender_name"]],
                ["Reason for export:\n" + "Vendita","Courier Service:\n" + invoice["courier_service"],"Recipient:\n" + invoice["recipient"]],
                ["Total Amount:\n" + str(invoice["total_amount"]),"IVA Amount:\n" + str(invoice["iva_amount"]),"Total Invoice EURO:\n" + str(invoice["total_invoice_euro"])],
            ]
            col_widths = [60,60,60]
            self.set_font('Arial', 'B', 10)
            self.cell(0, 8, 'Total Details', 0, 1, 'L')
            self.draw_table(data, col_widths, bold_headers=True)

        def thank_you_details(self, invoice):
            self.set_font('Arial', 'B', 10)
            self.cell(0, 8, 'Thank you for choosing our products!', 0, 1, 'R')

    def fetch_invoice_data(invoice_number):
        conn = _connection()
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM orders_fattura WHERE id = '{invoice_number}'")
        column_names = [description[0] for description in cursor.description]
        invoice_tuple = cursor.fetchone()
        conn.close()
        if invoice_tuple:
            invoice = dict(zip(column_names, invoice_tuple))
            return invoice
        return None

    def fetch_invoice_products(invoice_number):
        conn = _connection()
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM orders_fattura WHERE id = '{invoice_number}'")
        column_names = [description[0] for description in cursor.description]
        product_rows = cursor.fetchall()
        conn.close()
        products = [dict(zip(column_names, product_row)) for product_row in product_rows]
        return products

    def generate_invoice_pdf(order):
        invoice_number = order[0]
        invoice = fetch_invoice_data(invoice_number)
        products = fetch_invoice_products(invoice_number)
        if not invoice:
            messagebox.showerror("Errore", "Dati della fattura non trovati.")
            return

        # Popup window to edit and confirm invoice data
        def edit_and_confirm():
            def save_changes():
                # Update invoice data from the entry fields
                invoice["number"] = entry_number.get()
                invoice["Date_1"] = entry_date.get()
                invoice["customer_phone"] = entry_customer_phone.get()
                invoice["customer_address"] = entry_customer_address.get()
                invoice["id"] = entry_id.get()
                invoice["customer_email"] = entry_customer_email.get()
                invoice["vat_number"] = entry_vat_number.get()
                invoice["nation"] = entry_nation.get()
                invoice["iban"] = entry_iban.get()
                invoice["swift"] = entry_swift.get()
                invoice["agent_name"] = entry_agent_name.get()
                invoice["payment_condition"] = entry_payment_condition.get()
                invoice["sender_name"] = entry_sender_name.get()
                invoice["sender_giorno"] = entry_sender_giorno.get()
                invoice["courier_service"] = entry_courier_service.get()
                invoice["recipient"] = entry_recipient.get()
                invoice["total_quantity"] = entry_total_quantity.get()
                invoice["iva_amount"] = entry_iva_amount.get()
                invoice["total_amount"] = entry_total_amount.get()
                invoice["total_invoice_euro"] = entry_total_invoice_euro.get()

                # Close the popup window after saving changes
                popup.destroy()

                # Open a separate window to choose save path
                choose_save_path()

            # Ottiene la data corrente e la formatta come YYYYMMDD
            current_date = datetime.now().strftime("%Y%m%d")
            filename = f"Fattura_{current_date}"

            def choose_save_path():
                save_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")], initialfile=filename)
                if save_path:
                    pdf = PDF()
                    pdf.add_page()
                    pdf.invoice_header(invoice)
                    pdf.customer_details(invoice)
                    pdf.customer_details_1(invoice)
                    pdf.bank_details(invoice)
                    pdf.agent_details(invoice)
                    pdf.product_table(products)
                    pdf.total_details(invoice)
                    pdf.thank_you_details(invoice)
                    pdf.output(save_path)
                    messagebox.showinfo("Successo", f"Fattura salvata come {os.path.basename(save_path)}")
                else:
                    messagebox.showwarning("Salvataggio Annullato", "Il salvataggio della fattura è stato annullato.")

            # Create the popup window
            popup = tk.Toplevel()
            popup.title("Modifica i dettagli della fattura")
            popup.geometry("800x900")  # Set the size of the popup window
            popup.resizable(True, True)  # Prevent resizing for consistent layout

            # Add padding and spacing between elements
            padding = {'padx': 10, 'pady': 5}
            # Set the width of entry fields and the font size
            entry_width = 40
            font_size = ("Arial", 14)  # Font family Arial, size 14

            # Create and grid labels and entries for each invoice field with larger font
            tk.Label(popup, text="Numero:", font=font_size).grid(row=0, column=0, **padding)
            entry_number = tk.Entry(popup, width=entry_width, font=font_size)
            entry_number.grid(row=0, column=1, **padding)
            entry_number.insert(0, invoice["number"])

            tk.Label(popup, text="Data:", font=font_size).grid(row=1, column=0, **padding)
            entry_date = tk.Entry(popup, width=entry_width, font=font_size)
            entry_date.grid(row=1, column=1, **padding)
            entry_date.insert(0, invoice["Date_1"])

            tk.Label(popup, text="Telefono Cliente:", font=font_size).grid(row=2, column=0, **padding)
            entry_customer_phone = tk.Entry(popup, width=entry_width, font=font_size)
            entry_customer_phone.grid(row=2, column=1, **padding)
            entry_customer_phone.insert(0, invoice["customer_phone"])

            tk.Label(popup, text="Indirizzo Cliente:", font=font_size).grid(row=3, column=0, **padding)
            entry_customer_address = tk.Entry(popup, width=entry_width, font=font_size)
            entry_customer_address.grid(row=3, column=1, **padding)
            entry_customer_address.insert(0, invoice["customer_address"])

            tk.Label(popup, text="ID:", font=font_size).grid(row=4, column=0, **padding)
            entry_id = tk.Entry(popup, width=entry_width, font=font_size)
            entry_id.grid(row=4, column=1, **padding)
            entry_id.insert(0, invoice["id"])

            tk.Label(popup, text="Email Cliente:", font=font_size).grid(row=5, column=0, **padding)
            entry_customer_email = tk.Entry(popup, width=entry_width, font=font_size)
            entry_customer_email.grid(row=5, column=1, **padding)
            entry_customer_email.insert(0, invoice["customer_email"])

            tk.Label(popup, text="Partita IVA:", font=font_size).grid(row=6, column=0, **padding)
            entry_vat_number = tk.Entry(popup, width=entry_width, font=font_size)
            entry_vat_number.grid(row=6, column=1, **padding)
            entry_vat_number.insert(0, invoice["vat_number"])

            tk.Label(popup, text="Nazione:", font=font_size).grid(row=7, column=0, **padding)
            entry_nation = tk.Entry(popup, width=entry_width, font=font_size)
            entry_nation.grid(row=7, column=1, **padding)
            entry_nation.insert(0, invoice["nation"])

            tk.Label(popup, text="IBAN:", font=font_size).grid(row=8, column=0, **padding)
            entry_iban = tk.Entry(popup, width=entry_width, font=font_size)
            entry_iban.grid(row=8, column=1, **padding)
            entry_iban.insert(0, invoice["iban"])

            tk.Label(popup, text="Swift:", font=font_size).grid(row=9, column=0, **padding)
            entry_swift = tk.Entry(popup, width=entry_width, font=font_size)
            entry_swift.grid(row=9, column=1, **padding)
            entry_swift.insert(0, invoice["swift"])

            tk.Label(popup, text="Agente:", font=font_size).grid(row=10, column=0, **padding)
            entry_agent_name = tk.Entry(popup, width=entry_width, font=font_size)
            entry_agent_name.grid(row=10, column=1, **padding)
            entry_agent_name.insert(0, invoice["agent_name"])

            tk.Label(popup, text="Condizioni di Pagamento:", font=font_size).grid(row=11, column=0, **padding)
            entry_payment_condition = tk.Entry(popup, width=entry_width, font=font_size)
            entry_payment_condition.grid(row=11, column=1, **padding)
            entry_payment_condition.insert(0, invoice["payment_condition"])

            tk.Label(popup, text="Mittente:", font=font_size).grid(row=12, column=0, **padding)
            entry_sender_name = tk.Entry(popup, width=entry_width, font=font_size)
            entry_sender_name.grid(row=12, column=1, **padding)
            entry_sender_name.insert(0, invoice["sender_name"])

            tk.Label(popup, text="Giorno di chiusura:", font=font_size).grid(row=13, column=0, **padding)
            entry_sender_giorno = tk.Entry(popup, width=entry_width, font=font_size)
            entry_sender_giorno.grid(row=13, column=1, **padding)
            entry_sender_giorno.insert(0, invoice["sender_giorno"])

            tk.Label(popup, text="Servizio di Corriere:", font=font_size).grid(row=14, column=0, **padding)
            entry_courier_service = tk.Entry(popup, width=entry_width, font=font_size)
            entry_courier_service.grid(row=14, column=1, **padding)
            entry_courier_service.insert(0, invoice["courier_service"])

            tk.Label(popup, text="Destinatario:", font=font_size).grid(row=15, column=0, **padding)
            entry_recipient = tk.Entry(popup, width=entry_width, font=font_size)
            entry_recipient.grid(row=15, column=1, **padding)
            entry_recipient.insert(0, invoice["recipient"])

            tk.Label(popup, text="Quantità Totale:", font=font_size).grid(row=16, column=0, **padding)
            entry_total_quantity = tk.Entry(popup, width=entry_width, font=font_size)
            entry_total_quantity.grid(row=16, column=1, **padding)
            entry_total_quantity.insert(0, invoice["total_quantity"])

            tk.Label(popup, text="IVA:", font=font_size).grid(row=17, column=0, **padding)
            entry_iva_amount = tk.Entry(popup, width=entry_width, font=font_size)
            entry_iva_amount.grid(row=17, column=1, **padding)
            entry_iva_amount.insert(0, invoice["iva_amount"])

            tk.Label(popup, text="Importo Totale:", font=font_size).grid(row=18, column=0, **padding)
            entry_total_amount = tk.Entry(popup, width=entry_width, font=font_size)
            entry_total_amount.grid(row=18, column=1, **padding)
            entry_total_amount.insert(0, invoice["total_amount"])

            tk.Label(popup, text="Totale Fattura EURO:", font=font_size).grid(row=19, column=0, **padding)
            entry_total_invoice_euro = tk.Entry(popup, width=entry_width, font=font_size)
            entry_total_invoice_euro.grid(row=19, column=1, **padding)
            entry_total_invoice_euro.insert(0, invoice["total_invoice_euro"])

            # Add save changes button at the bottom
            tk.Button(popup, text="Salva modifiche e conferma", command=save_changes, font=font_size).grid(row=22, columnspan=2,
                                                                                           pady=30)

        edit_and_confirm()
        # Continue to generate the PDF after the popup is closed
        pdf = PDF()
        pdf.add_page()
        pdf.invoice_header(invoice)
        pdf.customer_details(invoice)
        pdf.customer_details_1(invoice)
        pdf.bank_details(invoice)
        pdf.agent_details(invoice)
        pdf.product_table(products)
        pdf.total_details(invoice)
        pdf.thank_you_details(invoice)

    parent_frame.grid_columnconfigure(0, weight=1)
    parent_frame.grid_columnconfigure(1, weight=1)

    button = ctk.CTkButton(parent_frame, text="Stampa Fattura", command=print_invoice)
    button.pack(side="left", padx=10, pady=20)

def handle_action_1(action, current_agente,current_id,current_cliente_connesso,current_quantita, tree):
    if action == "modify":
        details = ask_details_1(current_cliente_connesso,f"Inserisci il nuovo id, cliente e quantità per '{current_cliente_connesso}':", current_id,current_cliente_connesso,current_quantita)
        if details and details['quantity'] is not None:
            #update_record_andria(product, details['quantity'], details['cartons'])
            show_agente_clienti_conn(current_agente, tree, tree.table_frame)
            messagebox.showinfo("Aggiunta Cliente Connesso", f"Hai aggiunto {details['quantity']} prodotti di '{current_cliente_connesso}'.")
    elif action == "delete":
        confirm = messagebox.askokcancel("Conferma Eliminazione", f"Sei sicuro di voler eliminare '{current_cliente_connesso}'?")
        if confirm:
            #delete_record_andria(product)
            show_agente_info(current_agente, tree, tree.table_frame)
            messagebox.showinfo("Eliminazione Cliente Connesso", f"Il cliente connesso '{current_cliente_connesso}' è stato eliminato.")

def ask_details_1(current_agente, prompt,current_id,current_cliente_connesso,current_quantita):
    dialog = tk.Toplevel()
    dialog.title("Scegli le nuove info del cliente connesso")
    dialog.grab_set()
    dialog.transient()

    label = tk.Label(dialog, text=prompt, font=("Arial", 14))
    label.pack(pady=10)

    # Inizializza i campi con i valori correnti
    id_var = tk.IntVar(value=current_id)
    cliente_connesso_var = tk.IntVar(value=current_cliente_connesso)
    quantity_var = tk.IntVar(value=current_quantita)

    id_label = tk.Label(dialog, text=f"Id di '{current_cliente_connesso}':", font=("Arial", 14))
    id_label.pack(pady=5)
    id_entry = tk.Entry(dialog, textvariable=id_var, font=("Arial", 14))
    id_entry.pack(pady=5)

    cliente_conn_label = tk.Label(dialog, text=f"Descrizione di '{current_cliente_connesso}':", font=("Arial", 14))
    cliente_conn_label.pack(pady=5)
    cliente_conn_entry = tk.Entry(dialog, textvariable=cliente_connesso_var, font=("Arial", 14))
    cliente_conn_entry.pack(pady=5)

    quantity_label = tk.Label(dialog, text=f"Quantità di '{current_quantita}':",  font= ("Arial", 14))
    quantity_label.pack(pady=5)
    quantity_entry = tk.Entry(dialog, textvariable=quantity_var, font=("Arial", 14))
    quantity_entry.pack(pady=5)

    def on_confirm():
        dialog.destroy()
        dialog.details = {'id': id_var.get(), 'cliente connesso': cliente_connesso_var.get(), 'quantity': quantity_var.get()}

    confirm_button = tk.Button(dialog, text="Conferma",  font=("Arial", 12) , command=on_confirm)
    confirm_button.pack(pady=10)

    center_window(dialog,600,400)
    dialog.wait_window()
    return getattr(dialog, 'details', None)




def show_dashboard4(parent_frame):
    global combobox, agenti, tree
    for widget in parent_frame.winfo_children():
        widget.destroy()

    parent_frame_color = parent_frame.cget("fg_color")

    # Crea un frame per contenere sia il titolo che la legenda
    top_frame = ctk.CTkFrame(parent_frame, corner_radius=5, fg_color=parent_frame_color)
    top_frame.pack(fill="x", pady=0)  # Si espande orizzontalmente

    # Usa il layout grid nel top_frame
    top_frame.columnconfigure(0, weight=0)  # La colonna 0 non si espande
    top_frame.columnconfigure(1, weight=1)  # La colonna 1 si espande

    instruction_label = ctk.CTkLabel(top_frame, text="Scegliere Agente:", font=('Arial', 14))
    instruction_label.grid(row=0, column=1, padx=10)

    #cambiare funzione per agenti
    agenti = get_all_clienti_names()
    selected_agente = tk.StringVar()

    # Frame per allineare il menù a tendina e i pulsanti "Esporta in Excel"
    search_frame = ctk.CTkFrame(parent_frame,corner_radius=5)
    search_frame.pack(pady=10)

    # Style for larger font in Combobox
    style = ttk.Style()
    style.configure("TCombobox", font=('Arial', 16))  # Adjust font size as needed
    style.configure("TCombobox*Listbox*Font", font=('Arial', 16))  # Ensure larger font size for dropdown menu items

    # Creazione del menù a tendina con ricerca incrementale
    combobox = ttk.Combobox(search_frame, textvariable=selected_agente, values=agenti, font=('Arial', 16))
    combobox.configure(width=20)
    combobox.pack(side="left", padx=10)

    # Abilita la ricerca incrementale nel menù a tendina
    combobox.bind('<KeyRelease>', on_keyrelease)
    combobox.pack(pady=10)

    def on_agente_selected(event):
        agente = selected_agente.get()
        show_agente_info(agente, tree, table_frame)
        show_agente_clienti_conn(agente, tree_1, table_frame_1)
        mostra_frame()
        
    combobox.bind("<<ComboboxSelected>>", on_agente_selected)

    # Create a style for the Listbox within the Combobox dropdown
    combobox.option_add('*TCombobox*Listbox*Font', ('Arial', 16))
    '''
    def on_agente_selected(event):
        if listbox.curselection():  # Controlla se c'è una selezione
            agente = listbox.get(listbox.curselection())
            selected_agente.set(agente)  # Memorizza il cliente selezionato nell'Entry
            # Nasconde la Listbox dopo la selezione
            listbox_window.withdraw()
            show_agente_info(agente, tree, table_frame)
            show_agente_clienti_conn(agente, tree_1, table_frame_1)
            mostra_frame()

    # Funzione per aggiornare la Listbox in base alla ricerca
    def update_listbox(*args):
        search_term = entry.get().lower()  # Prende il testo inserito nella barra di ricerca
        listbox.delete(0, tk.END)  # Svuota la Listbox

        if search_term == "":  # Se la barra di ricerca è vuota, nascondi la Listbox
            listbox_window.withdraw()
            return

        # Filtra i clienti e aggiorna la Listbox
        filtered_agenti = [agente for agente in agenti if search_term in agente.lower()]

        if filtered_agenti:
            # Mostra la Listbox sotto la Entry solo se ci sono risultati
            listbox_window.geometry(f"300x200+{parent_frame.winfo_x() + search_frame.winfo_x() + entry.winfo_x() + 20}+{parent_frame.winfo_y() + search_frame.winfo_y() +entry.winfo_y() + search_frame.winfo_height() + entry.winfo_height() + 40}")
            listbox_window.deiconify()

            for agente in filtered_agenti:
                listbox.insert(tk.END, agente)
        else:
            listbox_window.withdraw()
            #listbox.place_forget()
            # Nasconde la Listbox se non ci sono risultati

    # Creazione della barra di ricerca (Entry)
    entry = tk.Entry(search_frame, textvariable=selected_agente, font=('Arial', 16))
    entry.pack(side="left", padx=10,pady=10)
    entry.config(width=25)
    entry.bind('<KeyRelease>', update_listbox)  # Ogni volta che si digita, aggiorna la Listbox

    listbox_window = tk.Toplevel(search_frame)
    listbox_window.wm_overrideredirect(True)  # Rimuove bordi e titoli della finestra
    listbox_window.geometry("300x200")
    listbox_window.withdraw()

    # Creazione della Listbox per visualizzare i risultati (inizialmente nascosta)
    listbox = tk.Listbox(listbox_window, font=('Arial', 16), height=8, selectmode=tk.SINGLE)
    listbox.pack(side="left", fill="both", expand=True)


    # Quando si seleziona un cliente dalla Listbox
    listbox.bind('<<ListboxSelect>>', on_agente_selected)
    '''

    def mostra_frame():
        table_frame.pack(side="left", padx=5, pady=10, fill="both", expand=True)  # Mostra il primo frame
        table_frame_1.pack(side="left", pady=10, fill="both", expand=True)  # Mostra il secondo frame affiancato

    # Pulsante "Esporta Excel Fornitori" accanto al pulsante prodotti
    export_fornitori_button = ctk.CTkButton(search_frame, text="Esporta Excel Agenti", command=open_agenti_selection_popup, font=('Arial', 14))
    export_fornitori_button.pack(side="left", padx=10)

    export_lista_button = ctk.CTkButton(search_frame, text="Esporta Excel Clienti Connessi", command=open_clienti_conn_popup, font=('Arial', 14))
    export_lista_button.pack(side="left", padx=10)

    export_lista_button = ctk.CTkButton(search_frame, text="Aggiungi Agente",command = lambda:open_add_popup("Agente"), font=('Arial', 14))
    export_lista_button.pack(side="left", padx=10)

    provvigioni_button = ctk.CTkButton(search_frame, text="Calcola Provvigioni", command = lambda:open_add_popup("Provvigioni"), font=('Arial', 14))
    provvigioni_button.pack(side="left", padx=10)

    global table_frame, table_frame_1
    table_frame = ctk.CTkFrame(parent_frame, corner_radius=5)
    table_frame.pack(side="left", padx=5, pady=10, fill="both", expand=True)
    table_frame.pack_forget()

    table_frame_1 = ctk.CTkFrame(parent_frame, corner_radius=5)
    table_frame_1.pack(side="left", pady=10, fill="both", expand=True)
    table_frame_1.pack_forget()

    table_title = ctk.CTkLabel(table_frame, text="Dati Agente:", font=('Arial', 14, 'bold'))
    table_title.pack(pady=10)

    table_title_1 = ctk.CTkLabel(table_frame_1, text="Clienti connessi:", font=('Arial', 14, 'bold'))
    table_title_1.pack(pady=10)

    # Create a frame to hold the table and scrollbars
    table_container = tk.Frame(table_frame, width=500, height=400)
    table_container.pack_propagate(False)  # Prevent the frame from resizing to fit its children
    table_container.pack(pady=10, fill="both", expand=True)

    table_container_1 = tk.Frame(table_frame_1, width=800, height=400)
    table_container_1.pack_propagate(False)  # Prevent the frame from resizing to fit its children
    table_container_1.pack(pady=10, fill="both", expand=True)

    # Use only one column for the Treeview to display data vertically
    global columns
    columns = ("",)

    global columns_1
    columns_1 = ("Cliente Connesso",)

    global tree
    tree = ttk.Treeview(table_container, columns=columns, show="headings")
    tree.table_container = table_container  # Save the frame in the treeview for reference
    setup_context_menu(tree)
    tree.bind("<B1-Motion>", lambda event: on_mouse_drag(event, tree))


    global tree_1
    tree_1 = ttk.Treeview(table_container_1, columns=columns_1, show="headings")
    tree_1.table_container = table_container_1  # Save the frame in the treeview for reference
    setup_context_menu(tree_1)
    tree_1.bind("<B1-Motion>", lambda event: on_mouse_drag(event, tree_1))

    style = ttk.Style()
    style.configure("Treeview",
                    rowheight=30,
                    font=('Arial', 14),
                    background="#f1f8e9",
                    foreground="#004d40",
                    fieldbackground="#f1f8e9",
                    bordercolor="#000000",
                    relief="solid",
                    borderwidth=1)
    style.configure("Treeview.Heading",
                    font=('Arial', 16, 'bold'),
                    background="#a5d6a7",
                    foreground="#004d40")
    style.map("Treeview",
              background=[('selected', '#c8e6c9')],
              foreground=[('selected', '#004d40')])

    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=1500, anchor="w")

    for col_1 in columns_1:
        tree_1.heading(col_1, text=col_1)
        tree_1.column(col_1, width=100, anchor="center")
    # Set a larger width to accommodate longer text

    yscrollbar = ttk.Scrollbar(table_container, orient=tk.VERTICAL, command=tree.yview)
    yscrollbar.pack(side="right", fill="y")
    tree.configure(yscroll=yscrollbar.set)

    yscrollbar_1 = ttk.Scrollbar(table_container_1, orient=tk.VERTICAL, command=tree.yview)
    yscrollbar_1.pack(side="right", fill="y")
    tree_1.configure(yscroll=yscrollbar_1.set)

    tree.pack(side="left", fill="both", expand=True)
    tree.bind("<Double-1>", lambda event: on_double_click(event, tree))

    tree_1.pack(side="left", fill="both", expand=True)
    tree_1.bind("<Double-1>", lambda event: on_double_click_1(event, tree_1))

    def fetch_orders():
        conn = _connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM orders")
        orders = cursor.fetchall()
        conn.close()
        return orders

    orders = fetch_orders()

    def update_treeview(order_list):
        for row in tree.get_children():
            tree.delete(row)
        for order in order_list:
            tree.insert("", tk.END, values=order)

    update_treeview(orders)

    def print_invoice():
        selected_item = tree.selection()
        if selected_item:
            order = tree.item(selected_item, "values")
            generate_invoice_pdf(order)
        else:
            messagebox.showwarning("Seleziona Ordine", "Per favore, seleziona un ordine per stampare la fattura.")

    class PDF(FPDF):
        def header(self):

            # Aggiungi il logo in alto a sinistra
            self.image('resources/LogoCINCOTTI.jpg', 8, 2, 77, 44, 'JPG')  # Modifica il percorso e le dimensioni

            # Aggiungi l'immagine in alto a destra
            self.image('resources/Logojpeg.jpg', 120, 8, 77, 33, 'JPG')  # Modifica il percorso e le dimensioni

            self.ln(35)

        def footer(self):
            # Aggiungi l'immagine in basso al centro
            self.set_y(-40)  # Posiziona a 30 mm dal fondo della pagina
            self.image('resources/LOGOMOZZABELLAECINCOTTI.jpg', 75, self.get_y(), 70, 40,
                       'JPG')  # Modifica il percorso e le dimensioni

        def draw_table(self, data, col_widths, headers=None, bold_headers=False):
            if headers:
                if bold_headers:
                    self.set_font('Arial', 'B', 10)
                else:
                    self.set_font('Arial', '', 10)

                for header, width in zip(headers, col_widths):
                    self.cell(width, 8, header, 1, 0, 'C')
                self.ln()

            if bold_headers:
                self.set_font('Arial', '', 10)

            for row in data:
                # Prima passata per determinare l'altezza massima della riga
                row_heights = []
                for datum, width in zip(row, col_widths):
                    line_height = 6  # Altezza per linea, può essere aggiustata
                    num_lines = self.get_string_width(datum) / width  # Calcola il numero di righe necessarie
                    num_lines = datum.count('\n') + 1  # Contiamo le nuove linee
                    height = line_height * num_lines
                    row_heights.append(height)
                max_row_height = max(row_heights)  # Altezza massima tra tutte le celle della riga

                # Seconda passata per disegnare le celle con l'altezza della riga determinata
                x_start = self.get_x()  # Posizione iniziale X per la riga
                y_start = self.get_y()  # Posizione iniziale Y per la riga
                for datum, width in zip(row, col_widths):
                    # Disegna la cella utilizzando multi_cell per il testo con ritorni a capo
                    self.multi_cell(width, line_height, datum, 1, 'L', False)
                    # Posiziona il cursore all'inizio della prossima cella nella riga
                    self.set_xy(x_start + width, y_start)
                    x_start += width  # Avanza per la larghezza della cella
                self.ln(max_row_height)  # Salta alla prossima riga usando l'altezza massima

        def invoice_header(self, invoice):
            self.set_font('Arial', '', 10)
            data = [
                f"Document: INVOICE", f"n. {invoice['number']}", f"Data: {invoice['Date_1']}"
            ]
            col_widths = [70, 40, 70]  # Regola le larghezze delle colonne come necessario
            self.draw_table([data], col_widths)

        def customer_details(self, invoice):
            data = [
                ["Goods destination:\n" + invoice["customer_phone"], "Customer:\n" + invoice["customer_address"]],
            ]
            col_widths = [70, 110]  # Regola le larghezze delle colonne come necessario
            self.draw_table(data, col_widths, bold_headers=True)

        def customer_details_1(self, invoice):
            data = [
                ["ID:\n" + invoice["id"], "Email:\n" + invoice["customer_email"],"VAT:\n" + invoice["vat_number"],"Nation:\n" + invoice["nation"]],
            ]
            col_widths = [25,65,45,45]  # Regola le larghezze delle colonne come necessario
            self.draw_table(data, col_widths, bold_headers=True)

        def bank_details(self, invoice):
            data = [
                ["IBAN:\n"+invoice["iban"], "Swift:\n"+invoice["swift"]]
            ]
            col_widths = [90,90]
            self.set_font('Arial', '', 8)
            self.draw_table(data, col_widths, bold_headers=True)

        def agent_details(self, invoice):
            data = [
                ["Agent:\n"+ invoice["agent_name"],"Payment Condition:\n"+ invoice["payment_condition"],"Sender:\n"+ invoice["sender_name"], "Giorno di chiusura:\n"+ invoice["sender_giorno"]]
            ]
            col_widths = [55,35,45,45]
            self.set_font('Arial', '', 8)
            self.draw_table(data, col_widths, bold_headers=True)

        def product_table(self, products):
            headers = ["Article", "Description", "Quantity", "Price", "Discount", "Amount"]
            data = [
                [product['article'], product['product_description'], str(product['product_quantity']),
                 str(product['product_price']), str(product['product_discount']), str(product['product_amount'])]
                for product in products
            ]
            col_widths = [20, 70, 20, 20, 20, 30]
            self.set_font('Arial', 'B', 10)
            self.cell(0, 8, 'Product Details', 0, 1, 'L')
            self.draw_table(data, col_widths, headers, bold_headers=True)

        def total_details(self, invoice):
            data = [
                ["Total Quantity:\n" + str(invoice["total_quantity"]),"IVA:\n" + str(invoice["iva_amount"]), "Sender:\n" + invoice["sender_name"]],
                ["Reason for export:\n" + "Vendita","Courier Service:\n" + invoice["courier_service"],"Recipient:\n" + invoice["recipient"]],
                ["Total Amount:\n" + str(invoice["total_amount"]),"IVA Amount:\n" + str(invoice["iva_amount"]),"Total Invoice EURO:\n" + str(invoice["total_invoice_euro"])],
            ]
            col_widths = [60,60,60]
            self.set_font('Arial', 'B', 10)
            self.cell(0, 8, 'Total Details', 0, 1, 'L')
            self.draw_table(data, col_widths, bold_headers=True)

        def thank_you_details(self, invoice):
            self.set_font('Arial', 'B', 10)
            self.cell(0, 8, 'Thank you for choosing our products!', 0, 1, 'R')

    def fetch_invoice_data(invoice_number):
        conn = _connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM `orders_fattura` WHERE `id` = %s", (invoice_number,))
        column_names = [description[0] for description in cursor.description]
        invoice_tuple = cursor.fetchone()
        conn.close()
        if invoice_tuple:
            invoice = dict(zip(column_names, invoice_tuple))
            return invoice
        return None

    def fetch_invoice_products(invoice_number):
        conn = _connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM `orders_fattura` WHERE `id` = %s", (invoice_number,))
        column_names = [description[0] for description in cursor.description]
        product_rows = cursor.fetchall()
        conn.close()
        products = [dict(zip(column_names, product_row)) for product_row in product_rows]
        return products

    def generate_invoice_pdf(order):
        invoice_number = order[0]
        invoice = fetch_invoice_data(invoice_number)
        products = fetch_invoice_products(invoice_number)
        if not invoice:
            messagebox.showerror("Errore", "Dati della fattura non trovati.")
            return

        # Popup window to edit and confirm invoice data
        def edit_and_confirm():
            def save_changes():
                # Update invoice data from the entry fields
                invoice["number"] = entry_number.get()
                invoice["Date_1"] = entry_date.get()
                invoice["customer_phone"] = entry_customer_phone.get()
                invoice["customer_address"] = entry_customer_address.get()
                invoice["id"] = entry_id.get()
                invoice["customer_email"] = entry_customer_email.get()
                invoice["vat_number"] = entry_vat_number.get()
                invoice["nation"] = entry_nation.get()
                invoice["iban"] = entry_iban.get()
                invoice["swift"] = entry_swift.get()
                invoice["agent_name"] = entry_agent_name.get()
                invoice["payment_condition"] = entry_payment_condition.get()
                invoice["sender_name"] = entry_sender_name.get()
                invoice["sender_giorno"] = entry_sender_giorno.get()
                invoice["courier_service"] = entry_courier_service.get()
                invoice["recipient"] = entry_recipient.get()
                invoice["total_quantity"] = entry_total_quantity.get()
                invoice["iva_amount"] = entry_iva_amount.get()
                invoice["total_amount"] = entry_total_amount.get()
                invoice["total_invoice_euro"] = entry_total_invoice_euro.get()

                # Close the popup window after saving changes
                popup.destroy()

                # Open a separate window to choose save path
                choose_save_path()

            # Ottiene la data corrente e la formatta come YYYYMMDD
            current_date = datetime.now().strftime("%Y%m%d")
            filename = f"Fattura_{current_date}"

            def choose_save_path():
                save_path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")], initialfile=filename)
                if save_path:
                    pdf = PDF()
                    pdf.add_page()
                    pdf.invoice_header(invoice)
                    pdf.customer_details(invoice)
                    pdf.customer_details_1(invoice)
                    pdf.bank_details(invoice)
                    pdf.agent_details(invoice)
                    pdf.product_table(products)
                    pdf.total_details(invoice)
                    pdf.thank_you_details(invoice)
                    pdf.output(save_path)
                    messagebox.showinfo("Successo", f"Fattura salvata come {os.path.basename(save_path)}")
                else:
                    messagebox.showwarning("Salvataggio Annullato", "Il salvataggio della fattura è stato annullato.")

            # Create the popup window
            popup = tk.Toplevel()
            popup.title("Modifica i dettagli della fattura")
            popup.geometry("800x900")  # Set the size of the popup window
            popup.resizable(True, True)  # Prevent resizing for consistent layout

            # Add padding and spacing between elements
            padding = {'padx': 10, 'pady': 5}
            # Set the width of entry fields and the font size
            entry_width = 40
            font_size = ("Arial", 14)  # Font family Arial, size 14

            # Create and grid labels and entries for each invoice field with larger font
            tk.Label(popup, text="Numero:", font=font_size).grid(row=0, column=0, **padding)
            entry_number = tk.Entry(popup, width=entry_width, font=font_size)
            entry_number.grid(row=0, column=1, **padding)
            entry_number.insert(0, invoice["number"])

            tk.Label(popup, text="Data:", font=font_size).grid(row=1, column=0, **padding)
            entry_date = tk.Entry(popup, width=entry_width, font=font_size)
            entry_date.grid(row=1, column=1, **padding)
            entry_date.insert(0, invoice["Date_1"])

            tk.Label(popup, text="Telefono Cliente:", font=font_size).grid(row=2, column=0, **padding)
            entry_customer_phone = tk.Entry(popup, width=entry_width, font=font_size)
            entry_customer_phone.grid(row=2, column=1, **padding)
            entry_customer_phone.insert(0, invoice["customer_phone"])

            tk.Label(popup, text="Indirizzo Cliente:", font=font_size).grid(row=3, column=0, **padding)
            entry_customer_address = tk.Entry(popup, width=entry_width, font=font_size)
            entry_customer_address.grid(row=3, column=1, **padding)
            entry_customer_address.insert(0, invoice["customer_address"])

            tk.Label(popup, text="ID:", font=font_size).grid(row=4, column=0, **padding)
            entry_id = tk.Entry(popup, width=entry_width, font=font_size)
            entry_id.grid(row=4, column=1, **padding)
            entry_id.insert(0, invoice["id"])

            tk.Label(popup, text="Email Cliente:", font=font_size).grid(row=5, column=0, **padding)
            entry_customer_email = tk.Entry(popup, width=entry_width, font=font_size)
            entry_customer_email.grid(row=5, column=1, **padding)
            entry_customer_email.insert(0, invoice["customer_email"])

            tk.Label(popup, text="Partita IVA:", font=font_size).grid(row=6, column=0, **padding)
            entry_vat_number = tk.Entry(popup, width=entry_width, font=font_size)
            entry_vat_number.grid(row=6, column=1, **padding)
            entry_vat_number.insert(0, invoice["vat_number"])

            tk.Label(popup, text="Nazione:", font=font_size).grid(row=7, column=0, **padding)
            entry_nation = tk.Entry(popup, width=entry_width, font=font_size)
            entry_nation.grid(row=7, column=1, **padding)
            entry_nation.insert(0, invoice["nation"])

            tk.Label(popup, text="IBAN:", font=font_size).grid(row=8, column=0, **padding)
            entry_iban = tk.Entry(popup, width=entry_width, font=font_size)
            entry_iban.grid(row=8, column=1, **padding)
            entry_iban.insert(0, invoice["iban"])

            tk.Label(popup, text="Swift:", font=font_size).grid(row=9, column=0, **padding)
            entry_swift = tk.Entry(popup, width=entry_width, font=font_size)
            entry_swift.grid(row=9, column=1, **padding)
            entry_swift.insert(0, invoice["swift"])

            tk.Label(popup, text="Agente:", font=font_size).grid(row=10, column=0, **padding)
            entry_agent_name = tk.Entry(popup, width=entry_width, font=font_size)
            entry_agent_name.grid(row=10, column=1, **padding)
            entry_agent_name.insert(0, invoice["agent_name"])

            tk.Label(popup, text="Condizioni di Pagamento:", font=font_size).grid(row=11, column=0, **padding)
            entry_payment_condition = tk.Entry(popup, width=entry_width, font=font_size)
            entry_payment_condition.grid(row=11, column=1, **padding)
            entry_payment_condition.insert(0, invoice["payment_condition"])

            tk.Label(popup, text="Mittente:", font=font_size).grid(row=12, column=0, **padding)
            entry_sender_name = tk.Entry(popup, width=entry_width, font=font_size)
            entry_sender_name.grid(row=12, column=1, **padding)
            entry_sender_name.insert(0, invoice["sender_name"])

            tk.Label(popup, text="Giorno di chiusura:", font=font_size).grid(row=13, column=0, **padding)
            entry_sender_giorno = tk.Entry(popup, width=entry_width, font=font_size)
            entry_sender_giorno.grid(row=13, column=1, **padding)
            entry_sender_giorno.insert(0, invoice["sender_giorno"])

            tk.Label(popup, text="Servizio di Corriere:", font=font_size).grid(row=14, column=0, **padding)
            entry_courier_service = tk.Entry(popup, width=entry_width, font=font_size)
            entry_courier_service.grid(row=14, column=1, **padding)
            entry_courier_service.insert(0, invoice["courier_service"])

            tk.Label(popup, text="Destinatario:", font=font_size).grid(row=15, column=0, **padding)
            entry_recipient = tk.Entry(popup, width=entry_width, font=font_size)
            entry_recipient.grid(row=15, column=1, **padding)
            entry_recipient.insert(0, invoice["recipient"])

            tk.Label(popup, text="Quantità Totale:", font=font_size).grid(row=16, column=0, **padding)
            entry_total_quantity = tk.Entry(popup, width=entry_width, font=font_size)
            entry_total_quantity.grid(row=16, column=1, **padding)
            entry_total_quantity.insert(0, invoice["total_quantity"])

            tk.Label(popup, text="IVA:", font=font_size).grid(row=17, column=0, **padding)
            entry_iva_amount = tk.Entry(popup, width=entry_width, font=font_size)
            entry_iva_amount.grid(row=17, column=1, **padding)
            entry_iva_amount.insert(0, invoice["iva_amount"])

            tk.Label(popup, text="Importo Totale:", font=font_size).grid(row=18, column=0, **padding)
            entry_total_amount = tk.Entry(popup, width=entry_width, font=font_size)
            entry_total_amount.grid(row=18, column=1, **padding)
            entry_total_amount.insert(0, invoice["total_amount"])

            tk.Label(popup, text="Totale Fattura EURO:", font=font_size).grid(row=19, column=0, **padding)
            entry_total_invoice_euro = tk.Entry(popup, width=entry_width, font=font_size)
            entry_total_invoice_euro.grid(row=19, column=1, **padding)
            entry_total_invoice_euro.insert(0, invoice["total_invoice_euro"])

            # Add save changes button at the bottom
            tk.Button(popup, text="Salva modifiche e conferma", command=save_changes, font=font_size).grid(row=22, columnspan=2,
                                                                                           pady=30)

        edit_and_confirm()
        # Continue to generate the PDF after the popup is closed
        pdf = PDF()
        pdf.add_page()
        pdf.invoice_header(invoice)
        pdf.customer_details(invoice)
        pdf.customer_details_1(invoice)
        pdf.bank_details(invoice)
        pdf.agent_details(invoice)
        pdf.product_table(products)
        pdf.total_details(invoice)
        pdf.thank_you_details(invoice)

    #button = ctk.CTkButton(search_frame, text="Stampa Fattura", command=print_invoice, font=(('Arial', 14)))
    #button.pack(side="left", padx=10)




# Other components of your application remain unchanged